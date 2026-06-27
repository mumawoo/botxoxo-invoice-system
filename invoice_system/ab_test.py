from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

from .config import Settings
from .excel_store import InvoiceWorkbook
from .image_splitter import OpenCVInvoiceSplitter, iter_images
from .models import InvoiceRecord, OCRAuditRow, OCRResult
from .paddle_vl_scan import PaddleOCRVLRecognizer
from .qwen_scan import QwenScanRecognizer
from .recognizers import Recognizer


@dataclass(frozen=True)
class ABTestSummary:
    source_images: int
    crops: int
    paddle_vl_rows: int
    qwen_rows: int
    output_dir: Path
    comparison_path: Path


def run_ab_test(
    settings: Settings,
    input_path: Path,
    output_dir: Path,
    paddle_vl_recognizer: Recognizer | None = None,
    qwen_recognizer: Recognizer | None = None,
) -> ABTestSummary:
    source_images = iter_images(input_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    crops_dir = output_dir / "crops"
    splitter = OpenCVInvoiceSplitter(crops_dir)
    crops = []
    for image in source_images:
        crops.extend(splitter.split(image))

    paddle_vl = paddle_vl_recognizer or PaddleOCRVLRecognizer()
    qwen = qwen_recognizer or QwenScanRecognizer(settings)
    paddle_vl_records, paddle_vl_audits, paddle_vl_results = _scan_crops(crops, paddle_vl)
    qwen_records, qwen_audits, qwen_results = _scan_crops(crops, qwen)

    paddle_vl_dir = output_dir / "paddleocr_vl"
    qwen_dir = output_dir / "qwen"
    paddle_vl_rows = InvoiceWorkbook(paddle_vl_dir / "Invoice_Output.xlsx").write_records(paddle_vl_records, paddle_vl_audits)
    qwen_rows = InvoiceWorkbook(qwen_dir / "Invoice_Output.xlsx").write_records(qwen_records, qwen_audits)
    comparison_path = output_dir / "AB_Comparison.xlsx"
    _write_comparison(comparison_path, paddle_vl_results, qwen_results)
    return ABTestSummary(len(source_images), len(crops), paddle_vl_rows, qwen_rows, output_dir, comparison_path)


def telegram_ab_input(settings: Settings, user_id: int, day: str | None = None) -> Path:
    user_root = settings.inbound_dir / "telegram" / str(user_id)
    if day:
        return user_root / day
    days = sorted(path for path in user_root.iterdir() if path.is_dir()) if user_root.exists() else []
    return days[-1] if days else user_root


def _scan_crops(crops, recognizer: Recognizer) -> tuple[list[InvoiceRecord], list[OCRAuditRow], list[OCRResult]]:
    records: list[InvoiceRecord] = []
    audits: list[OCRAuditRow] = []
    results: list[OCRResult] = []
    for crop in crops:
        result = recognizer.recognize(crop.crop_path)
        results.append(result)
        record = result.parsed_invoice or InvoiceRecord(remarks=f"{recognizer.engine} produced no valid invoice")
        record.source_image = str(crop.source_path)
        record.crop_image = str(crop.crop_path)
        if result.parsed_invoice is not None:
            records.append(record)
        audits.append(
            OCRAuditRow(
                source_image=str(crop.source_path),
                crop_image=str(crop.crop_path),
                decision=f"{recognizer.engine} direct A/B scan",
                used_codex=recognizer.engine == "qwen_scan",
                paddle_confidence=result.confidence if recognizer.engine == "paddleocr_vl" else 0.0,
                codex_confidence=result.confidence if recognizer.engine == "qwen_scan" else 0.0,
                paddle_error=result.error if recognizer.engine == "paddleocr_vl" else "",
                codex_error=result.error if recognizer.engine == "qwen_scan" else "",
                paddle_text=result.text if recognizer.engine == "paddleocr_vl" else "",
                codex_text=result.text if recognizer.engine == "qwen_scan" else "",
            )
        )
    return records, audits, results


def _write_comparison(path: Path, paddle_vl_results: list[OCRResult], qwen_results: list[OCRResult]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AB_Comparison"
    ws.append(
        [
            "Crop No.",
            "Field",
            "PaddleOCR-VL",
            "Qwen",
            "Match",
            "PaddleOCR-VL Confidence",
            "Qwen Confidence",
            "PaddleOCR-VL Error",
            "Qwen Error",
        ]
    )
    fields = [
        ("invoice_date", "Invoice Date"),
        ("seller", "Seller"),
        ("currency", "Currency"),
        ("total_amount", "Total Amount"),
        ("vat_amount", "VAT Amount"),
        ("sales_tax", "Sales Tax"),
        ("tips", "Tips"),
        ("expense_category", "Expense Category"),
    ]
    for index, (paddle_vl, qwen) in enumerate(zip(paddle_vl_results, qwen_results), start=1):
        left = paddle_vl.parsed_invoice
        right = qwen.parsed_invoice
        for attr, label in fields:
            local_value = getattr(left, attr, "") if left else ""
            qwen_value = getattr(right, attr, "") if right else ""
            ws.append(
                [
                    index,
                    label,
                    local_value,
                    qwen_value,
                    "yes" if _same_value(local_value, qwen_value) else "no",
                    round(paddle_vl.confidence, 3),
                    round(qwen.confidence, 3),
                    paddle_vl.error,
                    qwen.error,
                ]
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _same_value(left: object, right: object) -> bool:
    if isinstance(left, (int, float)) or isinstance(right, (int, float)):
        try:
            return abs(float(left or 0) - float(right or 0)) <= 0.50
        except (TypeError, ValueError):
            return False
    return str(left or "").strip().casefold() == str(right or "").strip().casefold()
