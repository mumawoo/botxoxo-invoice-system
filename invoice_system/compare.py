from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .excel_store import load_invoice_records
from .image_splitter import iter_images
from .models import InvoiceRecord
from .parsing import fuzzy_match, normalize_date, normalize_text


def compare_outputs(baseline: Path, candidate: Path, output: Path) -> Path:
    baseline_workbook = _find_excel(baseline)
    candidate_workbook = _find_excel(candidate)
    baseline_records = load_invoice_records(baseline_workbook)
    candidate_records = load_invoice_records(candidate_workbook)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    categories: dict[str, int] = {}
    ws.append(_comparison_headers())

    used_candidate: set[int] = set()
    for base in baseline_records:
        match_index, candidate_record = _best_match(base, candidate_records, used_candidate)
        if candidate_record is None:
            category = "missing in Windows"
            categories[category] = categories.get(category, 0) + 1
            ws.append(_comparison_row(category, base, None, []))
            continue
        used_candidate.add(match_index)
        diffs = _diffs(base, candidate_record)
        category = _comparison_category(base, candidate_record, diffs)
        categories[category] = categories.get(category, 0) + 1
        ws.append(_comparison_row(category, base, candidate_record, diffs))

    for index, record in enumerate(candidate_records):
        if index not in used_candidate:
            category = "extra in Windows"
            categories[category] = categories.get(category, 0) + 1
            ws.append(_comparison_row(category, None, record, []))

    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Baseline root", str(baseline)])
    summary_ws.append(["Windows root", str(candidate)])
    summary_ws.append(["Baseline workbook", str(baseline_workbook)])
    summary_ws.append(["Windows workbook", str(candidate_workbook)])
    summary_ws.append(["Report output", str(output)])
    summary_ws.append(["Baseline rows", len(baseline_records)])
    summary_ws.append(["Windows rows", len(candidate_records)])
    for category in (
        "exact match",
        "near match",
        "OCR disagreement resolved by Codex Scan",
        "field-level mismatch",
        "missing in Windows",
        "extra in Windows",
    ):
        summary_ws.append([category, categories.get(category, 0)])

    crop_ws = wb.create_sheet("Crop Counts")
    crop_ws.append(["Side", "Source", "Crop Count", "Evidence"])
    for side, root, workbook_path in (("baseline", baseline, baseline_workbook), ("candidate", candidate, candidate_workbook)):
        for source, count, evidence in _crop_count_rows(root, workbook_path):
            crop_ws.append([side, source, count, evidence])
    wb.save(output)
    return output


def _find_excel(path: Path) -> Path:
    if path.is_file():
        return path
    candidates = sorted(p for p in path.rglob("*.xlsx") if _is_candidate_workbook(p))
    output_candidates = [p for p in candidates if p.name.startswith("Invoice_Output")]
    if output_candidates:
        return output_candidates[0]
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"No invoice workbook found under {path}")


def _is_candidate_workbook(path: Path) -> bool:
    name = path.name
    lowered = name.casefold()
    if name.startswith("~$"):
        return False
    if "comparison_report" in lowered:
        return False
    if "manually_checked" in lowered:
        return False
    return path.suffix.casefold() == ".xlsx"


def _comparison_headers() -> list[str]:
    return [
        "Category",
        "Baseline Date",
        "Windows Date",
        "Baseline Seller",
        "Windows Seller",
        "Baseline Currency",
        "Windows Currency",
        "Baseline Total",
        "Windows Total",
        "Baseline Expense",
        "Windows Expense",
        "Baseline VAT",
        "Windows VAT",
        "Baseline Sales Tax",
        "Windows Sales Tax",
        "Baseline Tips",
        "Windows Tips",
        "Field Differences",
    ]


def _comparison_row(
    category: str,
    base: InvoiceRecord | None,
    candidate: InvoiceRecord | None,
    diffs: list[str],
) -> list[object]:
    return [
        category,
        base.invoice_date if base else "",
        candidate.invoice_date if candidate else "",
        base.seller if base else "",
        candidate.seller if candidate else "",
        base.currency if base else "",
        candidate.currency if candidate else "",
        base.total_amount if base else "",
        candidate.total_amount if candidate else "",
        base.expense_amount if base else "",
        candidate.expense_amount if candidate else "",
        base.vat_amount if base else "",
        candidate.vat_amount if candidate else "",
        base.sales_tax if base else "",
        candidate.sales_tax if candidate else "",
        base.tips if base else "",
        candidate.tips if candidate else "",
        "; ".join(diffs),
    ]


def _comparison_category(base: InvoiceRecord, candidate: InvoiceRecord, diffs: list[str]) -> str:
    if _resolved_ocr_disagreement_with_codex(candidate):
        return "OCR disagreement resolved by Codex Scan"
    if not diffs:
        return "exact match"
    if _near(base, candidate):
        return "near match"
    return "field-level mismatch"


def _resolved_ocr_disagreement_with_codex(record: InvoiceRecord) -> bool:
    remarks = (record.remarks or "").casefold()
    return "codex scan used" in remarks and "ocr mismatch" in remarks


def _crop_count_rows(root: Path, workbook_path: Path) -> list[tuple[str, int, str]]:
    audit_counts = _audit_crop_counts(workbook_path)
    if audit_counts:
        return [(source, count, "OCR_Audit") for source, count in sorted(audit_counts.items())]
    image_counts = _image_crop_counts(root)
    if image_counts:
        return [(source, count, "image filename inference") for source, count in sorted(image_counts.items())]
    return [("all images", 0, "image files")]


def _image_crop_counts(root: Path) -> dict[str, int]:
    counts: dict[str, set[str]] = {}
    for image in iter_images(root):
        source = _source_from_crop_filename(image)
        counts.setdefault(source, set()).add(str(image))
    return {source: len(crops) for source, crops in counts.items()}


def _source_from_crop_filename(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"_d\d+$", "", stem, flags=re.IGNORECASE)
    match = re.match(r"^\d{4}_\d{4}-\d{2}-\d{2}_[A-Za-z]{3}_\d+(?:\.\d+)?_(.+)$", stem)
    if match:
        return match.group(1)
    return stem or path.name


def _audit_crop_counts(workbook_path: Path) -> dict[str, int]:
    if not workbook_path.exists():
        return {}
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if "OCR_Audit" not in wb.sheetnames:
            return {}
        ws = wb["OCR_Audit"]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows, [])]
        try:
            source_index = headers.index("Source Image")
            crop_index = headers.index("Crop Image")
        except ValueError:
            return {}

        counts: dict[str, set[str]] = {}
        for row in rows:
            source = str(row[source_index] or "").strip()
            crop = str(row[crop_index] or "").strip()
            if not source or not crop:
                continue
            source_label = Path(source).name
            counts.setdefault(source_label, set()).add(crop)
        return {source: len(crops) for source, crops in counts.items()}
    finally:
        wb.close()


def _best_match(records_base: InvoiceRecord, candidates: list[InvoiceRecord], used: set[int]) -> tuple[int, InvoiceRecord | None]:
    best_score = -1
    best_index = -1
    best_record = None
    for index, candidate in enumerate(candidates):
        if index in used:
            continue
        score, corroborated = _match_score(records_base, candidate)
        if not corroborated:
            continue
        if score > best_score:
            best_score = score
            best_index = index
            best_record = candidate
    return best_index, best_record if best_score >= 3 else None


def _match_score(base: InvoiceRecord, candidate: InvoiceRecord) -> tuple[int, bool]:
    date_match = _dates_match(base.invoice_date, candidate.invoice_date)
    seller_match = _sellers_match(base.seller, candidate.seller)
    amount_match = abs(base.total_amount - candidate.total_amount) <= 0.50
    score = (2 if date_match else 0) + (2 if seller_match else 0) + (3 if amount_match else 0)
    corroborated = (amount_match and (date_match or seller_match)) or (date_match and seller_match)
    return score, corroborated


def _diffs(base: InvoiceRecord, candidate: InvoiceRecord) -> list[str]:
    diffs: list[str] = []
    if not _dates_match(base.invoice_date, candidate.invoice_date):
        diffs.append("invoice_date")
    if not _sellers_match(base.seller, candidate.seller):
        diffs.append("seller")
    if _normalize_currency(base.currency) != _normalize_currency(candidate.currency):
        diffs.append("currency")
    for name in ("total_amount", "vat_amount", "sales_tax", "tips"):
        if abs(getattr(base, name) - getattr(candidate, name)) > 0.50:
            diffs.append(name)
    return diffs


def _near(base: InvoiceRecord, candidate: InvoiceRecord) -> bool:
    return _dates_match(base.invoice_date, candidate.invoice_date) and abs(base.total_amount - candidate.total_amount) <= 0.50


def _dates_match(left: str, right: str) -> bool:
    normalized_left = _normalize_compare_date(left)
    normalized_right = _normalize_compare_date(right)
    return bool(normalized_left) and normalized_left == normalized_right


def _normalize_compare_date(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return normalize_date(text) or text[:10]


def _sellers_match(left: str, right: str) -> bool:
    return fuzzy_match(left or "", right or "")


def _normalize_currency(value: str) -> str:
    normalized = normalize_text(value or "").casefold()
    if normalized in {"m.n.", "mn", "peso", "pesos"}:
        return "MXN"
    return normalized.upper()
