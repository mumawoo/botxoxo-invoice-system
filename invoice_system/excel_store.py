from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import EXCEL_HEADERS, OCR_AUDIT_HEADERS, SOURCE_QA_HEADERS, InvoiceRecord, OCRAuditRow, SourceQARecord
from .parsing import normalize_date, normalize_text

LOCKED_STATUSES = {"manually checked", "deleted"}
RUN_SUMMARY_HEADERS = ["Metric", "Value"]


class InvoiceWorkbook:
    def __init__(self, workbook_path: Path, manual_checked_path: Path | None = None) -> None:
        self.workbook_path = workbook_path
        self.manual_checked_path = manual_checked_path or workbook_path.with_name("Invoice_Manually_Checked.xlsx")
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)

    def write_records(
        self,
        records: list[InvoiceRecord],
        audits: list[OCRAuditRow] | None = None,
        source_qas: list[SourceQARecord] | None = None,
    ) -> int:
        locked = self._locked_records()
        unlocked = [record for record in records if _lock_key(record) not in locked]

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(EXCEL_HEADERS)

        line_no = 1
        for record in locked.values():
            record.line_no = line_no
            ws.append(record.to_excel_row())
            line_no += 1

        for record in unlocked:
            record.line_no = line_no
            if record.expense_amount <= 0 and record.total_amount > 0:
                record.expense_amount = max(record.total_amount - record.vat_amount - record.sales_tax, 0.0)
            ws.append(record.to_excel_row())
            line_no += 1

        _format_invoice_sheet(ws)
        _autosize(ws)
        if audits is not None:
            summary_ws = wb.create_sheet("Run_Summary")
            summary_ws.append(RUN_SUMMARY_HEADERS)
            for metric, value in _run_summary_rows(records, locked, unlocked, audits):
                summary_ws.append([metric, value])
            _autosize(summary_ws)

            audit_ws = wb.create_sheet("OCR_Audit")
            audit_ws.append(OCR_AUDIT_HEADERS)
            for audit in audits:
                audit_ws.append(audit.to_excel_row())
            _autosize(audit_ws)

        if source_qas is not None:
            qa_ws = wb.create_sheet("Source_QA")
            qa_ws.append(SOURCE_QA_HEADERS)
            for source_qa in source_qas:
                qa_ws.append(source_qa.to_excel_row())
            _autosize(qa_ws)
        wb.save(self.workbook_path)
        wb.save(self.manual_checked_path)
        return len(unlocked)

    def _locked_records(self) -> dict[tuple[str, str, float], InvoiceRecord]:
        if not self.manual_checked_path.exists():
            return {}
        wb = load_workbook(self.manual_checked_path, data_only=True)
        try:
            ws = _invoice_sheet(wb)
            locked: dict[tuple[str, str, float], InvoiceRecord] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not _looks_like_invoice_row(row):
                    continue
                record = InvoiceRecord.from_excel_row(list(row))
                if record.status.strip().casefold() in LOCKED_STATUSES:
                    locked[_lock_key(record)] = record
            return locked
        finally:
            wb.close()


def load_invoice_records(path: Path) -> list[InvoiceRecord]:
    if path.is_dir():
        candidates = sorted(path.glob("*.xlsx"))
        path = next((p for p in candidates if "comparison_report" not in p.name), path / "Invoice_Output.xlsx")
    if not path.exists():
        return []
    try:
        from .reimbursement_excel import INVOICE_EXP_SHEET, load_reimbursement_records

        probe = load_workbook(path, read_only=True)
        try:
            if INVOICE_EXP_SHEET in probe.sheetnames:
                probe.close()
                return load_reimbursement_records(path)
        finally:
            try:
                probe.close()
            except Exception:
                pass
    except Exception:
        pass
    wb = load_workbook(path, data_only=True)
    try:
        ws = _invoice_sheet(wb)
        rows: list[InvoiceRecord] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not _looks_like_invoice_row(row):
                continue
            rows.append(InvoiceRecord.from_excel_row(list(row)))
        return rows
    finally:
        wb.close()


def _run_summary_rows(
    records: list[InvoiceRecord],
    locked: dict[tuple[str, str, float], InvoiceRecord],
    unlocked: list[InvoiceRecord],
    audits: list[OCRAuditRow],
) -> list[tuple[str, object]]:
    source_images = {audit.source_image for audit in audits if audit.source_image}
    noise_filtered = sum(1 for audit in audits if "noise filtered" in audit.decision.casefold())
    codex_attempted = sum(1 for audit in audits if audit.used_codex or audit.codex_error or audit.codex_text or audit.codex_confidence > 0)
    codex_used = sum(1 for audit in audits if audit.used_codex)
    local_agreement = sum(1 for audit in audits if "local OCR agreement".casefold() in audit.decision.casefold())
    return [
        ("Source images", len(source_images)),
        ("Crops processed", len(audits)),
        ("Invoice rows", len(locked) + len(unlocked)),
        ("New rows written", len(unlocked)),
        ("Locked rows preserved", len(locked)),
        ("Noise filtered crops", noise_filtered),
        ("Local OCR agreements", local_agreement),
        ("Codex Scan attempted", codex_attempted),
        ("Codex Scan used", codex_used),
    ]


def _autosize(ws) -> None:
    for column in ws.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)


def _format_invoice_sheet(ws) -> None:
    for cell in ws["A"][1:]:
        cell.number_format = "0000"


def _invoice_sheet(wb):
    if "Invoices" in wb.sheetnames:
        return wb["Invoices"]
    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(max_row=1, values_only=True), ())
        headers = [str(value or "") for value in first_row[: len(EXCEL_HEADERS)]]
        if headers == EXCEL_HEADERS:
            return ws
    return wb.active


def _looks_like_invoice_row(row: tuple[object, ...]) -> bool:
    padded = list(row[: len(EXCEL_HEADERS)]) + [None] * len(EXCEL_HEADERS)
    if not any(value not in (None, "") for value in padded[: len(EXCEL_HEADERS)]):
        return False
    date = str(padded[1] or "").strip()
    seller = str(padded[10] or "").strip()
    status = str(padded[12] or "").strip()
    amount = padded[5]
    return bool(date or seller or status or _to_float_like(amount) > 0)


def _to_float_like(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _lock_key(record: InvoiceRecord) -> tuple[str, str, float]:
    date = normalize_date(record.invoice_date) or record.invoice_date.strip()[:10]
    seller = normalize_text(record.seller).casefold()
    return (date, seller, round(record.total_amount, 2))
