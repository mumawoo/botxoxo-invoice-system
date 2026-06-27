from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .config import Settings
from .excel_store import load_invoice_records
from .queue_worker import telegram_user_output_dir, telegram_user_workbook
from .reimbursement_excel import REVIEW_CROPS_DIR, build_checked_outputs, checked_workbook_path

REIMBURSEMENT_FILE = "Reimbursement_Status.xlsx"
ACTIVE_BATCH_FILE = "active_batch_id.txt"

BATCH_HEADERS = [
    "Submit Batch ID",
    "Submitted At",
    "Telegram User ID",
    "Record Count",
    "Total MXN Amount",
    "Category Totals JSON",
    "Archived Excel",
    "Archived Final Crops",
]


@dataclass(frozen=True)
class ReimbursementSummary:
    record_count: int
    photo_count: int
    total_amount: float
    category_totals: dict[str, float]
    codex_used: int = 0
    failed_count: int = 0
    date_min: str = ""
    date_max: str = ""


@dataclass(frozen=True)
class SubmitResult:
    batch_id: str
    submitted_at: str
    record_count: int
    total_amount: float
    category_totals: dict[str, float]
    archived_excel: Path
    archived_crops: Path
    archived_manual_excel: Path | None = None
    archived_review_crops: Path | None = None
    missing_crops: tuple[str, ...] = ()


def reimbursement_path(settings: Settings, user_id: int) -> Path:
    return telegram_user_output_dir(settings, user_id) / REIMBURSEMENT_FILE


def sync_reimbursement_records(settings: Settings, user_id: int) -> Path:
    path = reimbursement_path(settings, user_id)
    wb = _load_or_create_status_workbook(path)
    wb.save(path)
    wb.close()
    return path


def refresh_checked_outputs(settings: Settings, user_id: int):
    return build_checked_outputs(telegram_user_output_dir(settings, user_id))


def unsubmitted_summary(settings: Settings, user_id: int) -> ReimbursementSummary:
    refresh_checked_outputs(settings, user_id)
    records = _active_records(settings, user_id)
    category_totals: dict[str, float] = {}
    total = 0.0
    dates: list[str] = []
    for record in records:
        amount = round(float(record.total_amount or 0), 2)
        category = record.expense_category or "Other"
        category_totals[category] = round(category_totals.get(category, 0.0) + amount, 2)
        total = round(total + amount, 2)
        if record.invoice_date:
            dates.append(record.invoice_date)
    return ReimbursementSummary(
        record_count=len(records),
        photo_count=0,
        total_amount=round(total, 2),
        category_totals=category_totals,
        date_min=min(dates) if dates else "",
        date_max=max(dates) if dates else "",
    )


def submitted_batches_text(settings: Settings, user_id: int) -> str:
    path = reimbursement_path(settings, user_id)
    if not path.exists():
        return "Submitted / 已提交\nNo submitted batches yet."
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb["Batches"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    finally:
        wb.close()
    if not rows:
        return "Submitted / 已提交\nNo submitted batches yet."
    lines = ["Submitted / 已提交"]
    for row in rows[-10:]:
        lines.append(f"{row[0]} | {row[1]} | records={row[3]} | total MXN={float(row[4] or 0):.2f}")
    return "\n".join(lines)


def submit_unsubmitted(settings: Settings, user_id: int) -> SubmitResult | None:
    checked = refresh_checked_outputs(settings, user_id)
    records = _active_records(settings, user_id)
    if not records:
        return None
    submitted_at = datetime.now().isoformat(timespec="seconds")
    path = reimbursement_path(settings, user_id)
    wb = _load_or_create_status_workbook(path)
    batch_id = _next_submit_batch_id(wb["Batches"], submitted_at)
    total, category_totals = _record_totals(records)
    archived_excel, archived_crops, archived_manual_excel, archived_review_crops = _archive_active_outputs(settings, user_id, batch_id)
    wb["Batches"].append(
        [
            batch_id,
            submitted_at,
            user_id,
            len(records),
            round(total, 2),
            json.dumps(category_totals, ensure_ascii=False, sort_keys=True),
            str(archived_excel),
            str(archived_crops),
        ]
    )
    _autosize(wb["Batches"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    _reset_active_scan_outputs(settings, user_id)
    return SubmitResult(
        batch_id,
        submitted_at,
        len(records),
        round(total, 2),
        category_totals,
        archived_excel,
        archived_crops,
        archived_manual_excel,
        archived_review_crops,
        tuple(checked.missing_crops),
    )


def format_reimbursement_summary(summary: ReimbursementSummary, title: str = "Report / 待报销") -> str:
    lines = [
        title,
        f"Photos: {summary.photo_count}",
        f"Invoice rows: {summary.record_count}",
        f"Total MXN: {summary.total_amount:.2f}",
    ]
    if summary.date_min or summary.date_max:
        lines.append(f"Date range: {summary.date_min or '-'} to {summary.date_max or '-'}")
    if summary.category_totals:
        lines.append("Category totals:")
        for category, amount in sorted(summary.category_totals.items()):
            lines.append(f"- {category}: {amount:.2f}")
    lines.append(f"Failed photos: {summary.failed_count}")
    return "\n".join(lines)


def format_submit_result(result: SubmitResult | None) -> str:
    if result is None:
        return "Submit / 提交\nNo unsubmitted records."
    lines = [
        "Submit / 提交完成",
        f"Batch: {result.batch_id}",
        f"Records: {result.record_count}",
        f"Total MXN: {result.total_amount:.2f}",
        f"Finance Excel: {result.archived_excel}",
        f"Finance crops: {result.archived_crops}",
        "Category totals:",
    ]
    for category, amount in sorted(result.category_totals.items()):
        lines.append(f"- {category}: {amount:.2f}")
    lines.append("Active Excel is archived. New scans start a fresh reimbursement workbook from 001.")
    if result.missing_crops:
        lines.append(f"Warning: missing crop links={len(result.missing_crops)}")
    return "\n".join(lines)


def active_scan_batch_id(output_dir: Path) -> str:
    path = output_dir / ACTIVE_BATCH_FILE
    if path.exists():
        value = path.read_text(encoding="utf-8").strip()
        if value:
            return value
    value = "SCAN-" + datetime.now().strftime("%Y%m%d-%H%M%S")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value, encoding="utf-8")
    return value


def _active_records(settings: Settings, user_id: int):
    output_dir = telegram_user_output_dir(settings, user_id)
    checked = checked_workbook_path(output_dir)
    source = checked if checked.exists() else telegram_user_workbook(settings, user_id)
    return [record for record in load_invoice_records(source) if record.total_amount > 0]


def _record_totals(records) -> tuple[float, dict[str, float]]:
    total = 0.0
    category_totals: dict[str, float] = {}
    for record in records:
        amount = round(float(record.total_amount or 0), 2)
        total = round(total + amount, 2)
        category = record.expense_category or "Other"
        category_totals[category] = round(category_totals.get(category, 0.0) + amount, 2)
    return total, category_totals


def _load_or_create_status_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        if "Batches" not in wb.sheetnames:
            ws = wb.create_sheet("Batches")
            ws.append(BATCH_HEADERS)
        return wb
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"
    ws.append(BATCH_HEADERS)
    return wb


def _next_submit_batch_id(batch_ws, submitted_at: str) -> str:
    date_part = submitted_at[:10].replace("-", "")
    existing = [str(row[0] or "") for row in batch_ws.iter_rows(min_row=2, values_only=True)]
    count = sum(1 for value in existing if value.startswith(f"SUB-{date_part}-"))
    return f"SUB-{date_part}-{count + 1:03d}"


def _archive_active_outputs(settings: Settings, user_id: int, batch_id: str) -> tuple[Path, Path, Path | None, Path | None]:
    output_dir = telegram_user_output_dir(settings, user_id)
    workbook = checked_workbook_path(output_dir)
    manual_workbook = telegram_user_workbook(settings, user_id)
    final_crops = output_dir / "final_crops"
    review_crops = output_dir / REVIEW_CROPS_DIR
    archive_dir = output_dir / "submitted" / batch_id
    archive_dir.mkdir(parents=True, exist_ok=True)
    archived_excel = archive_dir / workbook.name
    archived_crops = archive_dir / "final_crops"
    archived_manual_excel = archive_dir / "manual_check" / manual_workbook.name
    archived_review_crops = archive_dir / "manual_check" / REVIEW_CROPS_DIR
    if workbook.exists():
        shutil.copy2(workbook, archived_excel)
    if final_crops.exists():
        if archived_crops.exists():
            shutil.rmtree(archived_crops)
        shutil.copytree(final_crops, archived_crops)
    manual_saved = None
    review_saved = None
    if manual_workbook.exists():
        archived_manual_excel.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(manual_workbook, archived_manual_excel)
        manual_saved = archived_manual_excel
    if review_crops.exists():
        if archived_review_crops.exists():
            shutil.rmtree(archived_review_crops)
        archived_review_crops.parent.mkdir(parents=True, exist_ok=True)
        shutil.copytree(review_crops, archived_review_crops)
        review_saved = archived_review_crops
    for name in ("processing_state.json", "queue_state.json"):
        path = output_dir / name
        if path.exists():
            target = archive_dir / "manual_check" / name
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(path, target)
    return archived_excel, archived_crops, manual_saved, review_saved


def _reset_active_scan_outputs(settings: Settings, user_id: int) -> None:
    output_dir = telegram_user_output_dir(settings, user_id)
    names = (
        telegram_user_workbook(settings, user_id).name,
        checked_workbook_path(output_dir).name,
        "Invoice_Output.xlsx",
        "Invoice_Manually_Checked.xlsx",
        "processing_state.json",
        "crop_index_state.json",
        ACTIVE_BATCH_FILE,
    )
    for name in names:
        path = output_dir / name
        if path.exists():
            path.unlink()
    for folder in ("crops", "final_crops", REVIEW_CROPS_DIR):
        path = output_dir / folder
        if path.exists():
            shutil.rmtree(path)


def _autosize(ws) -> None:
    for column in ws.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 64)
