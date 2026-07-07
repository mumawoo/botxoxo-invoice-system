from __future__ import annotations

import re
import shutil
import json
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

from .expense_categories import EXPENSE_CATEGORIES, normalize_expense_category
from .fx_rates import ExchangeRate, SAFE_CURRENCY_CODES, fetch_safe_exchange_rates, nearest_rate_on_or_before, normalize_safe_currency
from .models import InvoiceRecord
from .pairing import pair_invoice_payment_slips
from .parsing import normalize_date

REIMBURSEMENT_WORKBOOK_NAME = "\u62a5\u9500\u660e\u7ec6_2026_xlsx.xlsx"
CHECKED_WORKBOOK_NAME = "\u62a5\u9500_checked_2026.xlsx"
INVOICE_EXP_SHEET = "Invoice exp"
SUMMARY_SHEET = "Summary"
FOOD_EXP_SHEET = "Food"
OTHER_EXP_SHEET = "Other"
EXCHANGE_RATE_SHEET = "exchange rate"
CROP_LINKS_SHEET = "_crop_links"
CORRECTED_MARKERS = {"corrected", "correct", "ok"}
DELETED_MARKERS = {"deleted", "delete", "\u5220\u9664", "\u5220\u6389"}
CROPS_DIR = "crops"
REVIEW_CROPS_DIR = "review_crops"
FINAL_CROPS_DIR = "final_crops"
FINAL_CROPS_MANIFEST = "final_crops_manifest.json"
GROUP_ARCHIVE_DIR = "group_archive"
MANUAL_STATUS_HEADER = "Manual status"
TRACE_ID_HEADER = "Trace ID"
SYSTEM_NOTE_HEADER = "System note"

REIMBURSEMENT_HEADERS = [
    "No.",
    MANUAL_STATUS_HEADER,
    "Date",
    "MXN Amount",
    "Type",
    "\u539f\u5e01\u79cd",
    "\u539f\u91d1\u989d",
    "\u6c47\u7387",
    "Merchant",
    "Detail",
    "Accounting Category",
    TRACE_ID_HEADER,
    SYSTEM_NOTE_HEADER,
    "Invoice link",
]

CHECKED_HEADERS = [
    header for header in REIMBURSEMENT_HEADERS if header not in {"Invoice link", MANUAL_STATUS_HEADER, SYSTEM_NOTE_HEADER}
] + ["Invoice link"]
LINK_INDEX = REIMBURSEMENT_HEADERS.index("Invoice link")

TYPE_LABELS_ZH = {
    "Food": "\u9910\u996e",
    "Gas": "\u6c7d\u6cb9",
    "Car repair": "\u8f66\u8f86\u7ef4\u4fee",
    "Toll/Parking": "\u8fc7\u8def\u8d39\u505c\u8f66\u8d39",
    "Utilities": "\u6c34\u7535\u7164",
    "Internet": "\u7f51\u7edc\u8d39",
    "Phone": "\u7535\u8bdd\u8d39",
    "Office supplies": "\u529e\u516c\u7528\u54c1",
    "Hotel": "\u4f4f\u5bbf",
    "Flight": "\u673a\u7968",
    "Other": "\u5176\u4ed6",
}

EXCHANGE_RATE_HEADERS = ["\u65e5\u671f", *SAFE_CURRENCY_CODES]
FetchRates = Callable[..., list[ExchangeRate]]

HEADER_ALIASES = {
    "No.": {"no.", "no", "number", "\u5e8f\u53f7"},
    "Invoice link": {"invoice link", "link", "crop", "final crop", "\u56fe\u7247", "\u56fe\u7247\u94fe\u63a5"},
    "Date": {"date", "invoice date", "\u65e5\u671f"},
    "MXN Amount": {"mxn amount", "amount mxn", "\u62a5\u9500\u91d1\u989d", "\u6bd4\u7d22\u91d1\u989d"},
    "Type": {"type", "\u7c7b\u578b", "\u8d39\u7528\u7c7b\u578b"},
    "\u539f\u5e01\u79cd": {"\u539f\u5e01\u79cd", "original currency", "currency"},
    "\u539f\u91d1\u989d": {"\u539f\u91d1\u989d", "original amount"},
    "\u6c47\u7387": {"\u6c47\u7387", "exchange rate", "rate"},
    "Merchant": {"merchant", "seller", "\u5546\u6237"},
    "Detail": {"detail", "contents", "\u660e\u7ec6"},
    "Accounting Category": {"accounting category", "category", "\u79d1\u76ee"},
    TRACE_ID_HEADER: {"trace id", "trace", "crop id", "original id", "\u8ffd\u8e2a\u7d22\u5f15", "\u539f\u59cb\u7f16\u53f7"},
    MANUAL_STATUS_HEADER: {"manual status", "status", "\u590d\u6838\u72b6\u6001", "\u4eba\u5de5\u72b6\u6001"},
    SYSTEM_NOTE_HEADER: {"system note", "note", "system remark", "remark", "\u7cfb\u7edf\u5907\u6ce8", "\u5907\u6ce8"},
}


@dataclass(frozen=True)
class ReimbursementWriteResult:
    rows_written: int
    workbook_path: Path
    rates_updated: bool
    fx_error: str = ""


@dataclass(frozen=True)
class ExchangeRateUpdateResult:
    workbook_path: Path
    start_date: date
    end_date: date
    existing_rows: int
    fetched_rows: int
    total_rows: int
    error: str = ""


@dataclass(frozen=True)
class CheckedBuildResult:
    workbook_path: Path
    final_crops_dir: Path
    records_written: int
    crops_written: int
    missing_crops: list[str]


@dataclass(frozen=True)
class RerunCheckedResult:
    archive_dir: Path
    workbook_path: Path
    checked_path: Path
    records_written: int
    crops_written: int
    moved: tuple[str, ...]
    changed: tuple[str, ...]
    warnings: tuple[str, ...] = ()


@dataclass(frozen=True)
class ManualChangeResult:
    crop_id: str
    row_idx: int
    merchant: str
    before: dict[str, object]
    after: dict[str, object]
    status: str
    workbook_path: Path


@dataclass(frozen=True)
class GroupReimbursementResult:
    crop_ids: tuple[str, ...]
    primary_id: str
    deleted_ids: tuple[str, ...]
    primary_row: int
    deleted_rows: tuple[int, ...]
    invoice_date: str
    seller: str
    category: str
    currency: str
    total_amount: float
    tips: float
    crop_path: Path | None
    archive_dir: Path | None
    warnings: tuple[str, ...]
    workbook_path: Path


@dataclass(frozen=True)
class _CheckedFinanceRow:
    number: str
    sheet_name: str
    values: list[object]
    row_idx: int


def reimbursement_workbook_path(output_dir: Path) -> Path:
    return output_dir / REIMBURSEMENT_WORKBOOK_NAME


def checked_workbook_path(output_dir: Path) -> Path:
    return output_dir / CHECKED_WORKBOOK_NAME


def next_manual_trace_id(workbook_path: Path) -> int:
    if not workbook_path.exists():
        return 1
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = _ensure_manual_sheet(wb)
        columns = _header_columns(ws)
        max_trace = 0
        for row_idx in range(2, ws.max_row + 1):
            values = _row_values_by_header(ws, row_idx, columns)
            if not _looks_like_reimbursement_row(values):
                continue
            trace = _trace_id_from_link(ws.cell(row_idx, columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))).value)
            trace = trace or str(ws.cell(row_idx, columns.get(TRACE_ID_HEADER, 0)).value or "")
            number = _to_int(trace) or _to_int(ws.cell(row_idx, columns.get("No.", 1)).value)
            max_trace = max(max_trace, number)
        return max(max_trace + 1, 1)
    finally:
        wb.close()


def _is_food_category(category: object, evidence: str = "") -> bool:
    return normalize_expense_category(str(category or ""), evidence) == "Food"


def _sheet_name_for_values(values: list[object]) -> str:
    category = _row_value(values, "Accounting Category") or _row_value(values, "Type") or ""
    return FOOD_EXP_SHEET if _is_food_category(category) else OTHER_EXP_SHEET


def _ensure_manual_sheet(wb):
    if INVOICE_EXP_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(INVOICE_EXP_SHEET, 0)
        ws.append(REIMBURSEMENT_HEADERS)
    ws = wb[INVOICE_EXP_SHEET]
    ws.sheet_state = "visible"
    _ensure_headers(ws)
    _merge_old_split_sheets_to_manual(wb, ws)
    _format_invoice_exp(ws)
    return ws


def load_reimbursement_records(path: Path) -> list[InvoiceRecord]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    try:
        ws = _ensure_manual_sheet(wb)
        records: list[InvoiceRecord] = []
        columns = _header_columns(ws)
        for row_idx in range(2, ws.max_row + 1):
            values = _row_values_by_header(ws, row_idx, columns)
            if _row_is_deleted(ws, row_idx) or not _looks_like_reimbursement_row(values):
                continue
            records.append(_record_from_reimbursement_row(values))
        return records
    finally:
        wb.close()


def change_reimbursement_record(
    output_dir: Path,
    crop_id: str,
    *,
    invoice_date: str | None = None,
    category: str | None = None,
    amount: float | None = None,
    currency: str | None = None,
    comment: str | None = None,
    status: str = "ok",
) -> ManualChangeResult:
    workbook_path = reimbursement_workbook_path(output_dir)
    if not workbook_path.exists():
        raise FileNotFoundError(str(workbook_path))
    normalized_crop_id = _normalize_crop_id(crop_id)
    wb = load_workbook(workbook_path)
    try:
        ws = _ensure_manual_sheet(wb)
        columns = _header_columns(ws)
        _rewrite_crop_links_to_review(ws, output_dir, columns)
        _reconcile_manual_links_to_processing_state(output_dir, ws, columns)
        columns = _header_columns(ws)
        _sync_no_to_crop_ids(ws, columns)
        row_idx = _find_row_by_crop_id(ws, columns, normalized_crop_id)
        if row_idx is None:
            raise LookupError(f"Cannot find crop {normalized_crop_id}")
        before = _row_snapshot(ws, row_idx, columns)
        rates = _rates_from_workbook(wb)
        _apply_manual_change(
            ws,
            row_idx,
            columns,
            rates,
            invoice_date=invoice_date,
            category=category,
            amount=amount,
            currency=currency,
            comment=comment,
            status=status,
        )
        after = _row_snapshot(ws, row_idx, columns)
        wb.save(workbook_path)
        return ManualChangeResult(
            normalized_crop_id,
            row_idx,
            str(after.get("Merchant") or before.get("Merchant") or "Unknown"),
            before,
            after,
            str(after.get(MANUAL_STATUS_HEADER) or status),
            workbook_path,
        )
    finally:
        wb.close()


def remove_reimbursement_rows_by_trace_ids(output_dir: Path, trace_ids: set[str]) -> int:
    workbook_path = reimbursement_workbook_path(output_dir)
    if not workbook_path.exists() or not trace_ids:
        return 0
    normalized = {_normalize_crop_id(trace_id) for trace_id in trace_ids}
    wb = load_workbook(workbook_path)
    try:
        ws = _ensure_manual_sheet(wb)
        columns = _header_columns(ws)
        _sync_no_to_crop_ids(ws, columns)
        removed = 0
        for row_idx in range(ws.max_row, 1, -1):
            row_trace = _row_trace_id(ws, row_idx, columns)
            if row_trace in normalized:
                ws.delete_rows(row_idx, 1)
                removed += 1
        crop_links = _crop_link_map(wb)
        if crop_links:
            crop_links = {
                primary: [support for support in supports if (_trace_id_from_link(support) or "") not in normalized]
                for primary, supports in crop_links.items()
                if (_trace_id_from_link(primary) or "") not in normalized
            }
            crop_links = {primary: supports for primary, supports in crop_links.items() if supports}
            _write_crop_links_sheet(wb, crop_links)
        _format_invoice_exp(ws)
        wb.save(workbook_path)
        return removed
    finally:
        wb.close()


def preview_reimbursement_group(output_dir: Path, crop_ids: list[str] | tuple[str, ...]) -> GroupReimbursementResult:
    workbook_path = reimbursement_workbook_path(output_dir)
    if not workbook_path.exists():
        raise FileNotFoundError(str(workbook_path))
    normalized = _normalize_group_crop_ids(crop_ids)
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = _ensure_manual_sheet(wb)
        columns = _header_columns(ws)
        _sync_no_to_crop_ids(ws, columns)
        group_rows = _group_rows_for_crop_ids(ws, columns, normalized, output_dir)
        records = [item[2] for item in group_rows]
        merged, warnings = _merged_group_record(records, normalized)
        primary_id = min(normalized, key=lambda value: int(value))
        primary_row = next(row_idx for crop_id, row_idx, _record, _path in group_rows if crop_id == primary_id)
        deleted_rows = tuple(row_idx for crop_id, row_idx, _record, _path in group_rows if crop_id != primary_id)
        return GroupReimbursementResult(
            tuple(normalized),
            primary_id,
            tuple(crop_id for crop_id in normalized if crop_id != primary_id),
            primary_row,
            deleted_rows,
            merged.invoice_date,
            merged.seller,
            normalize_expense_category(merged.expense_category, f"{merged.seller} {merged.contents}"),
            _normalize_currency(merged.currency),
            round(merged.total_amount, 2),
            round(merged.tips, 2),
            None,
            None,
            tuple(warnings),
            workbook_path,
        )
    finally:
        wb.close()


def apply_reimbursement_group(output_dir: Path, crop_ids: list[str] | tuple[str, ...]) -> GroupReimbursementResult:
    workbook_path = reimbursement_workbook_path(output_dir)
    if not workbook_path.exists():
        raise FileNotFoundError(str(workbook_path))
    normalized = _normalize_group_crop_ids(crop_ids)
    wb = load_workbook(workbook_path, data_only=False)
    try:
        ws = _ensure_manual_sheet(wb)
        columns = _header_columns(ws)
        _sync_no_to_crop_ids(ws, columns)
        columns = _header_columns(ws)
        group_rows = _group_rows_for_crop_ids(ws, columns, normalized, output_dir)
        records = [item[2] for item in group_rows]
        sources = [item[3] for item in group_rows if item[3] is not None]
        merged, warnings = _merged_group_record(records, normalized)
        primary_id = min(normalized, key=lambda value: int(value))
        primary_row = next(row_idx for crop_id, row_idx, _record, _path in group_rows if crop_id == primary_id)
        deleted_rows = tuple(row_idx for crop_id, row_idx, _record, _path in group_rows if crop_id != primary_id)
        deleted_ids = tuple(crop_id for crop_id in normalized if crop_id != primary_id)
        if not sources:
            raise FileNotFoundError("Cannot find group crop images")

        archive_dir = output_dir / GROUP_ARCHIVE_DIR / "_".join(normalized)
        archive_dir.mkdir(parents=True, exist_ok=True)
        for source in sources:
            if source.exists():
                shutil.copy2(source, archive_dir / source.name)

        rates = _rates_from_workbook(wb)
        category = normalize_expense_category(merged.expense_category, f"{merged.seller} {merged.contents}")
        target = output_dir / CROPS_DIR / _grouped_crop_name(primary_id, merged, category)
        target.parent.mkdir(parents=True, exist_ok=True)
        temp_target = target.with_name(f".{target.stem}.tmp{target.suffix}")
        if not _write_final_crop(sources, temp_target):
            raise FileNotFoundError("Cannot create combined group crop")
        if target.exists():
            target.unlink()
        temp_target.replace(target)

        merged.line_no = int(primary_id)
        merged.expense_category = category
        merged.crop_image = str(target)
        merged.supporting_crop_images = []
        _write_reimbursement_row(ws, primary_row, merged, rates)
        columns = _header_columns(ws)
        ws.cell(primary_row, columns[MANUAL_STATUS_HEADER]).value = "correct"
        ws.cell(primary_row, columns[SYSTEM_NOTE_HEADER]).value = f"Grouped {' + '.join(normalized)}; originals archived to {GROUP_ARCHIVE_DIR}/{'_'.join(normalized)}"
        ws.cell(primary_row, columns[TRACE_ID_HEADER]).value = primary_id

        for crop_id, row_idx, _record, _path in group_rows:
            if crop_id == primary_id:
                continue
            ws.cell(row_idx, columns[MANUAL_STATUS_HEADER]).value = "delete"
            ws.cell(row_idx, columns[SYSTEM_NOTE_HEADER]).value = f"Grouped into {primary_id}"

        crop_links = _crop_link_map(wb)
        for crop_id in normalized:
            for key in list(crop_links):
                if (_trace_id_from_link(key) or "") == crop_id:
                    crop_links.pop(key, None)
        _write_crop_links_sheet(wb, crop_links)
        _format_invoice_exp(ws)
        wb.save(workbook_path)
        _remove_group_source_files(sources, keep=target)
        return GroupReimbursementResult(
            tuple(normalized),
            primary_id,
            deleted_ids,
            primary_row,
            deleted_rows,
            merged.invoice_date,
            merged.seller,
            category,
            _normalize_currency(merged.currency),
            round(merged.total_amount, 2),
            round(merged.tips, 2),
            target,
            archive_dir,
            tuple(warnings),
            workbook_path,
        )
    finally:
        wb.close()


def available_crop_ids(output_dir: Path, limit: int = 8) -> list[str]:
    ids: list[str] = []
    for folder in (CROPS_DIR, REVIEW_CROPS_DIR, FINAL_CROPS_DIR):
        directory = output_dir / folder
        if not directory.exists():
            continue
        for path in sorted(directory.glob("*.jpg"), key=lambda item: (item.stat().st_mtime, item.name)):
            match = re.match(r"(\d{3,})_", path.name)
            if match and match.group(1) not in ids:
                ids.append(match.group(1))
    return ids[-limit:]


def build_checked_outputs(output_dir: Path, force: bool = False) -> CheckedBuildResult:
    workbook_path = reimbursement_workbook_path(output_dir)
    checked_path = checked_workbook_path(output_dir)
    final_dir = output_dir / FINAL_CROPS_DIR
    if not workbook_path.exists():
        return CheckedBuildResult(checked_path, final_dir, 0, 0, [])

    wb = load_workbook(workbook_path, data_only=False)
    try:
        manual_ws = _ensure_manual_sheet(wb)
        if manual_ws is None:
            return CheckedBuildResult(checked_path, final_dir, 0, 0, [])
        rates = _rates_from_workbook(wb)
        support_map = _crop_link_map(wb)
        rows: list[list[object]] = []
        row_sheets: list[str] = []
        crop_sources: list[list[Path]] = []
        missing: list[str] = []
        columns = _header_columns(manual_ws)
        _rewrite_crop_links_to_review(manual_ws, workbook_path.parent, columns)
        _reconcile_manual_links_to_processing_state(workbook_path.parent, manual_ws, columns)
        columns = _header_columns(manual_ws)
        _sync_no_to_crop_ids(manual_ws, columns)
        for row_idx in range(2, manual_ws.max_row + 1):
            values = _row_values_by_header(manual_ws, row_idx, columns)
            if _row_is_deleted(manual_ws, row_idx) or not _looks_like_reimbursement_row(values):
                continue
            rows.append(values)
            row_sheets.append(_sheet_name_for_values(values))
            link = _row_value(values, "Invoice link")
            source = _resolve_crop_source(workbook_path.parent, link)
            sources = [source] if source is not None else []
            for support in support_map.get(Path(str(link or "")).name, []):
                support_source = _resolve_crop_source(workbook_path.parent, support)
                if support_source is not None and support_source not in sources:
                    sources.append(support_source)
            crop_sources.append(sources)
            if source is None:
                missing.append(str(link or f"{manual_ws.title} row {row_idx}"))
        try:
            wb.save(workbook_path)
        except OSError:
            pass
    finally:
        wb.close()

    manifest_path = output_dir / FINAL_CROPS_MANIFEST
    force = force or not manifest_path.exists()
    if force:
        clear_generated_crops(final_dir, _source_names_inside_dir(crop_sources, final_dir))
    manifest = {} if force else _load_final_crops_manifest(manifest_path)
    next_manifest: dict[str, dict[str, object]] = {}
    current_final_paths: set[str] = set()
    checked_wb = Workbook()
    checked_summary_ws = checked_wb.active
    checked_summary_ws.title = SUMMARY_SHEET
    checked_food_ws = checked_wb.create_sheet(FOOD_EXP_SHEET)
    checked_food_ws.title = FOOD_EXP_SHEET
    checked_food_ws.append(CHECKED_HEADERS)
    checked_other_ws = checked_wb.create_sheet(OTHER_EXP_SHEET)
    checked_other_ws.append(CHECKED_HEADERS)
    checked_legacy_ws = checked_wb.create_sheet(INVOICE_EXP_SHEET)
    checked_legacy_ws.append(CHECKED_HEADERS)
    checked_legacy_ws.sheet_state = "hidden"
    checked_sheets = {FOOD_EXP_SHEET: checked_food_ws, OTHER_EXP_SHEET: checked_other_ws}
    crops_written = 0
    for index, values in enumerate(rows, start=1):
        output = list(values[: len(REIMBURSEMENT_HEADERS)])
        output += [None] * (len(REIMBURSEMENT_HEADERS) - len(output))
        output[0] = index
        sources = crop_sources[index - 1]
        final_subdir = _final_crop_subdir(row_sheets[index - 1])
        if sources:
            target = final_dir / final_subdir / _checked_crop_name(index, output, source=sources[0])
            target.parent.mkdir(parents=True, exist_ok=True)
            trace_id = _trace_id_from_link(sources[0]) or str(_row_value(output, TRACE_ID_HEADER) or f"{index:03d}").zfill(3)
            entry = _final_crop_manifest_entry(output_dir, output, row_sheets[index - 1], sources, target)
            if _ensure_final_crop(manifest, trace_id, entry, sources, target):
                output[LINK_INDEX] = f"{FINAL_CROPS_DIR}/{final_subdir}/{target.name}"
                crops_written += 1
                next_manifest[trace_id] = entry
                current_final_paths.add(str(entry["path"]))
        checked_output = _reorder_row_for_headers(output, REIMBURSEMENT_HEADERS, CHECKED_HEADERS)
        checked_ws = checked_sheets.get(row_sheets[index - 1], checked_other_ws)
        checked_ws.append(checked_output)
        _format_row(checked_ws, checked_ws.max_row, _header_columns(checked_ws))
        output_link = _row_value(output, "Invoice link")
        if output_link:
            link_col = _header_columns(checked_ws).get("Invoice link", len(CHECKED_HEADERS))
            link_cell = checked_ws.cell(checked_ws.max_row, link_col)
            link_cell.hyperlink = str(output_link)
            link_cell.style = "Hyperlink"
        checked_legacy_ws.append(checked_output)
        _format_row(checked_legacy_ws, checked_legacy_ws.max_row, _header_columns(checked_legacy_ws))
        if output_link:
            link_col = _header_columns(checked_legacy_ws).get("Invoice link", len(CHECKED_HEADERS))
            link_cell = checked_legacy_ws.cell(checked_legacy_ws.max_row, link_col)
            link_cell.hyperlink = str(output_link)
            link_cell.style = "Hyperlink"
    for ws in (checked_food_ws, checked_other_ws, checked_legacy_ws):
        _format_invoice_exp(ws, CHECKED_HEADERS)
        _autosize(ws, max_col=len(CHECKED_HEADERS))
    _write_food_summary_sheet(checked_summary_ws, rows, row_sheets)
    _write_exchange_rate_sheet(checked_wb, rates)
    checked_path.parent.mkdir(parents=True, exist_ok=True)
    checked_wb.save(checked_path)
    checked_wb.close()
    _prune_stale_final_crops(output_dir, manifest, next_manifest, current_final_paths)
    _save_final_crops_manifest(manifest_path, next_manifest)
    _save_checked_baseline(output_dir)
    return CheckedBuildResult(checked_path, final_dir, len(rows), crops_written, missing)


def _write_food_summary_sheet(ws, rows: list[list[object]], row_sheets: list[str]) -> None:
    ws.title = SUMMARY_SHEET
    ws["A1"] = "餐饮消费汇总"
    ws["A1"].font = Font(bold=True, size=14)
    dated_rows: list[tuple[date, list[object], str]] = []
    for values, sheet_name in zip(rows, row_sheets):
        invoice_date = _parse_date(_row_value(values, "Date"))
        if invoice_date is not None:
            dated_rows.append((invoice_date, values, sheet_name))
    if not dated_rows:
        ws["A3"] = "No dated reimbursement records"
        _autosize(ws, max_col=3)
        return

    start = min(item[0] for item in dated_rows)
    end = max(item[0] for item in dated_rows)
    day_count = (end - start).days + 1
    daily_food: dict[date, float] = {}
    for invoice_date, values, sheet_name in dated_rows:
        if sheet_name != FOOD_EXP_SHEET:
            continue
        amount = round(_to_float(_row_value(values, "MXN Amount")), 2)
        daily_food[invoice_date] = round(daily_food.get(invoice_date, 0.0) + amount, 2)

    all_days = [start + timedelta(days=offset) for offset in range(day_count)]
    food_total = round(sum(daily_food.values()), 2)
    average = round(food_total / day_count, 2) if day_count else 0.0
    missing_days = [day for day in all_days if round(daily_food.get(day, 0.0), 2) <= 0]

    summary_rows = [
        ("本次报销日期范围", f"{start.isoformat()} 至 {end.isoformat()}"),
        ("共", f"{day_count} 天"),
        ("餐饮总额", f"{food_total:.2f} MXN"),
        ("平均每天餐饮消费", f"{average:.2f} MXN"),
        ("未报餐费天数", f"{len(missing_days)} 天"),
        ("未报餐费日期", ", ".join(day.isoformat() for day in missing_days) if missing_days else "无"),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws.cell(row_idx, 1).value = label
        ws.cell(row_idx, 1).font = Font(bold=True)
        ws.cell(row_idx, 2).value = value

    table_row = 11
    headers = ["Date", "Food MXN", "Food reported?"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(table_row, col_idx)
        cell.value = header
        cell.font = Font(bold=True)
    for row_offset, day in enumerate(all_days, start=1):
        row_idx = table_row + row_offset
        food_amount = round(daily_food.get(day, 0.0), 2)
        ws.cell(row_idx, 1).value = day
        ws.cell(row_idx, 1).number_format = "yyyy-mm-dd"
        ws.cell(row_idx, 2).value = food_amount
        ws.cell(row_idx, 2).number_format = "#,##0.00"
        ws.cell(row_idx, 3).value = "Yes" if food_amount > 0 else "No"
    ws.freeze_panes = "A12"
    _autosize(ws, max_col=len(headers))


def rerun_checked_from_finance_edits(output_dir: Path) -> RerunCheckedResult:
    workbook_path = reimbursement_workbook_path(output_dir)
    checked_path = checked_workbook_path(output_dir)
    baseline_path = _checked_baseline_workbook_path(output_dir)
    if not checked_path.exists():
        raise FileNotFoundError(str(checked_path))
    if not baseline_path.exists():
        raise FileNotFoundError(str(baseline_path))

    baseline_rows = _read_checked_finance_rows(baseline_path)
    current_rows = _read_checked_finance_rows(checked_path)
    if not baseline_rows:
        raise ValueError("Baseline checked Excel has no finance rows")
    if not current_rows:
        raise ValueError("Current checked Excel has no finance rows")
    if set(baseline_rows) != set(current_rows):
        missing = sorted(set(baseline_rows) - set(current_rows))
        extra = sorted(set(current_rows) - set(baseline_rows))
        detail = []
        if missing:
            detail.append(f"missing checked No.: {', '.join(missing)}")
        if extra:
            detail.append(f"extra checked No.: {', '.join(extra)}")
        raise ValueError("; ".join(detail))

    changed: list[str] = []
    moved: list[str] = []
    warnings: list[str] = []
    for number in sorted(current_rows, key=_checked_number_sort_key):
        old = baseline_rows[number]
        new = current_rows[number]
        if old.sheet_name != new.sheet_name:
            moved.append(f"{number} {old.sheet_name} -> {new.sheet_name}")
        elif _checked_row_signature(old.values) != _checked_row_signature(new.values):
            changed.append(number)
            warnings.append(f"{number}: finance row values changed")
    if not moved and not changed:
        return RerunCheckedResult(output_dir, workbook_path, checked_path, len(current_rows), 0, (), (), ())

    archive_dir = _archive_rerun_inputs(output_dir)
    _migrate_current_finance_crops_to_review(output_dir, current_rows.values())
    _write_manual_workbook_from_checked(output_dir, current_rows, baseline_rows)
    result = build_checked_outputs(output_dir)
    return RerunCheckedResult(
        archive_dir,
        workbook_path,
        checked_path,
        result.records_written,
        result.crops_written,
        tuple(moved),
        tuple(changed),
        tuple(warnings),
    )


def _reorder_row_for_headers(values: list[object], source_headers: list[str], target_headers: list[str]) -> list[object]:
    by_header = {header: values[index] if index < len(values) else None for index, header in enumerate(source_headers)}
    return [by_header.get(header) for header in target_headers]


def _checked_baseline_dir(output_dir: Path) -> Path:
    return output_dir / "checked_baseline"


def _checked_baseline_workbook_path(output_dir: Path) -> Path:
    return _checked_baseline_dir(output_dir) / CHECKED_WORKBOOK_NAME


def _save_checked_baseline(output_dir: Path) -> None:
    baseline_dir = _checked_baseline_dir(output_dir)
    checked_path = checked_workbook_path(output_dir)
    final_dir = output_dir / FINAL_CROPS_DIR
    baseline_dir.mkdir(parents=True, exist_ok=True)
    if checked_path.exists():
        shutil.copy2(checked_path, baseline_dir / checked_path.name)
    baseline_crops = baseline_dir / FINAL_CROPS_DIR
    if baseline_crops.exists():
        shutil.rmtree(baseline_crops)
    if final_dir.exists():
        shutil.copytree(final_dir, baseline_crops)


def _read_checked_finance_rows(path: Path) -> dict[str, _CheckedFinanceRow]:
    wb = load_workbook(path, data_only=False)
    try:
        rows: dict[str, _CheckedFinanceRow] = {}
        for sheet_name in (FOOD_EXP_SHEET, OTHER_EXP_SHEET):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            columns = _checked_header_columns(ws)
            if not columns:
                continue
            for row_idx in range(2, ws.max_row + 1):
                values = _checked_row_values_by_header(ws, row_idx, columns)
                if not _looks_like_checked_row(values):
                    continue
                number = _checked_number(values)
                if not number:
                    raise ValueError(f"Missing checked No. in {sheet_name} row {row_idx}")
                if number in rows:
                    raise ValueError(f"Duplicate checked No.: {number}")
                link = str(_checked_value(values, "Invoice link") or "").strip()
                if not link:
                    raise ValueError(f"Missing Invoice link for checked No. {number}")
                rows[number] = _CheckedFinanceRow(number, sheet_name, values, row_idx)
        return rows
    finally:
        wb.close()


def _checked_header_columns(ws) -> dict[str, int]:
    raw = {str(ws.cell(1, col).value or "").strip(): col for col in range(1, ws.max_column + 1)}
    return {header: raw[header] for header in CHECKED_HEADERS if header in raw}


def _checked_row_values_by_header(ws, row_idx: int, columns: dict[str, int]) -> list[object]:
    values = [ws.cell(row_idx, columns[header]).value if header in columns else None for header in CHECKED_HEADERS]
    link_col = columns.get("Invoice link")
    if link_col:
        cell = ws.cell(row_idx, link_col)
        if cell.hyperlink and cell.hyperlink.target:
            values[CHECKED_HEADERS.index("Invoice link")] = cell.hyperlink.target
    return values


def _looks_like_checked_row(values: list[object]) -> bool:
    return bool(
        values
        and any(value not in (None, "") for value in values)
        and (
            _checked_value(values, "No.")
            or _checked_value(values, "Date")
            or _checked_value(values, "MXN Amount")
            or _checked_value(values, "Merchant")
        )
    )


def _checked_number(values: list[object]) -> str:
    number = _to_int(_checked_value(values, "No."))
    if number:
        return f"{number:03d}"
    link_id = _checked_number_from_link(_checked_value(values, "Invoice link"))
    return link_id


def _checked_number_from_link(value: object) -> str:
    match = re.match(r"(\d{3,})[a-z]?\_", Path(str(value or "")).name)
    return match.group(1) if match else ""


def _checked_number_sort_key(value: str) -> tuple[int, str]:
    return (_to_int(value) or 0, value)


def _checked_row_signature(values: list[object]) -> tuple[object, ...]:
    ignored = {"No.", TRACE_ID_HEADER, "Invoice link"}
    return tuple("" if _checked_value(values, header) is None else _checked_value(values, header) for header in CHECKED_HEADERS if header not in ignored)


def _archive_rerun_inputs(output_dir: Path) -> Path:
    archive_root = output_dir / "rerun_archive"
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    archive_dir = archive_root / stamp
    suffix = 1
    while archive_dir.exists():
        suffix += 1
        archive_dir = archive_root / f"{stamp}-{suffix:02d}"
    archive_dir.mkdir(parents=True, exist_ok=True)
    for path in [
        reimbursement_workbook_path(output_dir),
        checked_workbook_path(output_dir),
        output_dir / "processing_state.json",
        output_dir / "queue_state.json",
    ]:
        if path.exists():
            shutil.copy2(path, archive_dir / path.name)
    for folder in [output_dir / FINAL_CROPS_DIR, output_dir / CROPS_DIR, output_dir / REVIEW_CROPS_DIR, _checked_baseline_dir(output_dir)]:
        if folder.exists():
            shutil.copytree(folder, archive_dir / folder.name)
    return archive_dir


def _migrate_current_finance_crops_to_review(output_dir: Path, rows: Iterable[_CheckedFinanceRow]) -> None:
    crops_dir = output_dir / CROPS_DIR
    crops_dir.mkdir(parents=True, exist_ok=True)
    for row in rows:
        for source in _finance_crop_sources(output_dir, row):
            target = crops_dir / source.name
            if not target.exists():
                shutil.copy2(source, target)


def _finance_crop_sources(output_dir: Path, row: _CheckedFinanceRow) -> list[Path]:
    link = _checked_value(row.values, "Invoice link")
    source = _resolve_crop_source(output_dir, link)
    if source is None:
        raise ValueError(f"Cannot find finance crop for checked No. {row.number}: {link}")
    sources = [source]
    match = re.match(r"(\d{3,})([a-z])\_", source.name)
    if match:
        prefix = match.group(1)
        for sibling in sorted(source.parent.glob(f"{prefix}[a-z]_*")):
            if sibling.is_file() and sibling not in sources:
                sources.append(sibling)
    return sources


def _write_manual_workbook_from_checked(
    output_dir: Path,
    current_rows: dict[str, _CheckedFinanceRow],
    baseline_rows: dict[str, _CheckedFinanceRow],
) -> None:
    workbook_path = reimbursement_workbook_path(output_dir)
    old_rates = _read_existing_exchange_rate_rows(workbook_path)
    wb = Workbook()
    ws = wb.active
    ws.title = INVOICE_EXP_SHEET
    ws.append(REIMBURSEMENT_HEADERS)
    crop_links: dict[str, list[str]] = {}
    for number in sorted(current_rows, key=_checked_number_sort_key):
        row = current_rows[number]
        baseline = baseline_rows[number]
        manual_values = _manual_values_from_checked(output_dir, row, baseline)
        _append_values_with_link(ws, manual_values)
        support_links = _supporting_review_links_for_checked(output_dir, row)
        if support_links:
            crop_links[Path(str(_row_value(manual_values, "Invoice link"))).name] = support_links
    _format_invoice_exp(ws)
    _autosize(ws, max_col=len(REIMBURSEMENT_HEADERS))
    _write_exchange_rate_rows(wb, old_rates)
    _write_crop_links_sheet(wb, crop_links)
    wb.save(workbook_path)
    wb.close()


def _manual_values_from_checked(output_dir: Path, row: _CheckedFinanceRow, baseline: _CheckedFinanceRow) -> list[object]:
    values_by_header = {header: row.values[index] if index < len(row.values) else None for index, header in enumerate(CHECKED_HEADERS)}
    moved = row.sheet_name != baseline.sheet_name
    category = _rerun_category(row, moved)
    final_name = Path(str(values_by_header.get("Invoice link") or "")).name
    trace_id = str(values_by_header.get(TRACE_ID_HEADER) or _trace_id_from_link(final_name) or "").zfill(3)
    return [
        _to_int(trace_id) or _to_int(values_by_header.get("No.")) or _to_int(row.number),
        "correct" if moved or _checked_row_signature(row.values) != _checked_row_signature(baseline.values) else "",
        values_by_header.get("Date"),
        values_by_header.get("MXN Amount"),
        _type_label_zh(category),
        values_by_header.get("\u539f\u5e01\u79cd"),
        values_by_header.get("\u539f\u91d1\u989d"),
        values_by_header.get("\u6c47\u7387"),
        values_by_header.get("Merchant"),
        values_by_header.get("Detail"),
        category,
        trace_id,
        "",
        f"{CROPS_DIR}/{final_name}",
    ]


def _rerun_category(row: _CheckedFinanceRow, moved: bool) -> str:
    current_category = normalize_expense_category(str(_checked_value(row.values, "Accounting Category") or ""))
    if row.sheet_name == FOOD_EXP_SHEET:
        return FOOD_EXP_SHEET
    if moved:
        return OTHER_EXP_SHEET
    return OTHER_EXP_SHEET if current_category == FOOD_EXP_SHEET else current_category


def _supporting_review_links_for_checked(output_dir: Path, row: _CheckedFinanceRow) -> list[str]:
    sources = _finance_crop_sources(output_dir, row)
    return [_relative_crop_link(output_dir, source) for source in sources[1:]]


def _read_existing_exchange_rate_rows(workbook_path: Path) -> list[list[object]]:
    if not workbook_path.exists():
        return [EXCHANGE_RATE_HEADERS]
    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception:
        return [EXCHANGE_RATE_HEADERS]
    try:
        if EXCHANGE_RATE_SHEET not in wb.sheetnames:
            return [EXCHANGE_RATE_HEADERS]
        ws = wb[EXCHANGE_RATE_SHEET]
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        return rows or [EXCHANGE_RATE_HEADERS]
    finally:
        wb.close()


def _write_exchange_rate_rows(wb, rows: list[list[object]]) -> None:
    ws = wb.create_sheet(EXCHANGE_RATE_SHEET)
    for row in rows or [EXCHANGE_RATE_HEADERS]:
        ws.append(row)
    _autosize(ws, max_col=max(len(row) for row in rows or [EXCHANGE_RATE_HEADERS]))


def focus_reimbursement_workbook(path: Path, target_day: date | None = None) -> int:
    if not path.exists():
        return 0
    wb = load_workbook(path)
    try:
        target = target_day or date.today()
        ws = _ensure_manual_sheet(wb)
        row_idx = _focus_row(ws, target)
        wb.active = wb.sheetnames.index(ws.title)
        cell_ref = f"A{row_idx}"
        if ws.sheet_view.selection:
            ws.sheet_view.selection[0].activeCell = cell_ref
            ws.sheet_view.selection[0].sqref = cell_ref
        else:
            ws.sheet_view.selection = [Selection(activeCell=cell_ref, sqref=cell_ref)]
        ws.sheet_view.topLeftCell = f"A{max(row_idx - 3, 1)}"
        wb.save(path)
        return row_idx
    finally:
        wb.close()


def corrected_crop_names(workbook_path: Path) -> set[str]:
    if not workbook_path.exists():
        return set()
    wb = load_workbook(workbook_path, data_only=False)
    try:
        names: set[str] = set()
        ws = _ensure_manual_sheet(wb)
        columns = _header_columns(ws)
        link_col = columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))
        for row_idx in range(2, ws.max_row + 1):
            if not _row_is_protected(ws, row_idx):
                continue
            cell = ws.cell(row_idx, link_col)
            link = cell.hyperlink.target if cell.hyperlink else cell.value
            if link:
                names.add(Path(str(link)).name)
        return names
    finally:
        wb.close()


def _deleted_crop_names(ws) -> set[str]:
    return _marked_crop_names(ws, _row_is_deleted)


def _protected_crop_names(ws) -> set[str]:
    return _marked_crop_names(ws, _row_is_protected)


def _marked_crop_names(ws, predicate) -> set[str]:
    columns = _header_columns(ws)
    link_col = columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))
    names: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        if not predicate(ws, row_idx):
            continue
        cell = ws.cell(row_idx, link_col)
        link = str(cell.hyperlink.target if cell.hyperlink else cell.value or "").strip()
        if link:
            names.add(Path(link).name)
    return names


def _record_has_crop_name(record: InvoiceRecord, crop_names: set[str]) -> bool:
    if not crop_names:
        return False
    for crop_image in [record.crop_image, *list(getattr(record, "supporting_crop_images", []) or [])]:
        name = Path(str(crop_image or "")).name
        if name in crop_names:
            return True
    return False


def _record_has_any_crop(record: InvoiceRecord) -> bool:
    return any(str(crop_image or "").strip() for crop_image in [record.crop_image, *list(getattr(record, "supporting_crop_images", []) or [])])


def _record_is_protected_match(
    record: InvoiceRecord,
    rates: list[ExchangeRate],
    locked_keys: set[tuple[str, str, float]],
    protected_crops: set[str],
) -> bool:
    if _record_has_any_crop(record):
        return _record_has_crop_name(record, protected_crops)
    return _record_match_key(record, rates) in locked_keys


class ReimbursementWorkbook:
    def __init__(self, workbook_path: Path, fetch_rates: FetchRates | None = None) -> None:
        self.workbook_path = workbook_path
        self.fetch_rates = fetch_rates or fetch_safe_exchange_rates
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)

    def locked_numbers(self) -> set[int]:
        if not self.workbook_path.exists():
            return set()
        wb = load_workbook(self.workbook_path, data_only=False)
        try:
            ws = _ensure_manual_sheet(wb)
            numbers: set[int] = set()
            for row_idx in range(2, ws.max_row + 1):
                if _row_is_protected(ws, row_idx):
                    number = _to_int(ws.cell(row_idx, 1).value)
                    if number:
                        numbers.add(number)
            return numbers
        finally:
            wb.close()

    def unlocked_records(self, records: list[InvoiceRecord]) -> list[InvoiceRecord]:
        if not self.workbook_path.exists():
            return list(records)
        wb = load_workbook(self.workbook_path, data_only=False)
        try:
            ws = _ensure_manual_sheet(wb)
            rates = _rates_from_workbook(wb)
            locked_keys = _protected_record_keys(ws, rates)
            protected_crops = _protected_crop_names(ws)
            return [record for record in records if not _record_is_protected_match(record, rates, locked_keys, protected_crops)]
        finally:
            wb.close()

    def write_records(self, records: list[InvoiceRecord]) -> ReimbursementWriteResult:
        wb = _load_or_create_workbook(self.workbook_path)
        ws = _ensure_manual_sheet(wb)
        rates, rates_updated, fx_error = self._ensure_exchange_rates(wb, records)
        crop_links = _crop_link_map(wb)
        locked_rows = _protected_rows(ws)
        locked_keys = _protected_record_keys(ws, rates)
        protected_crops = _protected_crop_names(ws)
        output_records = [record for record in records if not _record_is_protected_match(record, rates, locked_keys, protected_crops)]
        _clear_unlocked_invoice_rows(ws, locked_rows)

        write_row = 2
        rows_written = 0
        for record in output_records:
            while write_row in locked_rows:
                write_row += 1
            _write_reimbursement_row(ws, write_row, record, rates)
            _set_crop_link_mapping(crop_links, record)
            write_row += 1
            rows_written += 1

        _sync_no_to_crop_ids(ws, _header_columns(ws))
        _write_crop_links_sheet(wb, crop_links)
        _format_invoice_exp(ws)
        _autosize(ws, max_col=len(REIMBURSEMENT_HEADERS))
        wb.save(self.workbook_path)
        wb.close()
        return ReimbursementWriteResult(rows_written, self.workbook_path, rates_updated, fx_error)

    def update_exchange_rates(self, start_date: date, end_date: date) -> ExchangeRateUpdateResult:
        wb = _load_or_create_workbook(self.workbook_path)
        existing = _rates_from_workbook(wb)
        fetched: list[ExchangeRate] = []
        error = ""
        try:
            fetched = self.fetch_rates(start_date, end_date)
        except TypeError:
            try:
                fetched = self.fetch_rates()
            except Exception as exc:
                error = str(exc).strip() or exc.__class__.__name__
        except Exception as exc:
            error = str(exc).strip() or exc.__class__.__name__
        merged = _merge_rates(existing, fetched)
        _write_exchange_rate_sheet(wb, merged)
        wb.save(self.workbook_path)
        wb.close()
        return ExchangeRateUpdateResult(self.workbook_path, start_date, end_date, len(existing), len(fetched), len(merged), error)

    def _ensure_exchange_rates(self, wb, records: list[InvoiceRecord]) -> tuple[list[ExchangeRate], bool, str]:
        existing = _rates_from_workbook(wb)
        fx_currencies = {
            _normalize_currency(record.currency)
            for record in records
            if _normalize_currency(record.currency) != "MXN"
        }
        fx_dates = [_record_date(record) for record in records if _normalize_currency(record.currency) != "MXN"]
        fx_dates = [value for value in fx_dates if value is not None]
        if not fx_dates:
            _write_exchange_rate_sheet(wb, existing)
            return existing, False, ""
        needed_start = min(fx_dates)
        needed_end = max(max(fx_dates), date.today())
        ranges = _missing_rate_ranges(existing, needed_start, needed_end)
        if any(_currency_needs_refresh(existing, currency, needed_start, needed_end) for currency in fx_currencies):
            ranges.append((needed_start, needed_end))
        ranges = _dedupe_ranges(ranges)
        fetched: list[ExchangeRate] = []
        errors: list[str] = []
        for start, end in ranges:
            try:
                fetched.extend(self.fetch_rates(start, end))
            except TypeError:
                try:
                    fetched.extend(self.fetch_rates())
                except Exception as exc:
                    errors.append(str(exc).strip() or exc.__class__.__name__)
            except Exception as exc:
                errors.append(str(exc).strip() or exc.__class__.__name__)
        merged = _merge_rates(existing, fetched)
        _write_exchange_rate_sheet(wb, merged)
        return merged, bool(fetched), "; ".join(errors)


def assign_available_line_numbers(records: list[InvoiceRecord], locked_numbers: Iterable[int]) -> None:
    used = {number for number in locked_numbers if number > 0}
    next_number = 1
    for record in records:
        while next_number in used:
            next_number += 1
        record.line_no = next_number
        used.add(next_number)
        next_number += 1


def clear_generated_crops(final_dir: Path, preserve_names: set[str] | None = None) -> None:
    preserve_names = preserve_names or set()
    if not final_dir.exists():
        final_dir.mkdir(parents=True, exist_ok=True)
        return
    for path in final_dir.rglob("*"):
        if path.is_dir():
            continue
        if path.name in preserve_names:
            continue
        path.unlink()
    for path in sorted((item for item in final_dir.rglob("*") if item.is_dir()), key=lambda item: len(item.parts), reverse=True):
        try:
            path.rmdir()
        except OSError:
            pass


def _migrate_final_crops_to_review(output_dir: Path) -> None:
    final_dir = output_dir / FINAL_CROPS_DIR
    review_dir = output_dir / REVIEW_CROPS_DIR
    if not final_dir.exists():
        review_dir.mkdir(parents=True, exist_ok=True)
        return
    review_dir.mkdir(parents=True, exist_ok=True)
    for path in final_dir.rglob("*"):
        if not path.is_file():
            continue
        target = review_dir / path.name
        if not target.exists():
            shutil.copy2(path, target)


def _resolve_crop_source(output_dir: Path, link_value: object) -> Path | None:
    text = str(link_value or "").strip()
    if not text:
        return None
    candidates: list[Path] = []
    path = Path(text)
    if path.is_absolute():
        candidates.append(path)
    else:
        candidates.append(output_dir / path)
        candidates.append(output_dir / CROPS_DIR / path.name)
        candidates.append(output_dir / REVIEW_CROPS_DIR / path.name)
        candidates.append(output_dir / FINAL_CROPS_DIR / path.name)
        candidates.append(output_dir / FINAL_CROPS_DIR / "food" / path.name)
        candidates.append(output_dir / FINAL_CROPS_DIR / "other" / path.name)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _source_names_inside_dir(crop_sources: list[list[Path]], directory: Path) -> set[str]:
    names: set[str] = set()
    try:
        resolved_dir = directory.resolve()
    except OSError:
        resolved_dir = directory
    for sources in crop_sources:
        for source in sources:
            try:
                source.resolve().relative_to(resolved_dir)
            except (OSError, ValueError):
                continue
            names.add(source.name)
    return names


def _relative_crop_link(output_dir: Path, source: Path) -> str:
    try:
        return source.resolve().relative_to(output_dir.resolve()).as_posix()
    except (OSError, ValueError):
        folder = source.parent.name or CROPS_DIR
        return f"{folder}/{source.name}"


def _crop_link_for_record(crop_image: str) -> str:
    path = Path(crop_image)
    folder = path.parent.name or CROPS_DIR
    return f"{folder}/{path.name}"


def _rewrite_crop_links_to_review(ws, output_dir: Path, columns: dict[str, int]) -> bool:
    link_col = columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))
    changed = False
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, link_col)
        value = str(cell.hyperlink.target if cell.hyperlink else cell.value or "").strip()
        if not value:
            continue
        path = Path(value)
        source = _resolve_crop_source(output_dir, value)
        if source is None:
            continue
        crop_path = output_dir / CROPS_DIR / source.name
        if crop_path.exists():
            target_value = f"{CROPS_DIR}/{source.name}"
        elif source.exists():
            target_value = _relative_crop_link(output_dir, source)
        else:
            continue
        if value == target_value:
            continue
        cell.value = target_value
        cell.hyperlink = target_value
        cell.style = "Hyperlink"
        changed = True
    return changed


def _reconcile_manual_links_to_processing_state(output_dir: Path, ws, columns: dict[str, int]) -> bool:
    state_path = output_dir / "processing_state.json"
    if not state_path.exists():
        return False
    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    records = [item for item in data.get("records", []) if isinstance(item, dict)]
    by_key: dict[tuple[str, str, str, float], list[dict]] = {}
    for record in records:
        crop_text = str(record.get("crop_image") or "").strip()
        if not crop_text:
            continue
        crop_path = Path(crop_text)
        if not crop_path.exists():
            continue
        key = _processing_record_key(record)
        by_key.setdefault(key, []).append(record)
    if not by_key:
        return False
    used: set[str] = set()
    changed = False
    link_col = columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))
    for row_idx in range(2, ws.max_row + 1):
        values = _row_values_by_header(ws, row_idx, columns)
        if not _looks_like_reimbursement_row(values):
            continue
        key = _row_crop_match_key(values)
        candidates = by_key.get(key, [])
        if not candidates:
            continue
        selected = None
        for candidate in candidates:
            name = Path(str(candidate.get("crop_image") or "")).name
            if name not in used:
                selected = candidate
                break
        if selected is None:
            selected = candidates[0]
        source = Path(str(selected.get("crop_image") or ""))
        if not source.exists():
            continue
        used.add(source.name)
        value = _relative_crop_link(output_dir, source)
        cell = ws.cell(row_idx, link_col)
        current_name = Path(str(cell.hyperlink.target if cell.hyperlink else cell.value or "")).name
        if current_name == source.name:
            continue
        cell.value = value
        cell.hyperlink = value
        cell.style = "Hyperlink"
        changed = True
    return changed


def _find_row_by_crop_id(ws, columns: dict[str, int], crop_id: str) -> int | None:
    link_col = columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))
    matches: list[int] = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, link_col)
        link = str(cell.hyperlink.target if cell.hyperlink else cell.value or "")
        name = Path(link).name
        if name.startswith(f"{crop_id}_"):
            matches.append(row_idx)
    if not matches:
        return None
    for row_idx in matches:
        if not _row_is_deleted(ws, row_idx):
            return row_idx
    return matches[0]


def _normalize_group_crop_ids(crop_ids: list[str] | tuple[str, ...]) -> list[str]:
    normalized: list[str] = []
    for crop_id in crop_ids:
        value = _normalize_crop_id(str(crop_id or ""))
        if not value:
            continue
        if value not in normalized:
            normalized.append(value)
    if len(normalized) < 2:
        raise ValueError("Group needs at least two crop IDs")
    return sorted(normalized, key=lambda value: int(value))


def _group_rows_for_crop_ids(ws, columns: dict[str, int], crop_ids: list[str], output_dir: Path) -> list[tuple[str, int, InvoiceRecord, Path | None]]:
    link_col = columns.get("Invoice link", len(REIMBURSEMENT_HEADERS))
    rows: list[tuple[str, int, InvoiceRecord, Path | None]] = []
    for crop_id in crop_ids:
        matches: list[int] = []
        for row_idx in range(2, ws.max_row + 1):
            link_cell = ws.cell(row_idx, link_col)
            link = str(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value or "")
            if Path(link).name.startswith(f"{crop_id}_"):
                matches.append(row_idx)
        if not matches:
            raise LookupError(f"Cannot find crop {crop_id}")
        active = [row_idx for row_idx in matches if not _row_is_deleted(ws, row_idx)]
        if not active:
            raise ValueError(f"Crop {crop_id} is already deleted")
        row_idx = active[0]
        link_cell = ws.cell(row_idx, link_col)
        link = str(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value or "")
        values = _row_values_by_header(ws, row_idx, columns)
        record = _record_from_reimbursement_row(values)
        record.line_no = int(crop_id)
        record.crop_image = link
        rows.append((crop_id, row_idx, record, _resolve_crop_source(output_dir, link)))
    return rows


def _merged_group_record(records: list[InvoiceRecord], crop_ids: list[str]) -> tuple[InvoiceRecord, list[str]]:
    warnings: list[str] = []
    primary_id = min(crop_ids, key=lambda value: int(value))
    primary = _record_for_trace(records, primary_id) or records[0]
    paired = pair_invoice_payment_slips([copy(record) for record in records])
    if len(paired) == 1:
        merged = paired[0]
    else:
        warnings.append("No confident payment pairing; grouped by user confirmation")
        amount_record = max(records, key=lambda record: float(record.total_amount or 0))
        merged = copy(primary)
        if _normalize_currency(amount_record.currency) == _normalize_currency(primary.currency):
            delta = round(float(amount_record.total_amount or 0) - float(primary.total_amount or 0), 2)
            if delta > 0.50:
                merged.tips = max(float(primary.tips or 0), delta)
            merged.total_amount = max(float(primary.total_amount or 0), float(amount_record.total_amount or 0))
            merged.expense_amount = max(merged.total_amount - float(merged.vat_amount or 0) - float(merged.sales_tax or 0), 0.0)
        if (not merged.seller or merged.seller == "Unknown") and amount_record.seller:
            merged.seller = amount_record.seller

    dates = {normalize_date(record.invoice_date) or str(record.invoice_date or "") for record in records if str(record.invoice_date or "").strip()}
    currencies = {_normalize_currency(record.currency) for record in records if str(record.currency or "").strip()}
    if len(dates) > 1:
        warnings.append("Grouped crops have different dates")
    if len(currencies) > 1:
        warnings.append("Grouped crops have different currencies")
    if not merged.invoice_date:
        merged.invoice_date = primary.invoice_date
    if not merged.currency:
        merged.currency = primary.currency or "MXN"
    merged.crop_image = primary.crop_image
    merged.supporting_crop_images = [record.crop_image for record in records if Path(str(record.crop_image or "")).name != Path(str(primary.crop_image or "")).name]
    details = [str(record.contents or "").strip() for record in records if str(record.contents or "").strip()]
    if details:
        merged.contents = " | ".join(dict.fromkeys(details))
    remarks = [str(merged.remarks or "").strip(), f"Manual group {' + '.join(crop_ids)}"]
    if warnings:
        remarks.append("; ".join(warnings))
    merged.remarks = "; ".join(item for item in remarks if item)
    return merged, warnings


def _record_for_trace(records: list[InvoiceRecord], trace_id: str) -> InvoiceRecord | None:
    for record in records:
        if (_trace_id_from_link(record.crop_image) or "") == trace_id:
            return record
    return None


def _grouped_crop_name(primary_id: str, record: InvoiceRecord, category: str) -> str:
    invoice_date = _safe_filename_part(record.invoice_date or "unknown-date")
    currency = _safe_filename_part(_normalize_currency(record.currency or "MXN"))
    amount = f"{float(record.total_amount or 0):.2f}"
    seller = _safe_filename_part(record.seller or category or "Unknown")[:80]
    return f"{primary_id}_{invoice_date}_{currency}_{amount}_{seller}.jpg"


def _remove_group_source_files(sources: list[Path], keep: Path) -> None:
    try:
        keep_resolved = keep.resolve()
    except OSError:
        keep_resolved = keep
    for source in sources:
        try:
            if source.resolve() == keep_resolved:
                continue
            if source.exists():
                source.unlink()
        except OSError:
            continue


def _sync_no_to_crop_ids(ws, columns: dict[str, int]) -> bool:
    no_col = columns.get("No.")
    link_col = columns.get("Invoice link")
    if not no_col or not link_col:
        return False
    trace_col = columns.get(TRACE_ID_HEADER)
    changed = False
    for row_idx in range(2, ws.max_row + 1):
        link_cell = ws.cell(row_idx, link_col)
        link = str(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value or "")
        trace_id = _trace_id_from_link(link)
        crop_no = _crop_no_from_link(link)
        if crop_no is None:
            continue
        no_cell = ws.cell(row_idx, no_col)
        if _to_int(no_cell.value) != crop_no:
            no_cell.value = crop_no
            changed = True
        if trace_col and trace_id and str(ws.cell(row_idx, trace_col).value or "").zfill(3) != trace_id:
            ws.cell(row_idx, trace_col).value = trace_id
            changed = True
    return changed


def _row_trace_id(ws, row_idx: int, columns: dict[str, int]) -> str:
    trace_col = columns.get(TRACE_ID_HEADER)
    if trace_col:
        trace = str(ws.cell(row_idx, trace_col).value or "").strip()
        if trace:
            return _normalize_crop_id(trace)
    link_col = columns.get("Invoice link")
    if link_col:
        link_cell = ws.cell(row_idx, link_col)
        link = str(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value or "")
        trace = _trace_id_from_link(link)
        if trace:
            return _normalize_crop_id(trace)
    no_col = columns.get("No.")
    if no_col:
        number = _to_int(ws.cell(row_idx, no_col).value)
        if number:
            return f"{number:03d}"
    return ""


def _crop_no_from_link(value: object) -> int | None:
    trace_id = _trace_id_from_link(value)
    if not trace_id:
        return None
    try:
        return int(trace_id)
    except ValueError:
        return None


def _trace_id_from_link(value: object) -> str | None:
    name = Path(str(value or "")).name
    trace_match = re.search(r"(?:^|_)trace(\d{3,})(?:_|\b)", name, re.I)
    if trace_match:
        return trace_match.group(1)
    match = re.match(r"(\d{3,})[a-z]?_", name)
    if not match:
        return None
    return match.group(1)


def _apply_manual_change(
    ws,
    row_idx: int,
    columns: dict[str, int],
    rates: list[ExchangeRate],
    *,
    invoice_date: str | None,
    category: str | None,
    amount: float | None,
    currency: str | None,
    comment: str | None,
    status: str,
) -> None:
    if invoice_date:
        parsed_date = _parse_date(invoice_date)
        if parsed_date is None:
            raise ValueError(f"Invalid date: {invoice_date}")
        ws.cell(row_idx, columns["Date"]).value = parsed_date
    if category:
        category_en = _manual_category(category)
        ws.cell(row_idx, columns["Accounting Category"]).value = category_en
        ws.cell(row_idx, columns["Type"]).value = _type_label_zh(category_en)
    if comment is not None:
        ws.cell(row_idx, columns["Detail"]).value = _merge_detail_comment(ws.cell(row_idx, columns["Detail"]).value, comment)
    current_currency = _normalize_currency(str(ws.cell(row_idx, columns.get("\u539f\u5e01\u79cd", 6)).value or "MXN"))
    current_original_amount = _to_float(ws.cell(row_idx, columns.get("\u539f\u91d1\u989d", 7)).value)
    current_mxn_amount = _to_float(ws.cell(row_idx, columns.get("MXN Amount", 4)).value)
    new_currency = _normalize_currency(currency or (current_currency if current_original_amount > 0 else "MXN"))
    new_amount = float(amount) if amount is not None else (current_original_amount if new_currency != "MXN" and current_original_amount > 0 else current_mxn_amount)
    rate_date = _parse_date(ws.cell(row_idx, columns.get("Date", 3)).value) or date.today()
    rate = _best_rate_for_date(rates, rate_date)
    mxn_amount, fx_multiplier = _mxn_amount(new_amount, new_currency, rate)
    if new_currency == "MXN":
        ws.cell(row_idx, columns["MXN Amount"]).value = round(new_amount, 2)
        ws.cell(row_idx, columns["\u539f\u5e01\u79cd"]).value = ""
        ws.cell(row_idx, columns["\u539f\u91d1\u989d"]).value = ""
        ws.cell(row_idx, columns["\u6c47\u7387"]).value = ""
    else:
        ws.cell(row_idx, columns["MXN Amount"]).value = mxn_amount
        ws.cell(row_idx, columns["\u539f\u5e01\u79cd"]).value = new_currency
        ws.cell(row_idx, columns["\u539f\u91d1\u989d"]).value = round(new_amount, 2)
        ws.cell(row_idx, columns["\u6c47\u7387"]).value = fx_multiplier
    ws.cell(row_idx, columns[MANUAL_STATUS_HEADER]).value = _normalize_manual_status(status)
    _format_row(ws, row_idx, columns)


def _merge_detail_comment(existing: object, comment: str) -> str:
    text = str(existing or "").strip()
    clean = str(comment or "").strip()
    if not clean:
        return text
    if not text:
        return clean
    if clean.casefold() in text.casefold():
        return text
    return f"{text} | {clean}"


def _manual_category(value: str) -> str:
    text = str(value or "").strip()
    for category in EXPENSE_CATEGORIES:
        if category.casefold() == text.casefold():
            return category
    category = normalize_expense_category(text)
    if category != "Other" or text.casefold() in {"other", "\u5176\u4ed6"}:
        return category
    raise ValueError(f"Invalid category: {value}")


def _row_snapshot(ws, row_idx: int, columns: dict[str, int]) -> dict[str, object]:
    return {header: ws.cell(row_idx, col).value for header, col in columns.items() if col <= ws.max_column}


def _normalize_crop_id(value: str) -> str:
    text = str(value or "").strip()
    match = re.search(r"\d+", text)
    if not match:
        raise ValueError("Missing crop id")
    return match.group(0).zfill(3)


def _normalize_manual_status(value: str) -> str:
    text = str(value or "ok").strip().casefold()
    if text in {"delete", "deleted", "\u5220\u9664", "\u5220\u6389"}:
        return "delete"
    if text in {"correct", "corrected"}:
        return "correct"
    return "ok"


def _row_crop_match_key(values: list[object]) -> tuple[str, str, str, float]:
    date_text = _date_to_text(values[2])
    merchant = re.sub(r"\s+", " ", str(values[8] or "Unknown")).strip().casefold()
    currency = _normalize_currency(str(values[5] or "MXN"))
    original_amount = _to_float(values[6])
    mxn_amount = _to_float(values[3])
    amount = original_amount if currency != "MXN" and original_amount > 0 else mxn_amount
    return (date_text, merchant, currency, round(amount, 2))


def _processing_record_key(record: dict) -> tuple[str, str, str, float]:
    date_text = _date_to_text(record.get("invoice_date"))
    merchant = re.sub(r"\s+", " ", str(record.get("seller") or "Unknown")).strip().casefold()
    currency = _normalize_currency(str(record.get("currency") or "MXN"))
    amount = _to_float(record.get("total_amount"))
    return (date_text, merchant, currency, round(amount, 2))


def _checked_crop_name(line_no: int, values: list[object], suffix: str = "", source: Path | None = None) -> str:
    invoice_date = _safe_filename_part(_date_to_text(_row_value(values, "Date")) or "unknown-date")
    original_currency = _normalize_currency(str(_row_value(values, "\u539f\u5e01\u79cd") or "MXN"))
    amount = _to_float(_row_value(values, "\u539f\u91d1\u989d")) if original_currency != "MXN" else _to_float(_row_value(values, "MXN Amount"))
    seller = _safe_filename_part(str(_row_value(values, "Merchant") or "Unknown"))[:80]
    trace_id = _trace_id_from_link(source or _row_value(values, TRACE_ID_HEADER) or _row_value(values, "Invoice link")) or "000"
    return f"{line_no:03d}{suffix}_trace{trace_id}_{invoice_date}_{original_currency}_{amount:.2f}_{seller}.jpg"


def _final_crop_subdir(sheet_name: str) -> str:
    return "food" if sheet_name == FOOD_EXP_SHEET else "other"


def _load_final_crops_manifest(path: Path) -> dict[str, dict[str, object]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    records = data.get("records") if isinstance(data, dict) else None
    if not isinstance(records, dict):
        return {}
    return {str(key): value for key, value in records.items() if isinstance(value, dict)}


def _save_final_crops_manifest(path: Path, records: dict[str, dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    data = {"version": 1, "records": records}
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    temp.replace(path)


def _final_crop_manifest_entry(
    output_dir: Path,
    values: list[object],
    sheet_name: str,
    sources: list[Path],
    target: Path,
) -> dict[str, object]:
    source_entries = []
    for source in sources:
        try:
            stat = source.stat()
            size = stat.st_size
            mtime_ns = stat.st_mtime_ns
        except OSError:
            size = 0
            mtime_ns = 0
        source_entries.append({"path": _manifest_relative_path(output_dir, source), "size": size, "mtime_ns": mtime_ns})
    trace_id = _trace_id_from_link(sources[0] if sources else None) or str(_row_value(values, TRACE_ID_HEADER) or "").zfill(3)
    payload = {
        "trace_id": trace_id,
        "sheet": sheet_name,
        "date": _date_to_text(_row_value(values, "Date")),
        "mxn_amount": round(_to_float(_row_value(values, "MXN Amount")), 2),
        "original_currency": _normalize_currency(str(_row_value(values, "\u539f\u5e01\u79cd") or "MXN")),
        "original_amount": round(_to_float(_row_value(values, "\u539f\u91d1\u989d")), 2),
        "merchant": str(_row_value(values, "Merchant") or ""),
        "category": str(_row_value(values, "Accounting Category") or ""),
        "sources": source_entries,
    }
    return {
        "path": _manifest_relative_path(output_dir, target),
        "signature": _manifest_signature(payload),
        "payload": payload,
    }


def _manifest_signature(payload: dict[str, object]) -> str:
    return json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _ensure_final_crop(
    manifest: dict[str, dict[str, object]],
    trace_id: str,
    entry: dict[str, object],
    sources: list[Path],
    target: Path,
) -> bool:
    old_entry = manifest.get(trace_id, {})
    old_signature = str(old_entry.get("signature") or "")
    old_relative = str(old_entry.get("path") or "")
    new_signature = str(entry.get("signature") or "")
    if old_signature == new_signature:
        old_path = target.parent.parent.parent / old_relative if old_relative else None
        if target.exists():
            return True
        if old_path is not None and old_path.exists():
            target.parent.mkdir(parents=True, exist_ok=True)
            if old_path.resolve() != target.resolve():
                if target.exists():
                    target.unlink()
                shutil.move(str(old_path), str(target))
            return True
    if old_relative:
        old_path = target.parent.parent.parent / old_relative
        if old_path.exists() and old_path.resolve() != target.resolve():
            old_path.unlink()
    return _write_final_crop(sources, target)


def _prune_stale_final_crops(
    output_dir: Path,
    old_manifest: dict[str, dict[str, object]],
    new_manifest: dict[str, dict[str, object]],
    current_paths: set[str],
) -> None:
    stale_paths = {
        str(entry.get("path") or "")
        for trace_id, entry in old_manifest.items()
        if trace_id not in new_manifest or str(entry.get("path") or "") not in current_paths
    }
    for relative in stale_paths:
        if not relative or relative in current_paths:
            continue
        path = output_dir / relative
        if path.exists():
            path.unlink()
    final_dir = output_dir / FINAL_CROPS_DIR
    if final_dir.exists():
        for path in sorted((item for item in final_dir.rglob("*") if item.is_dir()), key=lambda item: len(item.parts), reverse=True):
            try:
                path.rmdir()
            except OSError:
                pass


def _manifest_relative_path(output_dir: Path, path: Path) -> str:
    try:
        return path.resolve().relative_to(output_dir.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def _write_final_crop(sources: list[Path], target: Path) -> bool:
    existing = [source for source in sources if source.exists()]
    if not existing:
        return False
    if len(existing) == 1:
        if existing[0].resolve() != target.resolve():
            shutil.copy2(existing[0], target)
        return True
    try:
        _write_combined_crop(existing, target)
    except Exception:
        if existing[0].resolve() != target.resolve():
            shutil.copy2(existing[0], target)
    return True


def _write_combined_crop(sources: list[Path], target: Path) -> None:
    from PIL import Image, ImageOps

    images = [ImageOps.exif_transpose(Image.open(source)).convert("RGB") for source in sources]
    try:
        gap = 24
        if all(image.height >= image.width for image in images):
            width = sum(image.width for image in images) + gap * (len(images) - 1)
            height = max(image.height for image in images)
            canvas = Image.new("RGB", (width, height), "white")
            x = 0
            for image in images:
                y = (height - image.height) // 2
                canvas.paste(image, (x, y))
                x += image.width + gap
        else:
            width = max(image.width for image in images)
            height = sum(image.height for image in images) + gap * (len(images) - 1)
            canvas = Image.new("RGB", (width, height), "white")
            y = 0
            for image in images:
                x = (width - image.width) // 2
                canvas.paste(image, (x, y))
                y += image.height + gap
        canvas.save(target, "JPEG", quality=95, optimize=True)
    finally:
        for image in images:
            image.close()


def _crop_link_map(wb) -> dict[str, list[str]]:
    if CROP_LINKS_SHEET not in wb.sheetnames:
        return {}
    ws = wb[CROP_LINKS_SHEET]
    mapping: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        primary = str(row[0] or "").strip()
        supporting = str(row[1] or "").strip()
        if not primary or not supporting:
            continue
        mapping[Path(primary).name] = [part for part in supporting.split("|") if part]
    return mapping


def _set_crop_link_mapping(mapping: dict[str, list[str]], record: InvoiceRecord) -> None:
    primary = str(record.crop_image or "").strip()
    if not primary:
        return
    supports = []
    for value in getattr(record, "supporting_crop_images", []) or []:
        text = str(value or "").strip()
        if text and Path(text).name != Path(primary).name and text not in supports:
            supports.append(text)
    if supports:
        mapping[Path(primary).name] = supports
    else:
        mapping.pop(Path(primary).name, None)


def _write_crop_links_sheet(wb, mapping: dict[str, list[str]]) -> None:
    if CROP_LINKS_SHEET in wb.sheetnames:
        del wb[CROP_LINKS_SHEET]
    if not mapping:
        return
    ws = wb.create_sheet(CROP_LINKS_SHEET)
    ws.sheet_state = "hidden"
    ws.append(["Primary crop", "Supporting crops"])
    for primary, supports in sorted(mapping.items()):
        if supports:
            ws.append([primary, "|".join(supports)])


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
    return cleaned or "unknown"


def _load_or_create_workbook(path: Path):
    if path.exists():
        wb = load_workbook(path)
        _ensure_manual_sheet(wb)
        if EXCHANGE_RATE_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(EXCHANGE_RATE_SHEET)
            ws.append(EXCHANGE_RATE_HEADERS)
        return wb
    wb = Workbook()
    ws = wb.active
    ws.title = INVOICE_EXP_SHEET
    ws.append(REIMBURSEMENT_HEADERS)
    rates_ws = wb.create_sheet(EXCHANGE_RATE_SHEET)
    rates_ws.append(EXCHANGE_RATE_HEADERS)
    _format_invoice_exp(ws)
    return wb


def _merge_old_split_sheets_to_manual(wb, manual_ws) -> None:
    split_sheets = [wb[name] for name in (FOOD_EXP_SHEET, OTHER_EXP_SHEET) if name in wb.sheetnames and name != INVOICE_EXP_SHEET]
    if not split_sheets:
        return
    manual_columns = _header_columns(manual_ws)
    existing = {
        _row_identity(_row_values_by_header(manual_ws, row_idx, manual_columns))
        for row_idx in range(2, manual_ws.max_row + 1)
        if _looks_like_reimbursement_row(_row_values_by_header(manual_ws, row_idx, manual_columns))
    }
    for source in split_sheets:
        columns = _header_columns(source)
        for row_idx in range(2, source.max_row + 1):
            values = _row_values_by_header(source, row_idx, columns)
            if not _looks_like_reimbursement_row(values):
                continue
            identity = _row_identity(values)
            if identity in existing:
                continue
            _append_values_with_link(manual_ws, values)
            existing.add(identity)
    for source in split_sheets:
        if len(wb.sheetnames) > 1:
            del wb[source.title]


def _row_identity(values: list[object]) -> tuple[str, str, str, float, str]:
    crop_name = Path(str(_row_value(values, "Invoice link") or "")).name
    date_text = _date_to_text(values[2])
    merchant = re.sub(r"\s+", " ", str(values[8] or "Unknown")).strip().casefold()
    original_currency = _normalize_currency(str(values[5] or "MXN"))
    original_amount = _to_float(values[6])
    mxn_amount = _to_float(values[3])
    amount = original_amount if original_currency != "MXN" and original_amount > 0 else mxn_amount
    return (crop_name, date_text, merchant, round(amount, 2), original_currency)


def _append_values_with_link(ws, values: list[object]) -> None:
    values = list(values[: len(REIMBURSEMENT_HEADERS)])
    values += [None] * (len(REIMBURSEMENT_HEADERS) - len(values))
    ws.append(values)
    _format_row(ws, ws.max_row, _header_columns(ws))
    link = str(_row_value(values, "Invoice link") or "").strip()
    if link:
        link_col = _header_columns(ws).get("Invoice link", len(REIMBURSEMENT_HEADERS))
        cell = ws.cell(ws.max_row, link_col)
        cell.hyperlink = link
        cell.style = "Hyperlink"


def _ensure_headers(ws, headers: list[str] | None = None) -> None:
    headers = headers or REIMBURSEMENT_HEADERS
    _reorder_columns_to_headers(ws, headers)
    for col, header in enumerate(headers, start=1):
        if ws.cell(1, col).value in (None, ""):
            ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True)


def _reorder_columns_to_headers(ws, headers: list[str]) -> None:
    current = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if [str(value or "") for value in current[: len(headers)]] == headers:
        return
    source_columns: dict[str, int] = {}
    for col, value in enumerate(current, start=1):
        normalized = _normalize_header(value)
        for logical, aliases in HEADER_ALIASES.items():
            if normalized in aliases and logical in headers:
                source_columns.setdefault(logical, col)
                break
    if not source_columns:
        return
    max_row = max(ws.max_row, 1)
    cells: list[list[tuple[object, object, object]]] = []
    for row_idx in range(1, max_row + 1):
        row_cells: list[tuple[object, object, object]] = []
        for header in headers:
            source_col = source_columns.get(header)
            if row_idx == 1:
                row_cells.append((header, None, None))
                continue
            if source_col is None:
                row_cells.append((None, None, None))
                continue
            source = ws.cell(row_idx, source_col)
            row_cells.append((source.value, source.hyperlink.target if source.hyperlink else None, copy(source._style)))
        cells.append(row_cells)
    for row_idx, row_cells in enumerate(cells, start=1):
        for col, (value, hyperlink, style) in enumerate(row_cells, start=1):
            target = ws.cell(row_idx, col)
            target.value = value
            target.hyperlink = hyperlink
            if style is not None:
                target._style = style
            if hyperlink:
                target.style = "Hyperlink"


def _protected_rows(ws) -> set[int]:
    return {row_idx for row_idx in range(2, ws.max_row + 1) if _row_is_protected(ws, row_idx)}


def _row_is_corrected(ws, row_idx: int) -> bool:
    return _manual_status_contains(ws, row_idx, CORRECTED_MARKERS)


def _row_is_deleted(ws, row_idx: int) -> bool:
    return _manual_status_contains(ws, row_idx, DELETED_MARKERS)


def _row_is_protected(ws, row_idx: int) -> bool:
    return _row_is_corrected(ws, row_idx) or _row_is_deleted(ws, row_idx)


def _manual_status_contains(ws, row_idx: int, markers: set[str]) -> bool:
    columns = _header_columns(ws)
    status_col = columns.get(MANUAL_STATUS_HEADER)
    if not status_col:
        return False
    value = str(ws.cell(row_idx, status_col).value or "").casefold()
    return any(marker in value for marker in markers)


def _clear_unlocked_invoice_rows(ws, locked_rows: set[int]) -> None:
    for row_idx in range(2, ws.max_row + 1):
        if row_idx in locked_rows:
            continue
        for col in range(1, max(ws.max_column, len(REIMBURSEMENT_HEADERS)) + 1):
            cell = ws.cell(row_idx, col)
            cell.value = None
            cell.hyperlink = None


def _write_reimbursement_row(ws, row_idx: int, record: InvoiceRecord, rates: list[ExchangeRate]) -> None:
    columns = _header_columns(ws)
    values = _reimbursement_value_map(record, rates)
    for header, value in values.items():
        col = columns.get(header)
        if not col:
            continue
        ws.cell(row_idx, col).value = value
    _format_row(ws, row_idx, columns)
    if record.crop_image:
        target = _crop_link_for_record(record.crop_image)
        link_cell = ws.cell(row_idx, columns.get("Invoice link", len(REIMBURSEMENT_HEADERS)))
        link_cell.value = target
        link_cell.hyperlink = target
        link_cell.style = "Hyperlink"


def _reimbursement_value_map(record: InvoiceRecord, rates: list[ExchangeRate]) -> dict[str, object]:
    original_currency = _normalize_currency(record.currency)
    original_amount = round(record.total_amount, 2)
    rate_date = _record_date(record) or date.today()
    rate = _best_rate_for_date(rates, rate_date)
    mxn_amount, fx_multiplier = _mxn_amount(original_amount, original_currency, rate)
    category_en = normalize_expense_category(record.expense_category, f"{record.seller} {record.contents}")
    return {
        "No.": _crop_no_from_link(record.crop_image) or record.line_no,
        MANUAL_STATUS_HEADER: "",
        "Invoice link": Path(record.crop_image).name if record.crop_image else "",
        TRACE_ID_HEADER: _trace_id_from_link(record.crop_image) or (f"{record.line_no:03d}" if record.line_no else ""),
        "Date": _record_date(record) or record.invoice_date,
        "MXN Amount": mxn_amount,
        "Type": _type_label_zh(category_en),
        "\u539f\u5e01\u79cd": "" if original_currency == "MXN" else original_currency,
        "\u539f\u91d1\u989d": "" if original_currency == "MXN" else original_amount,
        "\u6c47\u7387": "" if original_currency == "MXN" else fx_multiplier,
        "Merchant": record.seller,
        "Detail": record.contents,
        "Accounting Category": category_en,
        SYSTEM_NOTE_HEADER: record.remarks,
    }


def _mxn_amount(amount: float, currency: str, rate: ExchangeRate | None) -> tuple[float | None, float | None]:
    if currency == "MXN":
        return round(amount, 2), None
    if rate:
        multiplier = rate.multiplier_to_mxn(currency)
        if multiplier is None:
            return None, None
        return round(amount * multiplier, 2), multiplier
    return None, None


def _record_from_reimbursement_row(values: list[object]) -> InvoiceRecord:
    category = str(values[10] or values[4] or "Other")
    currency = str(values[5] or "MXN")
    original_amount = _to_float(values[6])
    mxn_amount = _to_float(values[3])
    record = InvoiceRecord(
        line_no=_to_int(values[0]),
        invoice_date=_date_to_text(values[2]),
        expense_category=normalize_expense_category(category),
        contents=str(values[9] or ""),
        currency=currency,
        total_amount=round(mxn_amount if currency == "MXN" or original_amount <= 0 else mxn_amount, 2),
        seller=str(values[8] or "Unknown"),
        remarks=str(_row_value(values, SYSTEM_NOTE_HEADER) or ""),
        crop_image=str(_row_value(values, "Invoice link") or ""),
    )
    record.original_currency = currency
    record.original_amount = round(original_amount if original_amount > 0 else mxn_amount, 2)
    record.mxn_amount = round(mxn_amount, 2)
    return record


def _looks_like_reimbursement_row(values: list[object]) -> bool:
    if not any(value not in (None, "") for value in values):
        return False
    return bool(values[2] or values[3] or values[8])


def _focus_row(ws, target_day: date) -> int:
    last_data_row = 2
    columns = _header_columns(ws)
    date_col = columns.get("Date", 3)
    for row_idx in range(2, ws.max_row + 1):
        values = _row_values_by_header(ws, row_idx, columns)
        if _looks_like_reimbursement_row(values):
            last_data_row = row_idx
        if _parse_date(ws.cell(row_idx, date_col).value) == target_day:
            return row_idx
    return last_data_row


def _rates_from_workbook(wb) -> list[ExchangeRate]:
    if EXCHANGE_RATE_SHEET not in wb.sheetnames:
        return []
    ws = wb[EXCHANGE_RATE_SHEET]
    headers = [_normalize_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    code_by_col: dict[int, str] = {}
    for col, header in enumerate(headers, start=1):
        code = _exchange_header_to_code(header)
        if code:
            code_by_col[col] = code
    rates: list[ExchangeRate] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rate_date = _parse_date(row[0] if row else None)
        values: dict[str, float] = {}
        for col, code in code_by_col.items():
            value = _to_float(row[col - 1] if len(row) >= col else None)
            if value > 0:
                values[code] = value
        usd = values.get("USD", 0.0)
        mxn = values.get("MXN", 0.0)
        if rate_date and usd > 0 and mxn > 0:
            rates.append(ExchangeRate(rate_date, usd, mxn, rates=values))
    return rates


def _best_rate_for_date(rates: list[ExchangeRate], target: date) -> ExchangeRate | None:
    rate = nearest_rate_on_or_before(rates, target)
    if rate is not None:
        return rate
    if not rates:
        return None
    return min(rates, key=lambda item: abs((item.rate_date - target).days))


def _write_exchange_rate_sheet(wb, rates: list[ExchangeRate]) -> None:
    ws = wb[EXCHANGE_RATE_SHEET] if EXCHANGE_RATE_SHEET in wb.sheetnames else wb.create_sheet(EXCHANGE_RATE_SHEET)
    for col, header in enumerate(EXCHANGE_RATE_HEADERS, start=1):
        ws.cell(1, col).value = header
        ws.cell(1, col).font = Font(bold=True)
    existing_rows = {(_parse_date(ws.cell(row, 1).value) or date.min): row for row in range(2, ws.max_row + 1)}
    for rate in sorted(rates, key=lambda item: item.rate_date):
        row_idx = existing_rows.get(rate.rate_date)
        if row_idx is None:
            row_idx = ws.max_row + 1
        ws.cell(row_idx, 1).value = rate.rate_date
        ws.cell(row_idx, 1).number_format = "yyyy-mm-dd"
        for offset, code in enumerate(SAFE_CURRENCY_CODES, start=2):
            value = rate.rate_value(code)
            if value > 0:
                ws.cell(row_idx, offset).value = value
                ws.cell(row_idx, offset).number_format = "0.00"
    _autosize(ws, max_col=len(EXCHANGE_RATE_HEADERS))


def _merge_rates(existing: list[ExchangeRate], fetched: list[ExchangeRate]) -> list[ExchangeRate]:
    by_date = {rate.rate_date: rate for rate in existing}
    by_date.update({rate.rate_date: rate for rate in fetched})
    return [by_date[key] for key in sorted(by_date)]


def _currency_needs_refresh(existing: list[ExchangeRate], currency: str, needed_start: date, needed_end: date) -> bool:
    code = normalize_safe_currency(currency)
    if code in {"MXN", "CNY", "RMB"}:
        return False
    if code not in SAFE_CURRENCY_CODES:
        return False
    relevant = [rate for rate in existing if needed_start <= rate.rate_date <= needed_end]
    return not relevant or any(rate.rate_value(code) <= 0 for rate in relevant)


def _missing_rate_ranges(existing: list[ExchangeRate], needed_start: date, needed_end: date) -> list[tuple[date, date]]:
    if needed_start > needed_end:
        return []
    if not existing:
        return [(needed_start, needed_end)]
    existing_dates = [rate.rate_date for rate in existing]
    ranges: list[tuple[date, date]] = []
    min_existing = min(existing_dates)
    max_existing = max(existing_dates)
    if needed_start < min_existing:
        ranges.append((needed_start, min_existing - timedelta(days=1)))
    if needed_end > max_existing:
        ranges.append((max_existing + timedelta(days=1), needed_end))
    return [(start, end) for start, end in ranges if start <= end]


def _dedupe_ranges(ranges: list[tuple[date, date]]) -> list[tuple[date, date]]:
    deduped: list[tuple[date, date]] = []
    seen: set[tuple[date, date]] = set()
    for item in ranges:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return deduped


def _protected_record_keys(ws, rates: list[ExchangeRate]) -> set[tuple[str, str, float]]:
    columns = _header_columns(ws)
    keys: set[tuple[str, str, float]] = set()
    for row_idx in _protected_rows(ws):
        if _row_is_deleted(ws, row_idx):
            continue
        values = _row_values_by_header(ws, row_idx, columns)
        if not _looks_like_reimbursement_row(values):
            continue
        keys.add(_row_match_key(values))
    return keys


def _record_match_key(record: InvoiceRecord, rates: list[ExchangeRate]) -> tuple[str, str, float]:
    values = _reimbursement_value_map(record, rates)
    return _row_match_key([values.get(header) for header in REIMBURSEMENT_HEADERS])


def _row_match_key(values: list[object]) -> tuple[str, str, float]:
    date_text = _date_to_text(values[2])
    merchant = re.sub(r"\s+", " ", str(values[8] or "Unknown")).strip().casefold()
    original_currency = _normalize_currency(str(values[5] or "MXN"))
    original_amount = _to_float(values[6])
    mxn_amount = _to_float(values[3])
    amount = original_amount if original_currency != "MXN" and original_amount > 0 else mxn_amount
    return (date_text, merchant, round(amount, 2))


def _format_invoice_exp(ws, headers: list[str] | None = None) -> None:
    _ensure_headers(ws, headers)
    columns = _header_columns(ws)
    for row in range(2, ws.max_row + 1):
        _format_row(ws, row, columns)


def _format_row(ws, row: int, columns: dict[str, int]) -> None:
    if columns.get("No."):
        ws.cell(row, columns["No."]).number_format = "000"
    if columns.get("Date"):
        ws.cell(row, columns["Date"]).number_format = "yyyy-mm-dd"
    if columns.get("MXN Amount"):
        ws.cell(row, columns["MXN Amount"]).number_format = "#,##0.00"
    if columns.get("\u539f\u91d1\u989d"):
        ws.cell(row, columns["\u539f\u91d1\u989d"]).number_format = "#,##0.00"
    if columns.get("\u6c47\u7387"):
        ws.cell(row, columns["\u6c47\u7387"]).number_format = "0.000000"


def _header_columns(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, max(ws.max_column, len(REIMBURSEMENT_HEADERS)) + 1):
        normalized = _normalize_header(ws.cell(1, col).value)
        for logical, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapping.setdefault(logical, col)
                break
    for index, header in enumerate(REIMBURSEMENT_HEADERS, start=1):
        mapping.setdefault(header, index)
    return mapping


def _row_values_by_header(ws, row_idx: int, columns: dict[str, int]) -> list[object]:
    values: list[object] = []
    for index, header in enumerate(REIMBURSEMENT_HEADERS, start=1):
        if header in columns:
            values.append(ws.cell(row_idx, columns[header]).value)
        elif header in {MANUAL_STATUS_HEADER, SYSTEM_NOTE_HEADER}:
            values.append(None)
        else:
            values.append(ws.cell(row_idx, index).value)
    return values


def _row_value(values: list[object], header: str) -> object:
    try:
        index = REIMBURSEMENT_HEADERS.index(header)
    except ValueError:
        return None
    return values[index] if index < len(values) else None


def _checked_value(values: list[object], header: str) -> object:
    try:
        index = CHECKED_HEADERS.index(header)
    except ValueError:
        return None
    return values[index] if index < len(values) else None


def _normalize_header(value: object) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip().casefold()
    return text


def _exchange_header_to_code(header: str) -> str:
    aliases = {
        "\u7f8e\u5143": "USD",
        "--> \u7f8e\u5143": "USD",
        "\u6bd4\u7d22": "MXN",
        "--> \u6bd4\u7d22": "MXN",
    }
    if header.upper() in SAFE_CURRENCY_CODES:
        return header.upper()
    return aliases.get(header, "")


def _autosize(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_length = 0
        for cell in ws[letter]:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 42)


def _record_date(record: InvoiceRecord) -> date | None:
    normalized = normalize_date(record.invoice_date) or record.invoice_date
    return _parse_date(normalized)


def _parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    normalized = normalize_date(text) or text[:10]
    try:
        return date.fromisoformat(normalized)
    except ValueError:
        return None


def _date_to_text(value: object) -> str:
    parsed = _parse_date(value)
    return parsed.isoformat() if parsed else str(value or "")


def _normalize_currency(value: str) -> str:
    return normalize_safe_currency(value)


def _type_label_zh(category: str) -> str:
    normalized = normalize_expense_category(category)
    return TYPE_LABELS_ZH.get(normalized, TYPE_LABELS_ZH["Other"])


def _to_float(value: object) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: object) -> int | None:
    try:
        return int(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None
