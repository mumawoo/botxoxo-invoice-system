from __future__ import annotations

import asyncio
import importlib.util
import json
import os
import re
import errno
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .config import Settings
from .expense_categories import normalize_expense_category
from .models import InvoiceRecord, PipelineSummary
from .queue_worker import (
    DONE,
    FAILED_RETRYABLE,
    QueueItem,
    discover_and_enqueue,
    enqueue_photo,
    format_status,
    load_queue_state,
    queue_totals_for_day,
    retry_failed,
    start_background_worker,
    summarize_queue,
    telegram_user_day_dir,
    telegram_user_output_dir,
    telegram_user_queue_path,
    telegram_user_workbook,
)
from .reimbursement import (
    format_reimbursement_summary,
    format_submit_result,
    refresh_checked_outputs,
    rerun_finance_edits,
    submit_unsubmitted,
    submitted_batches_text,
    unsubmitted_summary,
)
from .reimbursement_excel import INVOICE_EXP_SHEET, REVIEW_CROPS_DIR, checked_workbook_path, focus_reimbursement_workbook
from .reimbursement_excel import available_crop_ids, change_reimbursement_record

SUBMIT_CONFIRM_WORDS = {"confirm", "yes", "ok", "\u786e\u8ba4", "\u63d0\u4ea4"}
SUBMIT_CANCEL_WORDS = {"cancel", "no", "\u53d6\u6d88", "\u4e0d\u63d0\u4ea4"}
LANG_EN = "en"
LANG_ZH = "zh"
LANG_CHOICES = {LANG_EN, LANG_ZH, "cn", "chinese", "english", "中文", "英文"}
PREFERENCES_FILE = "telegram_preferences.json"


def normalize_telegram_language(value: str | None) -> str:
    normalized = str(value or "").strip().casefold()
    if normalized in {"zh", "cn", "chinese", "中文"}:
        return LANG_ZH
    return LANG_EN


def user_language(settings: Settings, user_id: int | None = None) -> str:
    default = normalize_telegram_language(settings.telegram_language)
    if user_id is None:
        return default
    path = _telegram_preferences_path(settings, user_id)
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return default
    return normalize_telegram_language(str(data.get("language") or default))


def set_user_language(settings: Settings, user_id: int, language: str) -> str:
    selected = normalize_telegram_language(language)
    path = _telegram_preferences_path(settings, user_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps({"language": selected}, ensure_ascii=False, indent=2), encoding="utf-8")
    return selected


def _telegram_preferences_path(settings: Settings, user_id: int) -> Path:
    return telegram_user_output_dir(settings, user_id) / PREFERENCES_FILE


def is_zh(lang: str | None) -> bool:
    return normalize_telegram_language(lang) == LANG_ZH


def is_allowed_user(user_id: int, allowed_ids: set[int] | frozenset[int]) -> bool:
    return user_id in allowed_ids


def whoami_message(user_id: int, username: str | None = None, lang: str = LANG_EN) -> str:
    suffix = f" (@{username})" if username else ""
    if is_zh(lang):
        return f"你的 Telegram user ID 是 {user_id}{suffix}。"
    return f"Your Telegram user ID is {user_id}{suffix}."


def saved_photo_message(path: Path, lang: str = LANG_EN) -> str:
    if is_zh(lang):
        return "照片已保存"
    return "Saved photo"


def queued_photo_message(path: Path, worker_started: bool, lang: str = LANG_EN) -> str:
    if is_zh(lang):
        worker = "已启动" if worker_started else "正在运行"
        return f"照片已保存\n扫描器：{worker}"
    worker = "started" if worker_started else "already running"
    return f"Saved photo\nScanner: {worker}"


def processing_success_message(summary: PipelineSummary, lang: str = LANG_EN) -> str:
    if is_zh(lang):
        return f"照片已保存，并已处理当前 Telegram 批次。来源：{summary.source_images}。行数：{summary.records_written}。Excel：{summary.workbook_path}"
    return f"Saved photo and processed current Telegram batch. Sources: {summary.source_images}. Rows: {summary.records_written}. Excel: {summary.workbook_path}"


def processing_failure_message(path: Path, exc: Exception, lang: str = LANG_EN) -> str:
    detail = str(exc).strip() or exc.__class__.__name__
    if is_zh(lang):
        return f"照片已保存\n处理失败：{detail}"
    return f"Saved photo\nProcessing failed: {detail}"


def scan_completion_message(settings: Settings, user_id: int, item: QueueItem, records: list[InvoiceRecord], lang: str = LANG_EN) -> str:
    if item.status == FAILED_RETRYABLE:
        if is_zh(lang):
            return "\n".join(["扫描失败", f"照片：{Path(item.path).name}", f"错误：{item.error or 'unknown error'}", "用 /restart 重试。"])
        return "\n".join(["Scan failed", f"Photo: {Path(item.path).name}", f"Error: {item.error or 'unknown error'}", "Use /restart."])
    if item.status != DONE:
        return ""

    today = queue_totals_for_day(settings, user_id)
    queue = summarize_queue(settings, user_id)
    original_totals = _original_currency_totals(records)
    if is_zh(lang):
        lines = [
            "扫描完成",
            "",
            "本次扫描",
            f"新增行数：{len(records)}",
            "Crops：",
        ]
    else:
        lines = [
            "Scan complete",
            "",
            "This scan",
            f"Rows: {len(records)}",
            "Crops:",
        ]
    lines.extend(_format_scan_record_lines(records))
    if original_totals:
        lines.append("原币种合计：" if is_zh(lang) else "Original totals:")
        lines.extend(_format_currency_lines(original_totals))
    if is_zh(lang):
        lines.extend(["", "今天合计", f"行数：{today.record_count}", f"MXN 总额：{today.total_amount:.2f}"])
        lines.extend(_format_category_block(today.category_totals, lang))
        lines.extend(["", "队列", f"完成/等待/失败：{queue.done}/{queue.pending}/{queue.failed}", "人工检查：/excel"])
    else:
        lines.extend(["", "Today", f"Rows: {today.record_count}", f"MXN total: {today.total_amount:.2f}"])
        lines.extend(_format_category_block(today.category_totals, lang))
        lines.extend(["", "Queue", f"Done/Pending/Failed: {queue.done}/{queue.pending}/{queue.failed}", "Review: /excel"])
    return "\n".join(lines)


def telegram_photo_filename(timestamp: datetime, file_id: str, file_unique_id: str | None = None) -> str:
    identifier = _safe_filename_part(file_id) or _safe_filename_part(file_unique_id or "") or "telegram_photo"
    return f"{timestamp.strftime('%H%M%S')}_{identifier}.jpg"


def resolve_auto_process(settings: Settings, override: bool | None = None) -> bool:
    return settings.telegram_auto_process if override is None else override


def telegram_batch_source(day_dir: Path) -> Path:
    return day_dir


def telegram_polling_ready(settings: Settings) -> bool:
    return bool(settings.telegram_bot_token and telegram_package_ready())


def telegram_config_ready(settings: Settings) -> bool:
    return telegram_polling_ready(settings) and bool(settings.telegram_allowed_user_ids)


def telegram_package_ready() -> bool:
    return importlib.util.find_spec("telegram") is not None and importlib.util.find_spec("telegram.ext") is not None


def telegram_start_message(settings: Settings, lang: str = LANG_EN) -> str:
    if settings.telegram_allowed_user_ids:
        if is_zh(lang):
            return "发票机器人已就绪。请发送票据照片。用 /help 查看命令。"
        return "Invoice bot is ready. Send receipt photos. Use /help for commands."
    if is_zh(lang):
        return "发票机器人处于设置模式。用 /whoami 获取你的 user ID；配置 TELEGRAM_ALLOWED_USER_IDS 前会拒收照片。"
    return "Invoice bot setup mode. Use /whoami to get your user ID; receipt photos are rejected until TELEGRAM_ALLOWED_USER_IDS is configured."


def telegram_help_message(lang: str = LANG_EN) -> str:
    title = "命令" if is_zh(lang) else "Commands"
    return "\n".join([title, *[f"/{command} - {description}" for command, description in telegram_command_menu(lang)]])


def telegram_command_menu(lang: str = LANG_EN) -> list[tuple[str, str]]:
    if is_zh(lang):
        return [
            ("lang", "切换语言：/lang zh 或 /lang en"),
            ("change", "修改 crop：/change 021 type Other + 备注"),
            ("del", "删除 crop：/del 021"),
            ("excel", "下载当前人工复核 Excel"),
            ("checked", "下载财务版 Food/Other Excel"),
            ("crops", "发送最近 review crop 图片"),
            ("recent", "最近 2 次上传和 Excel 行"),
            ("rerun", "从修改后的财务 Excel 重建"),
            ("report", "未提交报销汇总"),
            ("restart", "重试失败照片"),
            ("submit", "先预览，再回复 confirm/cancel"),
            ("status", "队列状态"),
            ("whoami", "显示你的 Telegram user ID"),
            ("help", "显示命令"),
        ]
    return [
        ("lang", "switch language: /lang zh or /lang en"),
        ("change", "edit crop: /change 021 type Other + note"),
        ("del", "delete crop: /del 021"),
        ("excel", "download current reimbursement Excel"),
        ("checked", "download finance Food/Other Excel"),
        ("crops", "send latest review crop images"),
        ("recent", "last 2 uploads and Excel rows"),
        ("rerun", "rebuild from edited finance Excel"),
        ("report", "unsubmitted reimbursement summary"),
        ("restart", "retry failed photos"),
        ("submit", "preview, then reply confirm/cancel"),
        ("status", "queue status"),
        ("whoami", "show your Telegram user ID"),
        ("help", "show commands"),
    ]


def status_message(settings: Settings, user_id: int, lang: str = LANG_EN) -> str:
    discover_and_enqueue(settings, user_id)
    summary = summarize_queue(settings, user_id)
    return append_process_status(_format_status(summary, lang), lang=lang)


def append_process_status(text: str, pid: int | None = None, lang: str = LANG_EN) -> str:
    process_id = os.getpid() if pid is None else pid
    label = "Telegram bot PID" if not is_zh(lang) else "Telegram bot 进程 PID"
    return "\n".join([text, f"{label}: {process_id}"])


def restart_message(settings: Settings, user_id: int, lang: str = LANG_EN) -> str:
    count, summary = retry_failed(settings, user_id)
    started = start_background_worker(settings, user_id)
    if is_zh(lang):
        return "\n".join(
            [
                "重试失败照片",
                f"已重新入队：{count}",
                f"扫描器：{'已启动' if started else '正在运行'}",
                _format_status(summary, lang),
            ]
        )
    return "\n".join(
        [
            "Restart",
            f"Failed photos retried: {count}",
            f"Scanner: {'started' if started else 'already running'}",
            _format_status(summary, lang),
        ]
    )


def report_message(settings: Settings, user_id: int, lang: str = LANG_EN) -> str:
    queue = summarize_queue(settings, user_id)
    summary = unsubmitted_summary(settings, user_id)
    summary = type(summary)(
        record_count=summary.record_count,
        photo_count=queue.done,
        total_amount=summary.total_amount,
        category_totals=summary.category_totals,
        codex_used=summary.codex_used,
        failed_count=queue.failed,
        date_min=summary.date_min,
        date_max=summary.date_max,
    )
    if is_zh(lang):
        return _format_reimbursement_summary(summary, title="报销汇总 / 未提交", lang=lang)
    return format_reimbursement_summary(summary)


def submit_message(settings: Settings, user_id: int, args: list[str], lang: str = LANG_EN) -> str:
    if args:
        if is_zh(lang):
            return "请先发送 /submit，然后直接回复 confirm 确认，或 cancel 取消。"
        return "Use /submit first, then reply with plain text: confirm or cancel."
    summary = unsubmitted_summary(settings, user_id)
    if summary.record_count <= 0:
        if is_zh(lang):
            return "提交\n没有未提交记录。"
        return "Submit / 提交\nNo unsubmitted records."
    if is_zh(lang):
        return "\n".join(
            [
                _format_reimbursement_summary(summary, title="提交预览", lang=lang),
                "",
                "确认后会归档当前财务 Excel 和 final_crops，然后从 001 开始新批次。",
                "回复 confirm 确认提交，或 cancel 取消。",
            ]
        )
    lines = [
        format_reimbursement_summary(summary, title="Submit preview / 提交预览"),
        "",
        "This will archive the current finance Excel and final_crops, then start a new batch from 001.",
        "Reply confirm to submit, or cancel to stop.",
    ]
    return "\n".join(lines)


def submit_pending_confirmation(settings: Settings, user_id: int) -> bool:
    return unsubmitted_summary(settings, user_id).record_count > 0


def submit_confirmation_message(settings: Settings, user_id: int, text: str, lang: str = LANG_EN) -> tuple[bool, str]:
    normalized = str(text or "").strip().casefold()
    if normalized in SUBMIT_CONFIRM_WORDS:
        result = submit_unsubmitted(settings, user_id)
        if is_zh(lang):
            return True, _format_submit_result(result, lang)
        return True, format_submit_result(result)
    if normalized in SUBMIT_CANCEL_WORDS:
        if is_zh(lang):
            return True, "提交已取消。没有归档任何内容。"
        return True, "Submit cancelled. Nothing was archived."
    if is_zh(lang):
        return False, "提交等待确认。回复 confirm 提交，或 cancel 取消。"
    return False, "Submit pending. Reply confirm to submit, or cancel to stop."


def recent_message(settings: Settings, user_id: int, limit: int = 2, lang: str = LANG_EN) -> str:
    output_dir = telegram_user_output_dir(settings, user_id)
    state = load_queue_state(telegram_user_queue_path(settings, user_id))
    done_items = [item for item in state.items if item.status == DONE]
    done_items.sort(key=lambda item: (item.received_at or item.updated_at or "", item.updated_at or ""), reverse=True)
    if not done_items:
        if is_zh(lang):
            return "还没有完成的上传。"
        return "No completed uploads yet."

    excel_rows = _excel_rows_by_crop_id(telegram_user_workbook(settings, user_id))
    lines = [f"最近上传 {min(limit, len(done_items))} 次" if is_zh(lang) else f"Recent uploads / 最近 {min(limit, len(done_items))} 次"]
    for input_index, item in enumerate(done_items[:limit], start=1):
        crop_groups = _crop_groups_for_source(output_dir, item.path)
        excel_matches = sum(1 for crop_ids in crop_groups if crop_ids and excel_rows.get(crop_ids[0]) is not None)
        crop_total = len([crop for group in crop_groups for crop in group])
        lines.append("")
        if is_zh(lang):
            lines.append(f"输入 {input_index}: {item.status} | Excel 行 {excel_matches} | crops {crop_total}")
        else:
            lines.append(f"Input {input_index}: {item.status} | Excel rows {excel_matches} | crops {crop_total}")
        if not crop_groups:
            lines.append("- 没找到 crop 记录" if is_zh(lang) else "- no crop records found")
            continue
        for crop_ids in crop_groups:
            primary = crop_ids[0]
            row = excel_rows.get(primary)
            label = "+".join(crop_ids)
            if row is None:
                lines.append(f"- {label} -> 不在 Excel 中" if is_zh(lang) else f"- {label} -> not in Excel")
                continue
            status = str(row.get("Manual status") or "active")
            no = _format_excel_no(row.get("No."))
            excel_row = row.get("_row")
            excel_sheet = row.get("_sheet") or "Excel"
            if excel_sheet == INVOICE_EXP_SHEET:
                excel_sheet = "Excel"
            category = str(row.get("Accounting Category") or row.get("Type") or "Other")
            currency = str(row.get("\u539f\u5e01\u79cd") or "MXN")
            amount = float(row.get("\u539f\u91d1\u989d") or row.get("MXN Amount") or 0)
            merchant = str(row.get("Merchant") or "Unknown")
            combined = " | 合并为一行" if is_zh(lang) and len(crop_ids) > 1 else " | combined one row" if len(crop_ids) > 1 else ""
            if is_zh(lang):
                lines.append(f"- {label} -> {excel_sheet} 行 {excel_row}, No. {no} | {category} {currency} {amount:.2f} | {merchant} | {status}{combined}")
            else:
                lines.append(f"- {label} -> {excel_sheet} row {excel_row}, No. {no} | {category} {currency} {amount:.2f} | {merchant} | {status}{combined}")
    return "\n".join(lines)


def change_message(settings: Settings, user_id: int, args: list[str], lang: str = LANG_EN) -> str:
    if not args:
        return _change_usage_message(lang)
    output_dir = telegram_user_output_dir(settings, user_id)
    try:
        parsed = _parse_change_args(args)
        result = change_reimbursement_record(output_dir, **parsed)
        checked = refresh_checked_outputs(settings, user_id)
    except LookupError as exc:
        recent = ", ".join(available_crop_ids(output_dir)) or "none"
        if is_zh(lang):
            return f"{exc}\n最近 crops: {recent}"
        return f"{exc}\nRecent crops: {recent}"
    except FileNotFoundError:
        if is_zh(lang):
            return "还没有报销 Excel。请先发送照片。"
        return "No reimbursement Excel yet. Send photos first."
    except PermissionError:
        if is_zh(lang):
            return "Excel 已打开/锁定。请关闭后重新发送 /change。"
        return "Excel is open/locked. Close it and resend /change."
    except OSError as exc:
        if is_zh(lang):
            return f"Excel 已打开/锁定或不可用。请关闭后重新发送 /change。\n{exc}"
        return f"Excel is open/locked or unavailable. Close it and resend /change.\n{exc}"
    except ValueError as exc:
        if is_zh(lang):
            return f"{exc}\n用法: /change 021 type Other amount 33.35 currency USD + 备注"
        return f"{exc}\nUsage: /change 021 type Other amount 33.35 currency USD + comment"
    changed = _changed_fields_text(result.before, result.after)
    missing = f"\n缺失 crops: {len(checked.missing_crops)}" if is_zh(lang) and checked.missing_crops else f"\nMissing crops: {len(checked.missing_crops)}" if checked.missing_crops else ""
    if is_zh(lang):
        return "\n".join(
            [
                "修改已保存",
                f"Crop: {result.crop_id}",
                f"商户: {result.merchant}",
                f"状态: {result.status}",
                changed or "已修改: 仅状态",
                f"Checked 行/crops: {checked.records_written}/{checked.crops_written}{missing}",
            ]
        )
    return "\n".join(
        [
            "Change saved",
            f"Crop: {result.crop_id}",
            f"Merchant: {result.merchant}",
            f"Status: {result.status}",
            changed or "Changed: status only",
            f"Checked rows/crops: {checked.records_written}/{checked.crops_written}{missing}",
        ]
    )


def delete_message(settings: Settings, user_id: int, args: list[str], lang: str = LANG_EN) -> str:
    crop_ids = _delete_crop_ids(args)
    if not crop_ids:
        return _delete_usage_message(lang)
    output_dir = telegram_user_output_dir(settings, user_id)
    try:
        results = [change_reimbursement_record(output_dir, crop_id, status="delete") for crop_id in crop_ids]
        checked = refresh_checked_outputs(settings, user_id)
    except LookupError as exc:
        recent = ", ".join(available_crop_ids(output_dir)) or "none"
        if is_zh(lang):
            return f"{exc}\n最近 crops: {recent}"
        return f"{exc}\nRecent crops: {recent}"
    except FileNotFoundError:
        if is_zh(lang):
            return "还没有报销 Excel。请先发送照片。"
        return "No reimbursement Excel yet. Send photos first."
    except PermissionError:
        if is_zh(lang):
            return "Excel 已打开/锁定。请关闭后重新发送 /del。"
        return "Excel is open/locked. Close it and resend /del."
    except OSError as exc:
        if is_zh(lang):
            return f"Excel 已打开/锁定或不可用。请关闭后重新发送 /del。\n{exc}"
        return f"Excel is open/locked or unavailable. Close it and resend /del.\n{exc}"
    except ValueError as exc:
        if is_zh(lang):
            return f"{exc}\n用法: /del 021 或 /del 021 022"
        return f"{exc}\nUsage: /del 021 or /del 021 022"
    lines = ["已删除" if is_zh(lang) else "Deleted", f"Crops: {', '.join(result.crop_id for result in results)}"]
    if len(results) == 1:
        lines.append(f"{'商户' if is_zh(lang) else 'Merchant'}: {results[0].merchant}")
    else:
        lines.append(f"{'数量' if is_zh(lang) else 'Count'}: {len(results)}")
    lines.append(f"{'Checked 行/crops' if is_zh(lang) else 'Checked rows/crops'}: {checked.records_written}/{checked.crops_written}")
    return "\n".join(lines)


def rerun_message(settings: Settings, user_id: int, lang: str = LANG_EN) -> str:
    try:
        result = rerun_finance_edits(settings, user_id)
    except FileNotFoundError as exc:
        missing = Path(str(exc).strip() or "").name or "checked/baseline"
        if is_zh(lang):
            return f"无法 rerun：缺少 {missing}。\n请先用 /report 生成财务 Excel，然后修改财务 Excel 后再发 /rerun。"
        return f"Cannot rerun: missing {missing}.\nUse /report to generate the finance Excel, edit it, then send /rerun."
    except PermissionError:
        if is_zh(lang):
            return "Excel 已打开/锁定。请关闭财务 Excel 后重新发送 /rerun。"
        return "Excel is open/locked. Close the finance Excel and resend /rerun."
    except OSError as exc:
        if is_zh(lang):
            return f"Excel 已打开/锁定或不可用。请关闭后重新发送 /rerun。\n{exc}"
        return f"Excel is open/locked or unavailable. Close it and resend /rerun.\n{exc}"
    except ValueError as exc:
        if is_zh(lang):
            return f"无法 rerun：{exc}"
        return f"Cannot rerun: {exc}"
    if not result.moved and not result.changed:
        if is_zh(lang):
            return "Rerun：没有发现财务 Excel 和 baseline 的差异。未重建文件。"
        return "Rerun: no differences found between the finance Excel and baseline. Nothing rebuilt."
    if is_zh(lang):
        lines = [
            "Rerun 完成",
            f"Checked 行/crops: {result.records_written}/{result.crops_written}",
            f"归档: {result.archive_dir}",
        ]
        if result.moved:
            lines.append("移动:")
            lines.extend(f"- {item}" for item in result.moved)
        if result.changed:
            lines.append("财务行内容变化:")
            lines.extend(f"- {item}" for item in result.changed)
        if result.warnings:
            lines.append("警告:")
            lines.extend(f"- {item}" for item in result.warnings[:5])
        lines.append("新的人工表、checked Excel 和 final_crops 已重建。/change 和 /del 仍可继续使用。")
        return "\n".join(lines)
    lines = [
        "Rerun complete",
        f"Checked rows/crops: {result.records_written}/{result.crops_written}",
        f"Archive: {result.archive_dir}",
    ]
    if result.moved:
        lines.append("Moved:")
        lines.extend(f"- {item}" for item in result.moved)
    if result.changed:
        lines.append("Finance row value changes:")
        lines.extend(f"- {item}" for item in result.changed)
    if result.warnings:
        lines.append("Warnings:")
        lines.extend(f"- {item}" for item in result.warnings[:5])
    lines.append("New manual workbook, checked Excel, and final_crops were rebuilt. /change and /del still work.")
    return "\n".join(lines)


def _delete_crop_ids(args: list[str]) -> list[str]:
    crop_ids: list[str] = []
    seen: set[str] = set()
    for arg in args:
        for part in re.split(r"[,;\s]+", arg.strip()):
            crop_id = part.strip()
            if not crop_id or crop_id in seen:
                continue
            crop_ids.append(crop_id)
            seen.add(crop_id)
    return crop_ids


def _change_usage_message(lang: str = LANG_EN) -> str:
    if is_zh(lang):
        return "\n".join(
            [
                "修改 crop /change",
                "用法:",
                "/change 021 type Other",
                "/change 021 amount 33.35 currency USD",
                "/change 021 + 备注文字",
                "删除请用: /del 021",
            ]
        )
    return "\n".join(
        [
            "Change crop /change",
            "Usage:",
            "/change 021 type Other",
            "/change 021 amount 33.35 currency USD",
            "/change 021 + note text",
            "Use /del 021 to delete.",
        ]
    )


def _delete_usage_message(lang: str = LANG_EN) -> str:
    if is_zh(lang):
        return "\n".join(["删除 crop /del", "用法:", "/del 021", "/del 021 022"])
    return "\n".join(["Delete crop /del", "Usage:", "/del 021", "/del 021 022"])


def review_crop_paths(settings: Settings, user_id: int) -> list[Path]:
    refresh_checked_outputs(settings, user_id)
    output_dir = telegram_user_output_dir(settings, user_id)
    crops_dir = output_dir / REVIEW_CROPS_DIR
    crop_paths = _review_crop_files(crops_dir)
    latest = _latest_done_item(settings, user_id)
    if latest is None:
        return crop_paths
    names = _crop_names_for_source(output_dir, latest.path)
    if names:
        by_name = {path.name: path for path in crop_paths}
        matched = [by_name[name] for name in names if name in by_name]
        if matched:
            return matched
    source_paths = _crop_paths_for_source(output_dir, latest.path)
    if source_paths:
        return source_paths
    row_count = int(latest.row_count or 0)
    if row_count > 0:
        return crop_paths[-row_count:]
    return []


def _review_crop_files(crops_dir: Path) -> list[Path]:
    if not crops_dir.exists():
        return []
    return sorted(
        (path for path in crops_dir.glob("*.jpg") if path.is_file()),
        key=lambda path: (path.stat().st_mtime, path.name),
    )


def _parse_change_args(args: list[str]) -> dict[str, object]:
    if len(args) < 2:
        raise ValueError("Missing change details")
    crop_id = args[0]
    status = "ok"
    category: str | None = None
    amount: float | None = None
    currency: str | None = None
    comment: str | None = None
    tokens = args[1:]
    keys = {"type", "category", "cat", "amount", "amt", "currency", "cur"}
    statuses = {"ok", "correct", "corrected"}
    index = 0
    saw_change = False
    while index < len(tokens):
        token = tokens[index].casefold()
        if tokens[index] == "+":
            comment = " ".join(tokens[index + 1 :]).strip()
            saw_change = bool(comment) or saw_change
            break
        if token in {"delete", "deleted"}:
            raise ValueError("Use /del 021 to delete a crop")
        if token in statuses:
            status = "correct" if token in {"correct", "corrected"} else "ok"
            saw_change = True
            index += 1
            continue
        if token not in keys:
            raise ValueError(f"Unknown /change token: {tokens[index]}")
        index += 1
        value_parts: list[str] = []
        while index < len(tokens) and tokens[index] != "+" and tokens[index].casefold() not in keys and tokens[index].casefold() not in statuses:
            value_parts.append(tokens[index])
            index += 1
        if not value_parts:
            raise ValueError(f"Missing value after {token}")
        value = " ".join(value_parts)
        if token in {"type", "category", "cat"}:
            category = value
        elif token in {"amount", "amt"}:
            try:
                amount = float(value.replace(",", ""))
            except ValueError as exc:
                raise ValueError(f"Invalid amount: {value}") from exc
        elif token in {"currency", "cur"}:
            currency = value.upper()
        saw_change = True
    if not saw_change:
        raise ValueError("Missing change details")
    return {"crop_id": crop_id, "category": category, "amount": amount, "currency": currency, "comment": comment, "status": status}


def _changed_fields_text(before: dict[str, object], after: dict[str, object]) -> str:
    labels = ["Type", "Accounting Category", "MXN Amount", "\u539f\u5e01\u79cd", "\u539f\u91d1\u989d", "\u6c47\u7387", "Detail", "Manual status"]
    lines: list[str] = []
    for label in labels:
        old = before.get(label)
        new = after.get(label)
        if old != new:
            lines.append(f"{label}: {_display_value(old)} -> {_display_value(new)}")
    if not lines:
        return ""
    return "Changed:\n" + "\n".join(f"- {line}" for line in lines)


def _display_value(value: object) -> str:
    if value in (None, ""):
        return "(blank)"
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def format_telegram_config(settings: Settings, auto_process: bool | None = None) -> str:
    allowed_count = len(settings.telegram_allowed_user_ids)
    lines = [
        "Telegram bot config:",
        f"Bot token: {'configured' if settings.telegram_bot_token else 'missing'}",
        f"Telegram package: {'installed' if telegram_package_ready() else 'missing'}",
        f"Allowed user IDs: {allowed_count if allowed_count else 'missing'}",
        f"Auto process: {'enabled' if resolve_auto_process(settings, auto_process) else 'disabled'}",
        f"Telegram language: {normalize_telegram_language(settings.telegram_language)}",
        f"Inbound photo folder: {settings.inbound_dir / 'telegram' / '<telegram_user_id>' / 'YYYY-MM-DD'}",
        f"Qwen OCR: {'enabled' if settings.qwen_api_key else 'disabled'}",
        f"Codex Scan fallback: {'enabled' if settings.codex_scan_enabled and settings.openai_api_key else 'disabled'}",
        f"Polling startup: {'READY' if telegram_polling_ready(settings) else 'NOT READY'}",
        f"Photo ingestion: {'READY' if telegram_config_ready(settings) else 'NOT READY'}",
    ]
    if not settings.telegram_bot_token:
        lines.append("Set TELEGRAM_BOT_TOKEN in .env before starting polling.")
    if not telegram_package_ready():
        lines.append('Install python-telegram-bot with: python -m pip install "python-telegram-bot>=22.0"')
    if not settings.telegram_allowed_user_ids:
        lines.append("Set TELEGRAM_ALLOWED_USER_IDS before sending photos; empty allow-list rejects all photos.")
    if settings.telegram_bot_token and not settings.telegram_allowed_user_ids:
        lines.append("Polling can start for /whoami, but receipt photos will be rejected until allowed IDs are set.")
    lines.append(f"Status: {'READY' if telegram_config_ready(settings) else 'NOT READY'}")
    return "\n".join(lines)


def run_polling_bot(settings: Settings, auto_process: bool | None = None) -> None:
    if not settings.telegram_bot_token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN is not configured")
    lock_path = _acquire_telegram_instance_lock(settings)
    try:
        from telegram import BotCommand, MenuButtonCommands, Update
        from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters
    except ImportError as exc:
        _release_telegram_instance_lock(lock_path)
        raise RuntimeError("Install python-telegram-bot to use Telegram ingestion") from exc

    pending_submit_users: set[int] = set()

    async def post_init(application: Application) -> None:
        await application.bot.set_my_commands([BotCommand(command, description) for command, description in telegram_command_menu(user_language(settings))])
        await application.bot.set_chat_menu_button(menu_button=MenuButtonCommands())

    async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if update.effective_message:
            user = update.effective_user
            await update.effective_message.reply_text(telegram_start_message(settings, user_language(settings, user.id if user else None)))

    def allowed(update: Update) -> bool:
        user = update.effective_user
        return bool(user and is_allowed_user(user.id, settings.telegram_allowed_user_ids))

    async def reply_not_allowed(update: Update) -> bool:
        if allowed(update):
            return False
        if update.effective_message:
            user = update.effective_user
            lang = user_language(settings, user.id if user else None)
            await update.effective_message.reply_text("这个 Telegram 用户没有权限。" if is_zh(lang) else "This Telegram user is not allowed.")
        return True

    async def whoami(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        message = update.effective_message
        user = update.effective_user
        if message is not None and user is not None:
            await message.reply_text(whoami_message(user.id, user.username, user_language(settings, user.id)))

    async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if update.effective_message:
            user = update.effective_user
            await update.effective_message.reply_text(telegram_help_message(user_language(settings, user.id if user else None)))

    async def lang_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        if not context.args:
            current = user_language(settings, user.id)
            text = "当前语言：中文。使用 /lang en 切换英文。" if is_zh(current) else "Current language: English. Use /lang zh for Chinese."
            await message.reply_text(text)
            return
        selected = set_user_language(settings, user.id, context.args[0])
        await context.bot.set_my_commands([BotCommand(command, description) for command, description in telegram_command_menu(selected)])
        await message.reply_text("语言已切换为中文。" if selected == LANG_ZH else "Language switched to English.")

    async def status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        if update.effective_message and update.effective_user:
            await update.effective_message.reply_text(status_message(settings, update.effective_user.id, user_language(settings, update.effective_user.id)))

    async def restart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        if update.effective_message and update.effective_user:
            count, summary = retry_failed(settings, update.effective_user.id)
            started = start_background_worker(
                settings,
                update.effective_user.id,
                item_callback=_telegram_item_notifier(settings, context.bot, asyncio.get_running_loop()),
            )
            lang = user_language(settings, update.effective_user.id)
            await update.effective_message.reply_text(
                "\n".join(["重试失败照片", f"已重新入队：{count}", f"扫描器：{'已启动' if started else '正在运行'}", _format_status(summary, lang)])
                if is_zh(lang)
                else "\n".join(["Restart", f"Failed photos retried: {count}", f"Scanner: {'started' if started else 'already running'}", _format_status(summary, lang)])
            )

    async def report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        if update.effective_message and update.effective_user:
            await update.effective_message.reply_text(report_message(settings, update.effective_user.id, user_language(settings, update.effective_user.id)))

    async def recent(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        if update.effective_message and update.effective_user:
            await update.effective_message.reply_text(recent_message(settings, update.effective_user.id, lang=user_language(settings, update.effective_user.id)))

    async def submitted(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        if update.effective_message and update.effective_user:
            text = submitted_batches_text(settings, update.effective_user.id)
            if is_zh(user_language(settings, update.effective_user.id)):
                text = text.replace("No submitted batches yet.", "还没有已提交批次。").replace("Submitted / 已提交", "已提交")
            await update.effective_message.reply_text(text)

    async def submit(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        if update.effective_message and update.effective_user:
            lang = user_language(settings, update.effective_user.id)
            text = submit_message(settings, update.effective_user.id, list(context.args or []), lang)
            if not context.args and submit_pending_confirmation(settings, update.effective_user.id):
                pending_submit_users.add(update.effective_user.id)
            await update.effective_message.reply_text(text)

    async def text_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None or user.id not in pending_submit_users:
            return
        handled, text = submit_confirmation_message(settings, user.id, message.text or "", user_language(settings, user.id))
        if handled:
            pending_submit_users.discard(user.id)
        await message.reply_text(text)

    async def excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        workbook = telegram_user_workbook(settings, user.id)
        if not workbook.exists():
            await message.reply_text(status_message(settings, user.id, user_language(settings, user.id)))
            return
        refresh_checked_outputs(settings, user.id)
        try:
            focus_reimbursement_workbook(workbook)
        except OSError:
            lang = user_language(settings, user.id)
            await message.reply_text("Excel 已打开/锁定；如果想让文件打开时定位到今天的行，请先关闭它。" if is_zh(lang) else "Excel file is open/locked. Close it if you want the file to open at today's row.")
        with workbook.open("rb") as handle:
            await message.reply_document(document=handle, filename=workbook.name)

    async def checked_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        output_dir = telegram_user_output_dir(settings, user.id)
        workbook = checked_workbook_path(output_dir)
        try:
            refresh_checked_outputs(settings, user.id)
        except OSError:
            lang = user_language(settings, user.id)
            await message.reply_text("Excel 已打开/锁定。请关闭后重新发送 /checked。" if is_zh(lang) else "Excel is open/locked. Close it and resend /checked.")
            return
        if not workbook.exists():
            lang = user_language(settings, user.id)
            await message.reply_text("还没有财务版 checked Excel。请先发送照片或使用 /report 生成。" if is_zh(lang) else "No checked finance Excel yet. Send photos first or use /report.")
            return
        with workbook.open("rb") as handle:
            await message.reply_document(document=handle, filename=workbook.name)

    async def crops(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        lang = user_language(settings, user.id)
        crop_paths = review_crop_paths(settings, user.id)
        if not crop_paths:
            await message.reply_text("还没有 review crops。请先发送照片，或等待扫描完成。" if is_zh(lang) else "No review crops yet. Send photos first, or wait for scanning to finish.")
            return
        await message.reply_text(f"最近 review crops：{len(crop_paths)}" if is_zh(lang) else f"Latest review crops: {len(crop_paths)}")
        sent = 0
        failed: list[str] = []
        for index, crop_path in enumerate(crop_paths, start=1):
            caption = f"{index}/{len(crop_paths)} {crop_path.name}"
            try:
                with crop_path.open("rb") as handle:
                    await message.reply_photo(photo=handle, caption=caption)
                sent += 1
            except Exception as exc:
                try:
                    with crop_path.open("rb") as handle:
                        await message.reply_document(document=handle, filename=crop_path.name, caption=caption)
                    sent += 1
                except Exception as fallback_exc:
                    failed.append(f"{crop_path.name}: {fallback_exc or exc}")
        if failed:
            await message.reply_text(("Crop 发送警告:\n" if is_zh(lang) else "Crop send warning:\n") + "\n".join(failed[:5]))
        await message.reply_text(f"已发送 crops: {sent}/{len(crop_paths)}" if is_zh(lang) else f"Crops sent: {sent}/{len(crop_paths)}")

    async def change(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        await message.reply_text(change_message(settings, user.id, list(context.args or []), user_language(settings, user.id)))

    async def delete_crop(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        await message.reply_text(delete_message(settings, user.id, list(context.args or []), user_language(settings, user.id)))

    async def rerun(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        if await reply_not_allowed(update):
            return
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        await message.reply_text(rerun_message(settings, user.id, user_language(settings, user.id)))

    async def photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        message = update.effective_message
        user = update.effective_user
        if message is None or user is None:
            return
        if not is_allowed_user(user.id, settings.telegram_allowed_user_ids):
            lang = user_language(settings, user.id)
            await message.reply_text("这个 Telegram 用户没有权限。" if is_zh(lang) else "This Telegram user is not allowed.")
            return
        if not message.photo:
            lang = user_language(settings, user.id)
            await message.reply_text("请发送照片。" if is_zh(lang) else "Please send a photo.")
            return

        photo_size = message.photo[-1]
        telegram_file = await context.bot.get_file(photo_size.file_id)
        now = datetime.now()
        day_dir = telegram_user_day_dir(settings, user.id, now)
        day_dir.mkdir(parents=True, exist_ok=True)
        target = day_dir / telegram_photo_filename(now, photo_size.file_id, photo_size.file_unique_id)
        await telegram_file.download_to_drive(custom_path=Path(target))
        enqueue_photo(settings, user.id, target, now)

        should_process = resolve_auto_process(settings, auto_process)
        if should_process:
            started = start_background_worker(
                settings,
                user.id,
                item_callback=_telegram_item_notifier(settings, context.bot, asyncio.get_running_loop()),
            )
            await message.reply_text(queued_photo_message(target, started, user_language(settings, user.id)))
        else:
            await message.reply_text(saved_photo_message(target, user_language(settings, user.id)))

    app = Application.builder().token(settings.telegram_bot_token).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("lang", lang_command))
    app.add_handler(CommandHandler("whoami", whoami))
    app.add_handler(CommandHandler("status", status))
    app.add_handler(CommandHandler("restart", restart))
    app.add_handler(CommandHandler("excel", excel))
    app.add_handler(CommandHandler("checked", checked_excel))
    app.add_handler(CommandHandler("crops", crops))
    app.add_handler(CommandHandler("change", change))
    app.add_handler(CommandHandler("del", delete_crop))
    app.add_handler(CommandHandler("rerun", rerun))
    app.add_handler(CommandHandler("today_excel", excel))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("recent", recent))
    app.add_handler(CommandHandler("unsubmitted", report))
    app.add_handler(CommandHandler("submitted", submitted))
    app.add_handler(CommandHandler("submit", submit))
    app.add_handler(MessageHandler(filters.PHOTO, photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_message))
    try:
        app.run_polling()
    finally:
        _release_telegram_instance_lock(lock_path)


def _acquire_telegram_instance_lock(settings: Settings) -> Path:
    lock_path = settings.output_dir / "telegram_bot.pid"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    while True:
        try:
            fd = os.open(str(lock_path), os.O_WRONLY | os.O_CREAT | os.O_EXCL)
        except FileExistsError:
            existing_pid = _read_pid_file(lock_path)
            if existing_pid and _pid_is_running(existing_pid):
                raise RuntimeError(f"Telegram bot already running with PID {existing_pid}. Stop that process before starting another one.")
            try:
                lock_path.unlink()
            except OSError as exc:
                if exc.errno not in {errno.ENOENT, errno.EACCES, errno.EPERM}:
                    raise
                raise RuntimeError(f"Telegram bot lock exists but cannot be removed: {lock_path}") from exc
            continue
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            handle.write(str(os.getpid()))
        return lock_path


def _release_telegram_instance_lock(lock_path: Path) -> None:
    try:
        if _read_pid_file(lock_path) == os.getpid():
            lock_path.unlink()
    except OSError:
        pass


def _read_pid_file(path: Path) -> int | None:
    try:
        return int(path.read_text(encoding="utf-8").strip())
    except (OSError, ValueError):
        return None


def _pid_is_running(pid: int) -> bool:
    if pid <= 0:
        return False
    if os.name == "nt":
        try:
            import ctypes

            process_query_limited_information = 0x1000
            handle = ctypes.windll.kernel32.OpenProcess(process_query_limited_information, False, int(pid))
            if not handle:
                return False
            ctypes.windll.kernel32.CloseHandle(handle)
            return True
        except Exception:
            return False
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def _telegram_item_notifier(settings: Settings, bot, loop: asyncio.AbstractEventLoop):
    def notify(user_id: int, item: QueueItem, records: list[InvoiceRecord]) -> None:
        text = scan_completion_message(settings, user_id, item, records, user_language(settings, user_id))
        if not text:
            return
        asyncio.run_coroutine_threadsafe(bot.send_message(chat_id=user_id, text=text), loop)

    return notify


def _category_totals(records: list[InvoiceRecord]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for record in records:
        category = normalize_expense_category(record.expense_category)
        amount = float(record.total_amount or 0)
        totals[category] = round(totals.get(category, 0.0) + amount, 2)
    return {category: amount for category, amount in totals.items() if amount}


def _original_currency_totals(records: list[InvoiceRecord]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for record in records:
        currency = str(getattr(record, "original_currency", record.currency) or "MXN").upper()
        amount = float(getattr(record, "original_amount", record.total_amount) or 0)
        totals[currency] = round(totals.get(currency, 0.0) + amount, 2)
    return {currency: amount for currency, amount in totals.items() if amount}


def _format_currency_lines(totals: dict[str, float]) -> list[str]:
    if not totals:
        return ["- none"]
    return [f"- {currency}: {amount:.2f}" for currency, amount in sorted(totals.items()) if amount]


def _format_category_block(totals: dict[str, float], lang: str = LANG_EN) -> list[str]:
    lines = ["按类别：" if is_zh(lang) else "By category:"]
    amounts = [(category, amount) for category, amount in sorted(totals.items()) if amount]
    if not amounts:
        return lines + ["- none"]
    return lines + [f"- {category}: {amount:.2f}" for category, amount in amounts]


def _format_status(summary, lang: str = LANG_EN) -> str:
    if not is_zh(lang):
        return format_status(summary)
    lines = [
        "状态",
        f"用户: {summary.user_id}",
        f"扫描器: {summary.worker_status}",
        f"等待: {summary.pending}",
        f"处理中: {summary.processing}",
        f"完成: {summary.done}",
        f"失败: {summary.failed}",
        f"Excel: {summary.excel_path}",
    ]
    if summary.current_photo:
        lines.append(f"当前照片: {summary.current_photo}")
    if summary.last_error:
        lines.append(f"最近错误: {summary.last_error}")
    if summary.failed:
        lines.append("用 /restart 重试失败照片。")
    return "\n".join(lines)


def _format_reimbursement_summary(summary, title: str = "报销汇总 / 未提交", lang: str = LANG_EN) -> str:
    if not is_zh(lang):
        return format_reimbursement_summary(summary, title=title)
    lines = [
        title,
        f"照片数: {summary.photo_count}",
        f"发票行数: {summary.record_count}",
        f"MXN 总额: {summary.total_amount:.2f}",
    ]
    if summary.date_min or summary.date_max:
        lines.append(f"日期范围: {summary.date_min or '-'} 到 {summary.date_max or '-'}")
    if summary.category_totals:
        lines.append("按类别:")
        for category, amount in sorted(summary.category_totals.items()):
            lines.append(f"- {category}: {amount:.2f}")
    lines.append(f"失败照片: {summary.failed_count}")
    return "\n".join(lines)


def _format_submit_result(result, lang: str = LANG_EN) -> str:
    if not is_zh(lang):
        return format_submit_result(result)
    if result is None:
        return "提交\n没有未提交记录。"
    lines = [
        "提交完成",
        f"批次: {result.batch_id}",
        f"记录数: {result.record_count}",
        f"MXN 总额: {result.total_amount:.2f}",
        f"财务 Excel: {result.archived_excel}",
        f"财务 crops: {result.archived_crops}",
        "按类别:",
    ]
    for category, amount in sorted(result.category_totals.items()):
        lines.append(f"- {category}: {amount:.2f}")
    lines.append("当前 Excel 已归档。新扫描会从 001 开始新的报销表。")
    if result.missing_crops:
        lines.append(f"警告: 缺失 crop 链接={len(result.missing_crops)}")
    return "\n".join(lines)


def _format_scan_record_lines(records: list[InvoiceRecord]) -> list[str]:
    if not records:
        return ["- none"]
    lines: list[str] = []
    for position, record in enumerate(records, start=1):
        crop_ids = _crop_ids_from_record(record)
        index = "+".join(crop_ids) if crop_ids else (f"{record.line_no:03d}" if record.line_no else f"{position:03d}")
        category = normalize_expense_category(record.expense_category, f"{record.seller} {record.contents}")
        currency = str(getattr(record, "original_currency", record.currency) or "MXN").upper()
        amount = float(getattr(record, "original_amount", record.total_amount) or 0)
        seller = str(record.seller or "Unknown").strip() or "Unknown"
        invoice_date = str(record.invoice_date or "unknown-date").strip() or "unknown-date"
        lines.append(f"- {index} {invoice_date} {category}: {currency} {amount:.2f} | {seller}")
        if len(crop_ids) > 1:
            lines.append(f"  Combined: {' + '.join(crop_ids)} -> one Excel row; crops kept as a/b")
    return lines


def _crop_id_from_record(record: InvoiceRecord) -> str:
    name = Path(str(record.crop_image or "")).name
    match = re.match(r"(\d{3,})_", name)
    return match.group(1) if match else ""


def _crop_ids_from_record(record: InvoiceRecord) -> list[str]:
    ids: list[str] = []
    for crop_image in [record.crop_image, *list(getattr(record, "supporting_crop_images", []) or [])]:
        name = Path(str(crop_image or "")).name
        match = re.match(r"(\d{3,})_", name)
        if match and match.group(1) not in ids:
            ids.append(match.group(1))
    return ids


def _record_index_summary(records: list[InvoiceRecord], fallback_photo: Path) -> str:
    if not records:
        return fallback_photo.name
    labels: list[str] = []
    for record in records[:4]:
        index = f"{record.line_no:03d}" if record.line_no else "unknown"
        invoice_date = record.invoice_date or "unknown-date"
        labels.append(f"{index}_{invoice_date}")
    if len(records) > 4:
        labels.append(f"+{len(records) - 4} more")
    return ", ".join(labels)


def _latest_done_item(settings: Settings, user_id: int) -> QueueItem | None:
    state = load_queue_state(telegram_user_queue_path(settings, user_id))
    done = [item for item in state.items if item.status == DONE]
    if not done:
        return None
    return max(done, key=lambda item: (item.received_at or item.updated_at or "", item.updated_at or ""))


def _crop_names_for_source(output_dir: Path, source_image: str) -> list[str]:
    return [path.name for path in _crop_paths_for_source(output_dir, source_image, existing_only=False)]


def _crop_groups_for_source(output_dir: Path, source_image: str) -> list[list[str]]:
    state_path = output_dir / "processing_state.json"
    if not state_path.exists():
        return []
    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    source_key = _path_key_text(source_image)
    groups: list[list[str]] = []
    for record in data.get("records", []):
        if not isinstance(record, dict):
            continue
        if _path_key_text(str(record.get("source_image") or "")) != source_key:
            continue
        ids: list[str] = []
        for crop_text in [str(record.get("crop_image") or "").strip(), *[str(item or "").strip() for item in record.get("supporting_crop_images", []) or []]]:
            crop_id = _crop_id_from_path_text(crop_text)
            if crop_id and crop_id not in ids:
                ids.append(crop_id)
        if ids:
            groups.append(ids)
    return groups


def _crop_paths_for_source(output_dir: Path, source_image: str, existing_only: bool = True) -> list[Path]:
    state_path = output_dir / "processing_state.json"
    if not state_path.exists():
        return []
    try:
        data = json.loads(state_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    source_key = _path_key_text(source_image)
    paths: list[Path] = []
    for record in data.get("records", []):
        if not isinstance(record, dict):
            continue
        if _path_key_text(str(record.get("source_image") or "")) != source_key:
            continue
        for crop_text in [str(record.get("crop_image") or "").strip(), *[str(item or "").strip() for item in record.get("supporting_crop_images", []) or []]]:
            if not crop_text:
                continue
            path = Path(crop_text)
            if not path.is_absolute():
                path = output_dir / path
            if path in paths:
                continue
            if not existing_only or path.exists():
                paths.append(path)
    return paths


def _excel_rows_by_crop_id(workbook: Path) -> dict[str, dict[str, object]]:
    if not workbook.exists():
        return {}
    try:
        wb = load_workbook(workbook, data_only=True)
    except Exception:
        return {}
    try:
        rows: dict[str, dict[str, object]] = {}
        worksheets = [wb[INVOICE_EXP_SHEET]] if INVOICE_EXP_SHEET in wb.sheetnames else [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
        for ws in worksheets:
            headers = {str(ws.cell(1, col).value or ""): col for col in range(1, ws.max_column + 1)}
            if "Invoice link" not in headers or "MXN Amount" not in headers:
                continue
            link_col = headers.get("Invoice link", 2)
            for row_idx in range(2, ws.max_row + 1):
                link_cell = ws.cell(row_idx, link_col)
                link = str(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value or "")
                crop_id = _crop_id_from_path_text(link)
                if not crop_id:
                    continue
                row = {header: ws.cell(row_idx, col).value for header, col in headers.items() if header}
                row["_row"] = row_idx
                row["_sheet"] = ws.title
                rows[crop_id] = row
        return rows
    finally:
        wb.close()


def _crop_id_from_path_text(value: str) -> str:
    name = Path(str(value or "")).name
    match = re.match(r"(\d{3,})_", name)
    return match.group(1) if match else ""


def _format_excel_no(value: object) -> str:
    try:
        return f"{int(value):03d}"
    except (TypeError, ValueError):
        return str(value or "unknown")


def _path_key_text(value: str) -> str:
    try:
        return str(Path(value).resolve()).casefold()
    except OSError:
        return str(value or "").casefold()


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value.strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
    return cleaned[:120]
