import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from invoice_system.ab_test import run_ab_test, telegram_ab_input
from invoice_system.config import Settings
from invoice_system.models import InvoiceRecord, OCRResult, OCRTextLine


class ABTestTests(unittest.TestCase):
    def test_run_ab_test_writes_two_workbooks_and_comparison(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            image = root / "photo.jpg"
            image.write_bytes(b"not a real jpg")
            output = root / "ab"
            settings = Settings(root=root)

            summary = run_ab_test(
                settings,
                image,
                output,
                paddle_vl_recognizer=FakeRecognizer("paddleocr_vl", "Paddle Cafe", 120),
                qwen_recognizer=FakeRecognizer("qwen_scan", "Qwen Cafe", 126),
            )

            self.assertEqual(summary.source_images, 1)
            self.assertEqual(summary.crops, 1)
            self.assertTrue((output / "paddleocr_vl" / "Invoice_Output.xlsx").exists())
            self.assertTrue((output / "qwen" / "Invoice_Output.xlsx").exists())
            self.assertTrue(summary.comparison_path.exists())

            wb = load_workbook(summary.comparison_path, data_only=True)
            try:
                rows = list(wb["AB_Comparison"].iter_rows(values_only=True))
            finally:
                wb.close()
            self.assertEqual(rows[0][:4], ("Crop No.", "Field", "PaddleOCR-VL", "Qwen"))
            self.assertIn(("Seller", "Paddle Cafe", "Qwen Cafe"), [(row[1], row[2], row[3]) for row in rows])

    def test_telegram_ab_input_uses_date_or_latest_folder(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(root=root, inbound_dir=root / "data" / "inbound")
            first = settings.inbound_dir / "telegram" / "123" / "2026-06-15"
            second = settings.inbound_dir / "telegram" / "123" / "2026-06-16"
            first.mkdir(parents=True)
            second.mkdir(parents=True)

            self.assertEqual(telegram_ab_input(settings, 123), second)
            self.assertEqual(telegram_ab_input(settings, 123, "2026-06-15"), first)


class FakeRecognizer:
    def __init__(self, engine: str, seller: str, amount: float) -> None:
        self.engine = engine
        self.seller = seller
        self.amount = amount

    def recognize(self, image_path: Path) -> OCRResult:
        record = InvoiceRecord(
            invoice_date="2026-06-12",
            seller=self.seller,
            currency="MXN",
            total_amount=self.amount,
            expense_amount=self.amount,
        )
        return OCRResult(self.engine, [OCRTextLine(self.seller, 0.9)], record, 0.9)


if __name__ == "__main__":
    unittest.main()
