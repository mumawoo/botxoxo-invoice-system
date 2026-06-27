import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from invoice_system.compare import (
    _audit_crop_counts,
    _best_match,
    _comparison_category,
    _crop_count_rows,
    _diffs,
    _find_excel,
    _image_crop_counts,
    _match_score,
    _source_from_crop_filename,
    compare_outputs,
)
from invoice_system.excel_store import InvoiceWorkbook
from invoice_system.models import EXCEL_HEADERS, InvoiceRecord, OCRAuditRow


class CompareTests(unittest.TestCase):
    def test_summary_sheet_counts_exact_match(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            baseline.mkdir()
            candidate.mkdir()

            InvoiceWorkbook(baseline / "Invoice_Output.xlsx").write_records(
                [InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)]
            )
            InvoiceWorkbook(candidate / "Invoice_Output.xlsx").write_records(
                [InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)]
            )

            report = compare_outputs(baseline, candidate, root / "comparison.xlsx")

            wb = load_workbook(report, data_only=True)
            self.assertIn("Summary", wb.sheetnames)
            headers = [cell.value for cell in wb["Comparison"][1]]
            self.assertIn("Baseline VAT", headers)
            self.assertIn("Windows Tips", headers)
            summary = {row[0]: row[1] for row in wb["Summary"].iter_rows(values_only=True) if row[0]}
            self.assertEqual(summary["exact match"], 1)
            self.assertEqual(summary["Baseline rows"], 1)
            self.assertEqual(summary["Windows rows"], 1)
            self.assertEqual(summary["Baseline root"], str(baseline))
            self.assertEqual(summary["Windows root"], str(candidate))
            self.assertEqual(summary["Baseline workbook"], str(baseline / "Invoice_Output.xlsx"))
            self.assertEqual(summary["Windows workbook"], str(candidate / "Invoice_Output.xlsx"))
            self.assertEqual(summary["Report output"], str(root / "comparison.xlsx"))

    def test_compare_reads_invoice_sheet_when_baseline_active_sheet_is_summary(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            baseline.mkdir()
            candidate.mkdir()
            _write_workbook_with_summary_first(
                baseline / "Ubuntu_Output.xlsx",
                InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126),
            )
            InvoiceWorkbook(candidate / "Invoice_Output.xlsx").write_records(
                [InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)]
            )

            report = compare_outputs(baseline, candidate, root / "comparison.xlsx")

            wb = load_workbook(report, data_only=True)
            try:
                summary = {row[0]: row[1] for row in wb["Summary"].iter_rows(values_only=True) if row[0]}
                self.assertEqual(summary["Baseline rows"], 1)
                self.assertEqual(summary["exact match"], 1)
                self.assertEqual(summary["Baseline workbook"], str(baseline / "Ubuntu_Output.xlsx"))
            finally:
                wb.close()

    def test_summary_counts_codex_resolved_ocr_disagreement(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            baseline = root / "baseline"
            candidate = root / "candidate"
            baseline.mkdir()
            candidate.mkdir()

            InvoiceWorkbook(baseline / "Invoice_Output.xlsx").write_records(
                [InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)]
            )
            InvoiceWorkbook(candidate / "Invoice_Output.xlsx").write_records(
                [
                    InvoiceRecord(
                        invoice_date="2026-06-12",
                        seller="CAFE XUAN",
                        total_amount=126,
                        remarks="Codex Scan used: OCR mismatch",
                    )
                ]
            )

            report = compare_outputs(baseline, candidate, root / "comparison.xlsx")

            wb = load_workbook(report, data_only=True)
            try:
                summary = {row[0]: row[1] for row in wb["Summary"].iter_rows(values_only=True) if row[0]}
                self.assertEqual(summary["OCR disagreement resolved by Codex Scan"], 1)
                self.assertEqual(wb["Comparison"].cell(2, 1).value, "OCR disagreement resolved by Codex Scan")
            finally:
                wb.close()

    def test_find_excel_prefers_invoice_output(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "Invoice_Manually_Checked_Trial.xlsx").touch()
            (root / "comparison_report.xlsx").touch()
            output = root / "Invoice_Output_Trial.xlsx"
            output.touch()

            self.assertEqual(_find_excel(root), output)

    def test_find_excel_raises_when_missing(self):
        with tempfile.TemporaryDirectory() as temp:
            with self.assertRaises(FileNotFoundError):
                _find_excel(Path(temp))

    def test_crop_counts_use_ocr_audit_when_available(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / "Invoice_Output.xlsx"
            InvoiceWorkbook(workbook).write_records(
                [InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)],
                [
                    OCRAuditRow("photo_a.jpg", "crop_1.jpg", "local OCR agreement", False),
                    OCRAuditRow("photo_a.jpg", "crop_2.jpg", "local OCR agreement", False),
                    OCRAuditRow("photo_b.jpg", "crop_3.jpg", "local OCR agreement", False),
                ],
            )

            self.assertEqual(_audit_crop_counts(workbook), {"photo_a.jpg": 2, "photo_b.jpg": 1})

    def test_crop_counts_fall_back_to_image_filename_inference(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / "Invoice_Output.xlsx"
            InvoiceWorkbook(workbook).write_records([InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)])
            (root / "photo_a_d01.jpg").write_bytes(b"jpg")
            (root / "photo_a_d02.jpg").write_bytes(b"jpg")
            (root / "0003_2026-06-12_MXN_42.50_photo_b_d01.jpg").write_bytes(b"jpg")

            self.assertEqual(_image_crop_counts(root), {"photo_a": 2, "photo_b": 1})
            self.assertEqual(
                _crop_count_rows(root, workbook),
                [
                    ("photo_a", 2, "image filename inference"),
                    ("photo_b", 1, "image filename inference"),
                ],
            )

    def test_source_from_crop_filename_supports_raw_and_v2_names(self):
        self.assertEqual(_source_from_crop_filename(Path("photo_a_d03.jpg")), "photo_a")
        self.assertEqual(
            _source_from_crop_filename(Path("0007_2026-06-12_MXN_126.00_photo_a_d03.jpg")),
            "photo_a",
        )

    def test_best_match_rejects_amount_only_match(self):
        base = InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)
        candidate = InvoiceRecord(invoice_date="2026-06-13", seller="OTHER SHOP", total_amount=126)

        index, record = _best_match(base, [candidate], set())

        self.assertEqual(index, -1)
        self.assertIsNone(record)

    def test_best_match_accepts_amount_and_seller_match(self):
        base = InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)
        candidate = InvoiceRecord(invoice_date="2026-06-13", seller="CAFE XUAN", total_amount=126)

        index, record = _best_match(base, [candidate], set())

        self.assertEqual(index, 0)
        self.assertIs(record, candidate)

    def test_best_match_normalizes_date_and_fuzzy_seller(self):
        base = InvoiceRecord(invoice_date="12/06/2026", seller="Café Xuán", total_amount=126)
        candidate = InvoiceRecord(invoice_date="2026-06-12", seller="Cafe Xuan", total_amount=126)

        index, record = _best_match(base, [candidate], set())

        self.assertEqual(index, 0)
        self.assertIs(record, candidate)

    def test_match_score_uses_normalized_fields(self):
        base = InvoiceRecord(invoice_date="12/06/2026", seller="Café Xuán", total_amount=126)
        candidate = InvoiceRecord(invoice_date="2026-06-12 00:00:00", seller="Cafe Xuan", total_amount=126.49)

        score, corroborated = _match_score(base, candidate)

        self.assertEqual(score, 7)
        self.assertTrue(corroborated)

    def test_diffs_ignores_normalized_date_seller_and_currency_alias(self):
        base = InvoiceRecord(invoice_date="12/06/2026", seller="Café Xuán", currency="M.N.", total_amount=126)
        candidate = InvoiceRecord(invoice_date="2026-06-12", seller="Cafe Xuan", currency="MXN", total_amount=126.49)

        self.assertEqual(_diffs(base, candidate), [])

    def test_comparison_category_prioritizes_codex_ocr_mismatch(self):
        base = InvoiceRecord(invoice_date="2026-06-12", seller="CAFE XUAN", total_amount=126)
        candidate = InvoiceRecord(
            invoice_date="2026-06-12",
            seller="CAFE XUAN",
            total_amount=126,
            remarks="Codex Scan used: OCR mismatch",
        )

        category = _comparison_category(base, candidate, [])

        self.assertEqual(category, "OCR disagreement resolved by Codex Scan")


def _write_workbook_with_summary_first(path: Path, record: InvoiceRecord) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    ws = wb.create_sheet("Invoices")
    ws.append(EXCEL_HEADERS)
    record.line_no = 1
    ws.append(record.to_excel_row())
    wb.save(path)
    wb.close()


if __name__ == "__main__":
    unittest.main()
