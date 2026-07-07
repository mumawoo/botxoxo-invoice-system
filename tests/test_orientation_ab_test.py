import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from invoice_system.config import Settings
from invoice_system.models import CropResult, InvoiceRecord, OCRResult, OCRTextLine
from invoice_system.orientation_ab_test import prioritized_orientation_inputs, run_orientation_ab_test


class OrientationABTestTests(unittest.TestCase):
    def test_run_orientation_ab_test_writes_report_and_uses_splitter_modes(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "photo.jpg"
            source.write_bytes(b"fake image")
            output = root / "orientation_ab"
            FakeSplitter.created_modes.clear()

            with patch("invoice_system.orientation_ab_test.OpenCVInvoiceSplitter", FakeSplitter):
                summary = run_orientation_ab_test(Settings(root=root), source, output, limit=1, qwen_recognizer=FakeQwen())

            self.assertEqual(FakeSplitter.created_modes, [True, False])
            self.assertEqual(summary.source_images, 1)
            self.assertEqual(summary.local_crops, 1)
            self.assertEqual(summary.qwen_crops, 1)
            self.assertTrue((output / "local_orientation" / "crops").exists())
            self.assertTrue((output / "qwen_orientation" / "crops").exists())

            wb = load_workbook(summary.report_path, data_only=True)
            try:
                rows = list(wb["Orientation_AB_Report"].iter_rows(values_only=True))
            finally:
                wb.close()
            self.assertEqual(rows[0][:4], ("Crop No.", "Source Photo", "A Crop Path", "B Crop Path"))
            row = dict(zip(rows[0], rows[1]))
            self.assertEqual(row["A Qwen Rotate"], 180)
            self.assertEqual(row["B Qwen Rotate"], 180)
            self.assertEqual(row["Seller Match"], "yes")

    def test_prioritized_orientation_inputs_puts_orientation_records_first(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(root=root, inbound_dir=root / "data" / "inbound", output_dir=root / "data" / "output")
            day = settings.inbound_dir / "telegram" / "123" / "2026-07-03"
            day.mkdir(parents=True)
            normal = day / "normal.jpg"
            rotated = day / "rotated.jpg"
            normal.write_bytes(b"normal")
            rotated.write_bytes(b"rotated")
            state = settings.output_dir / "telegram" / "123" / "processing_state.json"
            state.parent.mkdir(parents=True)
            state.write_text(
                '{"records":[{"source_image":"' + str(rotated).replace("\\", "\\\\") + '","remarks":"qwen rotated crop 180deg 0.98"}]}',
                encoding="utf-8",
            )

            selected = prioritized_orientation_inputs(settings, 123, "2026-07-03")

            self.assertEqual(selected[0], rotated)
            self.assertIn(normal, selected)


class FakeSplitter:
    created_modes: list[bool] = []

    def __init__(self, crops_dir: Path, *, local_orientation: bool = True) -> None:
        self.crops_dir = crops_dir
        self.local_orientation = local_orientation
        self.created_modes.append(local_orientation)
        self.crops_dir.mkdir(parents=True, exist_ok=True)

    def split(self, image_path: Path) -> list[CropResult]:
        out = self.crops_dir / f"{image_path.stem}_d01.jpg"
        out.write_bytes(b"crop")
        return [CropResult(image_path, out, 1)]


class FakeQwen:
    engine = "qwen_scan"

    def recognize(self, image_path: Path) -> OCRResult:
        record = InvoiceRecord(
            invoice_date="2026-07-03",
            seller="Test Cafe",
            currency="MXN",
            total_amount=120.0,
            expense_category="Food",
        )
        return OCRResult(
            self.engine,
            [OCRTextLine("Test Cafe", 0.95)],
            record,
            0.95,
            rotate_degrees=180,
            orientation_confidence=0.98,
        )


if __name__ == "__main__":
    unittest.main()
