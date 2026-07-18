import tempfile
import unittest
import json
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from invoice_system.config import Settings
from invoice_system.dual_ocr import ResolvedScan
from invoice_system.models import InvoiceRecord, OCRResult, OCRTextLine
from invoice_system.pipeline import (
    InvoicePipeline,
    _apply_qwen_orientation,
    _copy_final_crops,
    _fill_missing_dates_from_neighbors,
    _final_crop_name,
    _source_qa_record,
    QwenOnlyResolver,
)
from invoice_system.reimbursement_excel import CHECKED_WORKBOOK_NAME, INVOICE_EXP_SHEET, reimbursement_workbook_path
from invoice_system.sample_data import create_synthetic_multi_receipt
from invoice_system.visual_count import VisualCountResult


class PipelineHelperTests(unittest.TestCase):
    def test_final_crop_name_uses_v2_pattern(self):
        record = InvoiceRecord(
            line_no=7,
            invoice_date="2026-06-12",
            currency="MXN",
            total_amount=126,
            seller="Cafe Xuan",
            source_image=r"C:\inbound\batch one.jpg",
        )

        name = _final_crop_name(record, Path("batch one_d03.jpg"))

        self.assertEqual(name, "007_2026-06-12_MXN_126.00_Cafe_Xuan.jpg")

    def test_copy_final_crops_copies_existing_crop_and_updates_record(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            crop = root / "raw" / "source_d02.jpg"
            crop.parent.mkdir()
            crop.write_bytes(b"jpg")
            record = InvoiceRecord(
                line_no=3,
                invoice_date="2026-06-12",
                currency="MXN",
                total_amount=42.5,
                seller="Market",
                source_image=str(root / "source.jpg"),
                crop_image=str(crop),
            )

            copied = _copy_final_crops([record], root / "final_crops")

            self.assertEqual(len(copied), 1)
            self.assertTrue(copied[0].exists())
            self.assertEqual(copied[0].name, "source_d02.jpg")
            self.assertEqual(record.crop_image, str(copied[0]))

    def test_copy_final_crops_copies_supporting_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            primary = root / "raw" / "040_invoice.jpg"
            supporting = root / "raw" / "041_payment.jpg"
            primary.parent.mkdir()
            primary.write_bytes(b"invoice")
            supporting.write_bytes(b"payment")
            record = InvoiceRecord(
                line_no=1,
                invoice_date="2026-06-12",
                currency="MXN",
                total_amount=126,
                seller="Cafe",
                crop_image=str(primary),
                supporting_crop_images=[str(supporting)],
            )

            copied = _copy_final_crops([record], root / "review_crops")

            self.assertEqual([path.name for path in copied], ["040_invoice.jpg", "041_payment.jpg"])
            self.assertEqual(Path(record.crop_image).name, "040_invoice.jpg")
            self.assertEqual([Path(path).name for path in record.supporting_crop_images], ["041_payment.jpg"])

    def test_apply_qwen_orientation_rotates_high_confidence_crop(self):
        from PIL import Image

        with tempfile.TemporaryDirectory() as temp:
            image_path = Path(temp) / "crop.png"
            image = Image.new("RGB", (12, 8), "white")
            image.putpixel((0, 0), (255, 0, 0))
            image.save(image_path)
            result = OCRResult("qwen_scan", rotate_degrees=180, orientation_confidence=0.98)

            note = _apply_qwen_orientation(image_path, result)

            with Image.open(image_path) as rotated:
                self.assertEqual(rotated.getpixel((11, 7)), (255, 0, 0))
            self.assertIn("rotated crop 180deg", note)

    def test_apply_qwen_orientation_uses_clockwise_90_convention(self):
        from PIL import Image

        with tempfile.TemporaryDirectory() as temp:
            image_path = Path(temp) / "crop.png"
            image = Image.new("RGB", (12, 8), "white")
            image.putpixel((0, 0), (255, 0, 0))
            image.save(image_path)
            result = OCRResult("qwen_scan", rotate_degrees=90, orientation_confidence=0.98)

            note = _apply_qwen_orientation(image_path, result)

            with Image.open(image_path) as rotated:
                self.assertEqual(rotated.size, (8, 12))
                self.assertEqual(rotated.getpixel((7, 0)), (255, 0, 0))
            self.assertIn("rotated crop 90deg clockwise", note)

    def test_apply_qwen_orientation_uses_clockwise_270_convention(self):
        from PIL import Image

        with tempfile.TemporaryDirectory() as temp:
            image_path = Path(temp) / "crop.png"
            image = Image.new("RGB", (12, 8), "white")
            image.putpixel((0, 0), (255, 0, 0))
            image.save(image_path)
            result = OCRResult("qwen_scan", rotate_degrees=270, orientation_confidence=0.98)

            note = _apply_qwen_orientation(image_path, result)

            with Image.open(image_path) as rotated:
                self.assertEqual(rotated.size, (8, 12))
                self.assertEqual(rotated.getpixel((0, 11)), (255, 0, 0))
            self.assertIn("rotated crop 270deg clockwise", note)

    def test_apply_qwen_orientation_does_not_rotate_low_confidence_crop(self):
        from PIL import Image

        with tempfile.TemporaryDirectory() as temp:
            image_path = Path(temp) / "crop.png"
            image = Image.new("RGB", (12, 8), "white")
            image.putpixel((0, 0), (255, 0, 0))
            image.save(image_path)
            before = image_path.read_bytes()
            result = OCRResult("qwen_scan", rotate_degrees=90, orientation_confidence=0.50)

            note = _apply_qwen_orientation(image_path, result)

            self.assertEqual(image_path.read_bytes(), before)
            self.assertIn("orientation uncertain 90deg", note)

    def test_copy_final_crops_removes_stale_generated_files(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            final_dir = root / "final_crops"
            final_dir.mkdir()
            stale = final_dir / "stale.jpg"
            stale.write_bytes(b"old")
            crop = root / "raw" / "source_d01.jpg"
            crop.parent.mkdir()
            crop.write_bytes(b"new")
            record = InvoiceRecord(
                line_no=1,
                invoice_date="2026-06-12",
                currency="MXN",
                total_amount=10,
                source_image=str(root / "source.jpg"),
                crop_image=str(crop),
            )

            copied = _copy_final_crops([record], final_dir)

            self.assertFalse(stale.exists())
            self.assertEqual(len(copied), 1)
            self.assertTrue(copied[0].exists())

    def test_processes_multi_receipt_photo_with_fake_resolver(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            sample = create_synthetic_multi_receipt(root / "multi.jpg")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                ai_visual_count_min_opencv_crops=2,
            )
            output = root / "out"
            pipeline = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=FakeResolver())

            summary = pipeline.process_path(sample)

            self.assertEqual(summary.source_images, 1)
            self.assertEqual(summary.crops, 2)
            self.assertEqual(summary.records_written, 2)
            self.assertTrue(summary.workbook_path.exists())

            wb = load_workbook(summary.workbook_path, data_only=True)
            try:
                invoices = wb[INVOICE_EXP_SHEET]
                self.assertEqual(invoices.max_row, 3)
                self.assertEqual(invoices.cell(2, 1).value, 1)
                self.assertEqual(invoices.cell(3, 1).value, 2)
                self.assertEqual(invoices.cell(2, 1).number_format, "000")
                self.assertEqual(invoices.cell(2, 9).value, "CAFE XUAN")
                self.assertEqual(invoices.cell(3, 9).value, "PANADERIA LUZ")
                self.assertEqual(invoices.cell(1, invoices.max_column).value, "Invoice link")
                self.assertEqual(invoices.cell(2, invoices.max_column).hyperlink.target, "crops/001_2026-06-12_MXN_116.00_CAFE_XUAN.jpg")
            finally:
                wb.close()

            review_crops = sorted((output / "crops").glob("*.jpg"))
            self.assertEqual(len(review_crops), 2)
            self.assertEqual(review_crops[0].name, "001_2026-06-12_MXN_116.00_CAFE_XUAN.jpg")
            self.assertEqual(review_crops[1].name, "002_2026-06-13_MXN_92.80_PANADERIA_LUZ.jpg")
            self.assertFalse((output / "review_crops").exists())

            self.assertFalse((output / CHECKED_WORKBOOK_NAME).exists())
            self.assertFalse((output / "final_crops").exists())

    def test_process_path_removes_stale_raw_crops_before_run(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            sample = create_synthetic_multi_receipt(root / "multi.jpg")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                ai_visual_count_min_opencv_crops=2,
            )
            output = root / "out"
            stale = output / "crops" / "stale.jpg"
            stale.parent.mkdir(parents=True)
            stale.write_bytes(b"old")
            pipeline = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=FakeResolver())

            pipeline.process_path(sample)

            self.assertFalse(stale.exists())
            current_crops = sorted((output / "crops").glob("*.jpg"))
            self.assertEqual(len(current_crops), 2)

    def test_noise_crop_is_deleted_and_not_written(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            sample = root / "noise.jpg"
            sample.write_bytes(b"not really a jpg")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            output = root / "out"

            summary = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=NoiseResolver()).process_path(sample)

            self.assertEqual(summary.crops, 1)
            self.assertEqual(summary.records_written, 0)
            self.assertEqual(list((output / "crops").glob("*.jpg")), [])
            wb = load_workbook(summary.workbook_path, data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].max_row, 1)
            finally:
                wb.close()

    def test_failed_qwen_receipt_crop_is_kept_for_human_review(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            sample = root / "payment.jpg"
            _write_receipt_like_image(sample)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            output = root / "out"

            summary = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=FailedQwenReceiptResolver()).process_path(sample)

            self.assertEqual(summary.crops, 1)
            self.assertEqual(summary.records_written, 1)
            crops = list((output / "crops").glob("*.jpg"))
            self.assertEqual(len(crops), 1)
            wb = load_workbook(summary.workbook_path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.max_row, 2)
                self.assertEqual(ws.cell(2, 1).value, 1)
                self.assertEqual(ws.cell(2, 9).value, "Unknown")
                headers = {cell.value: cell.column for cell in ws[1]}
                self.assertIn("needs human review", ws.cell(2, headers["System note"]).value)
            finally:
                wb.close()

    def test_resume_skips_completed_source_after_crash(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "a.jpg").write_bytes(b"not really a jpg")
            (input_dir / "b.jpg").write_bytes(b"not really a jpg b")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            output = root / "out"
            first_resolver = PathBasedResolver(crash_on_stem="b")
            pipeline = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=first_resolver)

            with self.assertRaises(RuntimeError):
                pipeline.process_path(input_dir, resume=True)

            self.assertTrue(first_resolver.calls[0].endswith("a_d01.jpg"))
            self.assertTrue(first_resolver.calls[1].endswith("b_d01.jpg"))
            self.assertTrue((output / "processing_state.json").exists())
            wb = load_workbook(reimbursement_workbook_path(output), data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].max_row, 2)
                self.assertEqual(wb[INVOICE_EXP_SHEET].cell(2, 9).value, "SELLER A")
            finally:
                wb.close()

            second_resolver = PathBasedResolver()
            resumed = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=second_resolver)
            summary = resumed.process_path(input_dir, resume=True)

            self.assertEqual(len(second_resolver.calls), 1)
            self.assertTrue(second_resolver.calls[0].endswith("b_d01.jpg"))
            self.assertEqual(summary.source_images, 2)
            wb = load_workbook(reimbursement_workbook_path(output), data_only=True)
            try:
                sellers = [wb[INVOICE_EXP_SHEET].cell(row, 9).value for row in range(2, wb[INVOICE_EXP_SHEET].max_row + 1)]
                self.assertEqual(sellers, ["SELLER A", "SELLER B"])
            finally:
                wb.close()

    def test_exact_duplicate_source_photo_is_skipped_by_hash(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "a.jpg").write_bytes(b"same photo bytes")
            (input_dir / "b.jpg").write_bytes(b"same photo bytes")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            output = root / "out"
            resolver = PathBasedResolver()
            pipeline = InvoicePipeline(settings=settings, trial=True, output_dir=output, resolver=resolver)

            summary = pipeline.process_path(input_dir, resume=True)

            self.assertEqual(len(resolver.calls), 1)
            self.assertTrue(resolver.calls[0].endswith("a_d01.jpg"))
            self.assertEqual(summary.records_written, 1)
            wb = load_workbook(reimbursement_workbook_path(output), data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].max_row, 2)
                state = json.loads((output / "processing_state.json").read_text(encoding="utf-8"))
                self.assertEqual(len(state["source_qas"]), 2)
                self.assertIn("Exact duplicate source photo skipped", state["source_qas"][1]["reason"])
            finally:
                wb.close()

    def test_crop_index_restarts_after_active_state_reset(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            output = root / "out"
            first = root / "a.jpg"
            second = root / "b.jpg"
            first.write_bytes(b"first")
            second.write_bytes(b"second")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )

            InvoicePipeline(settings=settings, trial=False, output_dir=output, resolver=PathBasedResolver()).process_path(
                first,
                resume=True,
            )
            (output / "processing_state.json").unlink()
            reimbursement_workbook_path(output).unlink()
            InvoicePipeline(settings=settings, trial=False, output_dir=output, resolver=PathBasedResolver()).process_path(
                second,
                resume=True,
            )

            wb = load_workbook(reimbursement_workbook_path(output), data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].cell(2, 1).value, 1)
            finally:
                wb.close()

    def test_source_qa_marks_visual_count_mismatch_for_review(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            sample = create_synthetic_multi_receipt(root / "multi.jpg")
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                ai_visual_count_min_opencv_crops=2,
            )
            output = root / "out"
            pipeline = InvoicePipeline(
                settings=settings,
                trial=True,
                output_dir=output,
                resolver=FakeResolver(),
                visual_counter=FakeVisualCounter(count=3),
            )

            pipeline.process_path(sample, resume=True)

            state = json.loads((output / "processing_state.json").read_text(encoding="utf-8"))
            qa = state["source_qas"][0]
            self.assertEqual(qa["ai_visual_count"], 3)
            self.assertEqual(qa["opencv_crop_count"], 2)
            self.assertTrue(qa["needs_human_review"])
            self.assertIn("AI count 3 != OpenCV crops 2", qa["reason"])

    def test_visual_count_runs_only_when_opencv_finds_four_or_more_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                ai_visual_count_min_opencv_crops=4,
            )
            counter = FakeVisualCounter(count=9)
            pipeline = InvoicePipeline(
                settings=settings,
                trial=True,
                output_dir=root / "out",
                resolver=FakeResolver(),
                visual_counter=counter,
            )

            skipped = pipeline._visual_count_for_image(root / "photo.jpg", 3)
            counted = pipeline._visual_count_for_image(root / "photo.jpg", 4)

            self.assertIsNone(skipped.count)
            self.assertIn("OpenCV crops 3 < 4", skipped.reason)
            self.assertEqual(counted.count, 9)
            self.assertEqual(counter.calls, [root / "photo.jpg"])

    def test_source_qa_marks_partial_qwen_failure_for_review(self):
        qa = _source_qa_record(
            Path("photo.jpg"),
            VisualCountResult(None, reason="count skipped"),
            3,
            3,
            failed_crop_count=1,
        )

        self.assertTrue(qa.needs_human_review)
        self.assertIn("Qwen validation failed for 1 crop", qa.reason)

    def test_qwen_only_resolver_retains_amount_when_seller_validation_fails(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            resolver = QwenOnlyResolver(settings)
            qwen = OCRResult(
                "qwen_scan",
                [
                    OCRTextLine(
                        '{"invoice_date":"","expense_category":"Food","currency":"MXN",'
                        '"total_amount":150,"expense_amount":150,"seller":"",'
                        '"remarks":"Handwritten Nota De Cuenta"}',
                        1.0,
                    )
                ],
                None,
                0.0,
                "Qwen Scan JSON failed validation: seller",
            )

            with patch.object(resolver.qwen, "recognize", return_value=qwen):
                resolved = resolver.scan(root / "crop.jpg")

            self.assertEqual(resolved.record.total_amount, 150)
            self.assertEqual(resolved.record.invoice_date, "")
            self.assertEqual(resolved.record.expense_category, "Food")
            self.assertIn("partial result retained", resolved.record.remarks)

    def test_missing_date_remains_blank_next_to_dated_receipt(self):
        records = [
            InvoiceRecord(invoice_date="2026-06-12", seller="Printed", total_amount=100),
            InvoiceRecord(invoice_date="", seller="Handwritten", total_amount=150, remarks="Qwen Scan used"),
            InvoiceRecord(invoice_date="2026-06-13", seller="Printed Later", total_amount=80),
        ]

        _fill_missing_dates_from_neighbors(records)

        self.assertEqual(records[1].invoice_date, "")
        self.assertNotIn("Missing date filled", records[1].remarks)

    def test_missing_date_does_not_fall_back_to_source_photo_date(self):
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "photo.jpg"
            source.write_bytes(b"jpg")
            records = [
                InvoiceRecord(invoice_date="", seller="Handwritten", total_amount=150, source_image=str(source)),
            ]

            _fill_missing_dates_from_neighbors(records)

            self.assertEqual(records[0].invoice_date, "")
            self.assertNotIn("Missing date filled", records[0].remarks)

    def test_missing_date_uses_unique_date_from_same_photo(self):
        source = "telegram/photo.jpg"
        dated = InvoiceRecord(invoice_date="2026-06-12", seller="Printed", total_amount=100, source_image=source)
        undated = InvoiceRecord(invoice_date="", seller="Handwritten", total_amount=150, source_image=source)

        _fill_missing_dates_from_neighbors([dated, undated], targets=[undated])

        self.assertEqual(undated.invoice_date, "2026-06-12")
        self.assertIn("Missing date filled from same photo: 2026-06-12", undated.remarks)

    def test_missing_date_uses_nearest_unambiguous_batch_date_for_new_record_only(self):
        historical_blank = InvoiceRecord(invoice_date="", seller="Old", total_amount=50)
        nearby = InvoiceRecord(invoice_date="2026-06-18", seller="Printed", total_amount=100)
        new_blank = InvoiceRecord(invoice_date="", seller="Handwritten", total_amount=150)

        _fill_missing_dates_from_neighbors(
            [historical_blank, nearby, new_blank],
            targets=[new_blank],
        )

        self.assertEqual(historical_blank.invoice_date, "")
        self.assertEqual(new_blank.invoice_date, "2026-06-18")
        self.assertIn("Missing date filled from nearby batch receipt: 2026-06-18", new_blank.remarks)


class FakeResolver:
    def __init__(self) -> None:
        self.index = 0

    def scan(self, image_path: Path) -> ResolvedScan:
        records = [
            InvoiceRecord(
                invoice_date="2026-06-12",
                seller="CAFE XUAN",
                currency="MXN",
                total_amount=116,
                expense_category="Food",
                expense_amount=100,
                vat_amount=16,
            ),
            InvoiceRecord(
                invoice_date="2026-06-13",
                seller="PANADERIA LUZ",
                currency="MXN",
                total_amount=92.8,
                expense_category="Food",
                expense_amount=80,
                vat_amount=12.8,
            ),
        ]
        record = records[self.index]
        self.index += 1
        paddle = OCRResult("paddleocr", [OCRTextLine(record.seller, 0.9)], record, 0.9)
        easy = OCRResult("easyocr", [OCRTextLine(record.seller, 0.9)], record, 0.9)
        return ResolvedScan(record, paddle, easy, None, False, "local OCR agreement")


class FakeVisualCounter:
    def __init__(self, count: int | None) -> None:
        self.count_value = count
        self.calls: list[Path] = []

    def count(self, image_path: Path) -> VisualCountResult:
        self.calls.append(image_path)
        return VisualCountResult(self.count_value, confidence=0.9, reason="test count")


class PathBasedResolver:
    def __init__(self, crash_on_stem: str | None = None) -> None:
        self.crash_on_stem = crash_on_stem
        self.calls: list[str] = []

    def scan(self, image_path: Path) -> ResolvedScan:
        self.calls.append(image_path.name)
        if self.crash_on_stem and (
            image_path.stem.startswith(self.crash_on_stem) or f"_{self.crash_on_stem}_" in image_path.stem
        ):
            raise RuntimeError("simulated crash")
        label = "A" if "_a_" in image_path.stem or image_path.stem.startswith("a") else "B"
        record = InvoiceRecord(
            invoice_date="2026-06-12",
            seller=f"SELLER {label}",
            currency="MXN",
            total_amount=100 if label == "A" else 200,
            expense_amount=100 if label == "A" else 200,
        )
        paddle = OCRResult("paddleocr", [OCRTextLine(record.seller, 0.9)], record, 0.9)
        easy = OCRResult("easyocr", [OCRTextLine(record.seller, 0.9)], record, 0.9)
        return ResolvedScan(record, paddle, easy, None, False, "local OCR agreement")


class NoiseResolver:
    def scan(self, image_path: Path) -> ResolvedScan:
        record = InvoiceRecord(seller="Unknown", total_amount=0)
        result = OCRResult("qwen_scan", [], record, 1.0)
        return ResolvedScan(record, result, result, result, True, "qwen scan only")


class FailedQwenReceiptResolver:
    def scan(self, image_path: Path) -> ResolvedScan:
        record = InvoiceRecord(seller="Unknown", total_amount=0, remarks="Qwen Scan failed: total_amount; needs human review")
        result = OCRResult(
            "qwen_scan",
            [OCRTextLine('{"seller":"Flap","total_amount":0,"remarks":"Importe visible but not parsed"}', 1.0)],
            None,
            0.0,
            "Qwen Scan JSON failed validation: total_amount",
        )
        empty = OCRResult("local_ocr_disabled")
        return ResolvedScan(record, empty, empty, result, False, "qwen scan failed")


def _write_receipt_like_image(path: Path) -> None:
    from PIL import Image, ImageDraw

    image = Image.new("RGB", (1200, 1600), "white")
    draw = ImageDraw.Draw(image)
    draw.text((120, 140), "COMISION FEDERAL DE ELECTRICIDAD", fill="black")
    draw.text((120, 260), "COMPROBANTE DE PAGO POR INTERNET", fill="black")
    draw.text((120, 380), "Fecha de pago: 11/07/2026", fill="black")
    draw.text((120, 500), "Importe: $58.00", fill="black")
    draw.text((120, 620), "CINCUENTA Y OCHO PESOS 00/100 MXP", fill="black")
    image.save(path)


if __name__ == "__main__":
    unittest.main()
