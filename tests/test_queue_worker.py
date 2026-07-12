import tempfile
import unittest
import json
from datetime import datetime
from pathlib import Path

from invoice_system.config import Settings
from invoice_system.grouping import mark_source_for_group_if_armed, arm_next_group
from invoice_system.models import InvoiceRecord, PipelineSummary
from invoice_system.queue_worker import (
    DONE,
    FAILED_RETRYABLE,
    PENDING,
    QueueItem,
    QueueState,
    _workbook_records_for_source,
    block_rollback_for_manual_edit,
    discover_and_enqueue,
    enqueue_photo,
    load_queue_state,
    process_user_queue_once,
    prepare_last_photo_rescan,
    queue_totals_for_day,
    reset_active_user_workspace,
    retry_failed,
    rollback_last_photo,
    save_queue_state,
    telegram_photo_quality,
    telegram_user_day_dir,
    telegram_user_output_dir,
    telegram_user_queue_path,
    telegram_user_workbook,
)
from invoice_system.reimbursement_excel import ReimbursementWorkbook
from invoice_system.reimbursement_excel import load_reimbursement_records


class QueueWorkerTests(unittest.TestCase):
    def test_telegram_photo_quality_classifies_sd_hd_and_unknown(self):
        self.assertEqual(telegram_photo_quality(960, 1280), "sd")
        self.assertEqual(telegram_photo_quality(1920, 2560), "hd")
        self.assertEqual(telegram_photo_quality(0, 0), "unknown")

    def test_enqueue_photo_writes_per_user_queue(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "photo.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")

            summary = enqueue_photo(settings, 123, photo)

            self.assertEqual(summary.pending, 1)
            self.assertTrue(telegram_user_queue_path(settings, 123).exists())
            self.assertFalse(telegram_user_queue_path(settings, 456).exists())

    def test_enqueue_photo_persists_telegram_quality_metadata(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "photo.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")

            enqueue_photo(
                settings,
                123,
                photo,
                image_width=960,
                image_height=1280,
                file_size=321,
                upload_quality="sd",
            )

            item = load_queue_state(telegram_user_queue_path(settings, 123)).items[0]
            self.assertEqual((item.image_width, item.image_height, item.file_size), (960, 1280, 321))
            self.assertEqual(item.upload_quality, "sd")

    def test_old_queue_item_without_quality_fields_loads_as_unknown(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "queue_state.json"
            path.write_text(json.dumps({"items": [{"path": "old.jpg", "status": "done"}]}), encoding="utf-8")

            item = load_queue_state(path).items[0]

            self.assertEqual(item.upload_quality, "unknown")
            self.assertEqual(item.detected_receipt_count, 0)

    def test_discover_and_enqueue_scans_user_root_only(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            own = telegram_user_day_dir(settings, 123) / "own.jpg"
            other = telegram_user_day_dir(settings, 456) / "other.jpg"
            own.parent.mkdir(parents=True)
            other.parent.mkdir(parents=True)
            own.write_bytes(b"jpg")
            other.write_bytes(b"jpg")

            summary = discover_and_enqueue(settings, 123)

            self.assertEqual(summary.pending, 1)
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            self.assertEqual(Path(state.items[0].path).name, "own.jpg")

    def test_worker_marks_failure_retryable_and_continues(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            good = telegram_user_day_dir(settings, 123) / "good.jpg"
            bad = telegram_user_day_dir(settings, 123) / "bad.jpg"
            good.parent.mkdir(parents=True)
            good.write_bytes(b"jpg")
            bad.write_bytes(b"jpg")

            summary = process_user_queue_once(settings, 123, pipeline_factory=FakeQueuePipeline)

            self.assertEqual(summary.done, 1)
            self.assertEqual(summary.failed, 1)
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            statuses = {Path(item.path).name: item.status for item in state.items}
            self.assertEqual(statuses["good.jpg"], DONE)
            self.assertEqual(statuses["bad.jpg"], FAILED_RETRYABLE)

    def test_retry_failed_moves_only_failed_to_pending(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            good = telegram_user_day_dir(settings, 123) / "good.jpg"
            bad = telegram_user_day_dir(settings, 123) / "bad.jpg"
            good.parent.mkdir(parents=True)
            good.write_bytes(b"jpg")
            bad.write_bytes(b"jpg")
            process_user_queue_once(settings, 123, pipeline_factory=FakeQueuePipeline)

            count, summary = retry_failed(settings, 123)

            self.assertEqual(count, 1)
            self.assertEqual(summary.pending, 1)
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            statuses = {Path(item.path).name: item.status for item in state.items}
            self.assertEqual(statuses["good.jpg"], DONE)
            self.assertEqual(statuses["bad.jpg"], PENDING)

    def test_worker_notifies_after_each_item(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            good = telegram_user_day_dir(settings, 123) / "good.jpg"
            bad = telegram_user_day_dir(settings, 123) / "bad.jpg"
            good.parent.mkdir(parents=True)
            good.write_bytes(b"jpg")
            bad.write_bytes(b"jpg")
            events = []

            process_user_queue_once(
                settings,
                123,
                pipeline_factory=FakeQueuePipeline,
                item_callback=lambda user_id, item, records: events.append((user_id, Path(item.path).name, item.status)),
            )

            self.assertEqual(set(events), {(123, "good.jpg", DONE), (123, "bad.jpg", FAILED_RETRYABLE)})

    def test_worker_forced_group_merges_next_upload_to_one_row(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "group.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"group")
            output_dir = telegram_user_output_dir(settings, 123)
            arm_next_group(output_dir)
            self.assertTrue(mark_source_for_group_if_armed(output_dir, photo))
            enqueue_photo(settings, 123, photo)

            summary = process_user_queue_once(settings, 123, pipeline_factory=FakeGroupedQueuePipeline)

            self.assertEqual(summary.done, 1)
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            item = state.items[0]
            self.assertEqual(item.status, DONE)
            self.assertEqual(item.row_count, 1)
            self.assertEqual(item.total_amount, 622.60)
            records = load_reimbursement_records(telegram_user_workbook(settings, 123))
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].line_no, 44)
            self.assertEqual(records[0].total_amount, 622.60)

    def test_workbook_records_for_source_matches_by_crop_id_not_row_position(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            output_dir = telegram_user_output_dir(settings, 123)
            source = telegram_user_day_dir(settings, 123) / "photo.jpg"
            source.parent.mkdir(parents=True)
            source.write_bytes(b"jpg")
            crop_a = output_dir / "crops" / "039_2026-05-09_MXN_25.00_DESIGNA.jpg"
            crop_b = output_dir / "crops" / "040_2026-06-18_MXN_165.00_Salsa.jpg"
            crop_a.parent.mkdir(parents=True)
            crop_a.write_bytes(b"crop")
            crop_b.write_bytes(b"crop")
            (output_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "source_image": str(source),
                                "crop_image": str(crop_a),
                                "supporting_crop_images": [],
                            },
                            {
                                "source_image": str(source),
                                "crop_image": str(crop_b),
                                "supporting_crop_images": [],
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )
            workbook_records = [
                InvoiceRecord(line_no=1, crop_image=str(output_dir / "crops" / "001_old.jpg"), total_amount=10),
                InvoiceRecord(line_no=40, crop_image=str(crop_b), total_amount=165),
            ]

            matched = _workbook_records_for_source(output_dir, source, workbook_records)

            self.assertEqual([record.line_no for record in matched], [40])

    def test_reset_active_user_workspace_archives_inbound_output_and_clears_queue(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "photo.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")
            output_dir = telegram_user_output_dir(settings, 123)
            final_crop = output_dir / "final_crops" / "001.jpg"
            final_crop.parent.mkdir(parents=True)
            final_crop.write_bytes(b"crop")
            enqueue_photo(settings, 123, photo)

            summary = reset_active_user_workspace(settings, 123)

            self.assertTrue(summary.archive_dir.exists())
            self.assertFalse(photo.exists())
            self.assertFalse(final_crop.exists())
            self.assertEqual(load_queue_state(telegram_user_queue_path(settings, 123)).items, [])
            archived_files = list(summary.archive_dir.rglob("*"))
            self.assertTrue(any(path.name == "photo.jpg" for path in archived_files))
            self.assertTrue(any(path.name == "001.jpg" for path in archived_files))

    def test_queue_totals_for_day_filters_checked_records_to_today_sources(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            output_dir = telegram_user_output_dir(settings, 123)
            review = output_dir / "review_crops"
            review.mkdir(parents=True)
            old_crop = review / "001_2026-06-10_MXN_100.00_Old.jpg"
            today_crop = review / "082_2026-06-27_MXN_215.00_Restaurante.jpg"
            old_crop.write_bytes(b"old")
            today_crop.write_bytes(b"today")
            old_photo = telegram_user_day_dir(settings, 123, datetime(2026, 6, 26, 10, 0)) / "old.jpg"
            today_photo = telegram_user_day_dir(settings, 123, datetime(2026, 6, 27, 10, 0)) / "today.jpg"
            old_photo.parent.mkdir(parents=True)
            today_photo.parent.mkdir(parents=True)
            old_photo.write_bytes(b"old")
            today_photo.write_bytes(b"today")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-10", expense_category="Food", seller="Old", total_amount=100, crop_image=str(old_crop)),
                    InvoiceRecord(line_no=82, invoice_date="2026-06-27", expense_category="Food", seller="Restaurante", total_amount=215, crop_image=str(today_crop)),
                ]
            )
            save_queue_state(
                telegram_user_queue_path(settings, 123),
                QueueState(
                    [
                        QueueItem(path=str(old_photo), status=DONE, updated_at="2026-06-26T11:00:00", row_count=1, total_amount=100, category_totals={"Food": 100}),
                        QueueItem(path=str(today_photo), status=DONE, updated_at="2026-06-27T11:00:00", row_count=1, total_amount=215, category_totals={"Food": 215}),
                    ]
                ),
            )
            (output_dir / "processing_state.json").write_text(
                '{"records":[{"source_image":"'
                + str(old_photo).replace("\\", "\\\\")
                + '","crop_image":"'
                + str(old_crop).replace("\\", "\\\\")
                + '"},{"source_image":"'
                + str(today_photo).replace("\\", "\\\\")
                + '","crop_image":"'
                + str(today_crop).replace("\\", "\\\\")
                + '"}]}',
                encoding="utf-8",
            )

            totals = queue_totals_for_day(settings, 123, "2026-06-27")

            self.assertEqual(totals.record_count, 1)
            self.assertEqual(totals.total_amount, 215)
            self.assertEqual(totals.category_totals, {"Food": 215})

    def test_rollback_last_pending_photo_removes_queue_and_source_only(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "pending.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([QueueItem(path=str(photo), status=PENDING)]))

            summary = rollback_last_photo(settings, 123)

            self.assertEqual(summary.status, PENDING)
            self.assertEqual(summary.queue_items_removed, 1)
            self.assertFalse(photo.exists())
            self.assertEqual(load_queue_state(telegram_user_queue_path(settings, 123)).items, [])
            self.assertFalse((telegram_user_output_dir(settings, 123) / "processing_state.json").exists())

    def test_rollback_last_processing_photo_is_rejected_without_mutation(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "processing.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([QueueItem(path=str(photo), status="processing")]))

            with self.assertRaisesRegex(RuntimeError, "still processing"):
                rollback_last_photo(settings, 123)

            self.assertTrue(photo.exists())
            self.assertEqual(load_queue_state(telegram_user_queue_path(settings, 123)).items[0].status, "processing")

    def test_rollback_pending_photo_is_not_blocked_by_older_manual_edit_marker(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            photo = telegram_user_day_dir(settings, 123) / "pending.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([QueueItem(path=str(photo), status=PENDING)]))

            block_rollback_for_manual_edit(settings, 123, "/change was used")

            summary = rollback_last_photo(settings, 123)

            self.assertEqual(summary.status, PENDING)
            self.assertFalse(photo.exists())
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            self.assertTrue(state.rollback_blocked)
            self.assertEqual(len(state.items), 0)

    def test_rollback_done_photo_uses_latest_photo_even_if_target_trace_was_manually_edited(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            output_dir = telegram_user_output_dir(settings, 123)
            review_dir = output_dir / "review_crops"
            review_dir.mkdir(parents=True)
            photo = telegram_user_day_dir(settings, 123, datetime(2026, 7, 1, 11, 0)) / "edited.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"edited")
            crop = review_dir / "002_2026-07-01_MXN_20.00_Bad.jpg"
            crop.write_bytes(b"crop")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=2, invoice_date="2026-07-01", expense_category="Food", seller="Bad", total_amount=20, crop_image=str(crop), source_image=str(photo))]
            )
            workbook = telegram_user_workbook(settings, 123)
            from openpyxl import load_workbook

            wb = load_workbook(workbook)
            try:
                ws = wb["Invoice exp"]
                ws.cell(2, 2).value = "ok"
                wb.save(workbook)
            finally:
                wb.close()
            (output_dir / "processing_state.json").write_text(
                json.dumps({"records": [{"source_image": str(photo), "crop_image": str(crop)}]}),
                encoding="utf-8",
            )
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([QueueItem(path=str(photo), status=DONE, row_count=1)]))

            summary = rollback_last_photo(settings, 123)

            self.assertEqual(summary.trace_ids, ("002",))
            self.assertFalse(photo.exists())
            self.assertFalse(crop.exists())

    def test_rollback_last_done_photo_removes_records_crops_and_reuses_trace_ids(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            output_dir = telegram_user_output_dir(settings, 123)
            review_dir = output_dir / "review_crops"
            crop_dir = output_dir / "crops"
            final_food_dir = output_dir / "final_crops" / "food"
            for directory in (review_dir, crop_dir, final_food_dir):
                directory.mkdir(parents=True)

            old_photo = telegram_user_day_dir(settings, 123, datetime(2026, 7, 1, 10, 0)) / "old.jpg"
            bad_photo = telegram_user_day_dir(settings, 123, datetime(2026, 7, 1, 11, 0)) / "bad.jpg"
            old_photo.parent.mkdir(parents=True)
            old_photo.write_bytes(b"old")
            bad_photo.write_bytes(b"bad")
            old_crop = review_dir / "001_2026-07-01_MXN_100.00_Cafe.jpg"
            bad_crop_2 = review_dir / "002_2026-07-01_MXN_20.00_Bad.jpg"
            bad_crop_3 = review_dir / "003_2026-07-01_MXN_30.00_Bad.jpg"
            raw_bad_crop = crop_dir / "003_2026-07-01_MXN_30.00_Bad.jpg"
            final_bad_crop = final_food_dir / "002_trace002_2026-07-01_MXN_20.00_Bad.jpg"
            for path, payload in (
                (old_crop, b"old crop"),
                (bad_crop_2, b"bad crop 2"),
                (bad_crop_3, b"bad crop 3"),
                (raw_bad_crop, b"raw bad"),
                (final_bad_crop, b"final bad"),
            ):
                path.write_bytes(payload)

            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-07-01", expense_category="Food", seller="Cafe", total_amount=100, crop_image=str(old_crop), source_image=str(old_photo)),
                    InvoiceRecord(line_no=2, invoice_date="2026-07-01", expense_category="Food", seller="Bad", total_amount=20, crop_image=str(bad_crop_2), source_image=str(bad_photo)),
                    InvoiceRecord(line_no=3, invoice_date="2026-07-01", expense_category="Food", seller="Bad", total_amount=30, crop_image=str(bad_crop_3), source_image=str(bad_photo)),
                ]
            )
            (output_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "version": 1,
                        "completed_sources": [str(old_photo), str(bad_photo)],
                        "source_hashes": ["oldhash", _sha256_bytes(b"bad")],
                        "records": [
                            {"source_image": str(old_photo), "crop_image": str(old_crop)},
                            {"source_image": str(bad_photo), "crop_image": str(bad_crop_2), "supporting_crop_images": [str(bad_crop_3)]},
                        ],
                        "audits": [
                            {"source_image": str(bad_photo), "crop_image": str(raw_bad_crop)},
                        ],
                        "source_qas": [
                            {"source_image": str(old_photo), "reason": "ok"},
                            {"source_image": str(bad_photo), "reason": "bad"},
                        ],
                        "next_crop_id": 4,
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            save_queue_state(
                telegram_user_queue_path(settings, 123),
                QueueState(
                    [
                        QueueItem(path=str(old_photo), status=DONE, row_count=1),
                        QueueItem(path=str(bad_photo), status=DONE, row_count=2),
                    ]
                ),
            )

            summary = rollback_last_photo(settings, 123)

            self.assertEqual(summary.trace_ids, ("002", "003"))
            self.assertEqual(summary.manual_rows_deleted, 2)
            self.assertFalse(bad_photo.exists())
            self.assertFalse(bad_crop_2.exists())
            self.assertFalse(bad_crop_3.exists())
            self.assertFalse(raw_bad_crop.exists())
            self.assertFalse(final_bad_crop.exists())
            self.assertTrue(old_photo.exists())
            self.assertTrue(old_crop.exists())
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            self.assertEqual(len(state.items), 1)
            self.assertEqual(Path(state.items[0].path), old_photo)
            processing = json.loads((output_dir / "processing_state.json").read_text(encoding="utf-8"))
            self.assertEqual(processing["completed_sources"], [str(old_photo)])
            self.assertEqual(len(processing["records"]), 1)
            self.assertEqual(processing["source_qas"], [{"source_image": str(old_photo), "reason": "ok"}])
            self.assertEqual(processing["next_crop_id"], 2)
            workbook_records = load_reimbursement_records(telegram_user_workbook(settings, 123))
            self.assertEqual([record.seller for record in workbook_records], ["Cafe"])

    def test_prepare_last_photo_rescan_preserves_source_and_requeues(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            output_dir = telegram_user_output_dir(settings, 123)
            review_dir = output_dir / "review_crops"
            review_dir.mkdir(parents=True)
            photo = telegram_user_day_dir(settings, 123, datetime(2026, 7, 2, 12, 0)) / "last.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"source")
            crop = review_dir / "026_2026-07-02_MXN_50.00_Cordia.jpg"
            crop.write_bytes(b"crop")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=26, invoice_date="2026-07-02", expense_category="Food", seller="Cordia", total_amount=50, crop_image=str(crop), source_image=str(photo))]
            )
            (output_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "completed_sources": [str(photo)],
                        "source_hashes": [_sha256_bytes(b"source")],
                        "records": [{"source_image": str(photo), "crop_image": str(crop)}],
                        "audits": [{"source_image": str(photo), "crop_image": str(crop)}],
                        "source_qas": [{"source_image": str(photo), "reason": "old"}],
                        "next_crop_id": 27,
                    }
                ),
                encoding="utf-8",
            )
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([QueueItem(path=str(photo), status=DONE, row_count=1)]))

            summary = prepare_last_photo_rescan(settings, 123)

            self.assertTrue(photo.exists())
            self.assertFalse(crop.exists())
            self.assertEqual(summary.trace_ids, ("026",))
            self.assertEqual(summary.manual_rows_deleted, 1)
            state = load_queue_state(telegram_user_queue_path(settings, 123))
            self.assertEqual(state.items[0].status, PENDING)
            processing = json.loads((output_dir / "processing_state.json").read_text(encoding="utf-8"))
            self.assertEqual(processing["records"], [])
            self.assertEqual(processing["completed_sources"], [])
            self.assertEqual(processing["next_crop_id"], 1)


class FakeQueuePipeline:
    def __init__(self, settings: Settings, trial: bool, output_dir: Path | None):
        self.output_dir = output_dir or Path(".")

    def process_path(self, input_path: Path | None = None, resume: bool = False):
        if input_path and input_path.name == "bad.jpg":
            raise RuntimeError("simulated OCR failure")
        workbook = self.output_dir / "Invoice_Output.xlsx"
        workbook.parent.mkdir(parents=True, exist_ok=True)
        workbook.write_bytes(b"xlsx")
        return PipelineSummary(1, 1, 1, workbook)


class FakeGroupedQueuePipeline:
    def __init__(self, settings: Settings, trial: bool, output_dir: Path | None):
        self.output_dir = output_dir or Path(".")

    def process_path(self, input_path: Path | None = None, resume: bool = False):
        from PIL import Image

        if input_path is None:
            raise RuntimeError("missing input")
        workbook = self.output_dir / "报销明细_2026_xlsx.xlsx"
        crops = self.output_dir / "crops"
        crops.mkdir(parents=True, exist_ok=True)
        detail_crop = crops / "044_2026-05-09_MXN_566.00_SUSHI_ROLL.jpg"
        card_crop = crops / "045_2026-05-09_MXN_622.60_SUSHI_ROLL_CARD.jpg"
        Image.new("RGB", (120, 240), "white").save(detail_crop)
        Image.new("RGB", (100, 220), "white").save(card_crop)
        ReimbursementWorkbook(workbook).write_records(
            [
                InvoiceRecord(line_no=44, invoice_date="2026-05-09", expense_category="Food", currency="MXN", total_amount=566, seller="SUSHI ROLL", crop_image=str(detail_crop), source_image=str(input_path)),
                InvoiceRecord(line_no=45, invoice_date="2026-05-09", expense_category="Food", currency="MXN", total_amount=622.60, tips=56.60, seller="SUSHI ROLL MIFEL", contents="venta con propina", crop_image=str(card_crop), source_image=str(input_path)),
            ]
        )
        (self.output_dir / "processing_state.json").write_text(
            json.dumps(
                {
                    "completed_sources": [str(input_path)],
                    "records": [
                        {"source_image": str(input_path), "crop_image": str(detail_crop), "supporting_crop_images": []},
                        {"source_image": str(input_path), "crop_image": str(card_crop), "supporting_crop_images": []},
                    ],
                    "audits": [],
                    "source_qas": [],
                    "next_crop_id": 46,
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        return PipelineSummary(1, 2, 2, workbook)


def _settings(root: Path) -> Settings:
    return Settings(
        root=root,
        inbound_dir=root / "data" / "inbound",
        trial_dir=root / "data" / "trial",
        output_dir=root / "data" / "output",
        baseline_dir=root / "data" / "baseline",
        telegram_allowed_user_ids=frozenset({123, 456}),
    )


def _sha256_bytes(payload: bytes) -> str:
    import hashlib

    return hashlib.sha256(payload).hexdigest()


if __name__ == "__main__":
    unittest.main()
