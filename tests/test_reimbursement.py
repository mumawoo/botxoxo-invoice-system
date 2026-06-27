import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from invoice_system.config import Settings
from invoice_system.models import InvoiceRecord
from invoice_system.queue_worker import telegram_user_output_dir, telegram_user_workbook
from invoice_system.reimbursement import (
    reimbursement_path,
    submit_unsubmitted,
    submitted_batches_text,
    sync_reimbursement_records,
    unsubmitted_summary,
)
from invoice_system.reimbursement_excel import CHECKED_WORKBOOK_NAME, INVOICE_EXP_SHEET, ReimbursementWorkbook


class ReimbursementTests(unittest.TestCase):
    def test_report_summarizes_unsubmitted_by_category(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            _write_reimbursement_workbook(settings, 123)

            summary = unsubmitted_summary(settings, 123)

            self.assertEqual(summary.record_count, 2)
            self.assertEqual(summary.total_amount, 300)
            self.assertEqual(summary.category_totals, {"Food": 100, "Gas": 200})

    def test_submit_archives_active_excel_and_resets_batch(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            _write_reimbursement_workbook(settings, 123)
            output_dir = telegram_user_output_dir(settings, 123)
            index_state = output_dir / "crop_index_state.json"
            index_state.write_text('{"version": 2, "next_crop_id": 2601022}', encoding="utf-8")

            result = submit_unsubmitted(settings, 123)

            self.assertIsNotNone(result)
            self.assertEqual(result.record_count, 2)
            self.assertTrue(result.archived_excel.exists())
            self.assertEqual(result.archived_excel.name, CHECKED_WORKBOOK_NAME)
            self.assertTrue((result.archived_crops / "food" / "001_2026-06-12_MXN_100.00_Cafe.jpg").exists())
            self.assertTrue((result.archived_crops / "other" / "002_2026-06-13_MXN_200.00_Pemex.jpg").exists())
            self.assertTrue(result.archived_manual_excel.exists())
            self.assertTrue((result.archived_review_crops / "001_2026-06-12_MXN_100.00_Cafe.jpg").exists())
            self.assertFalse(telegram_user_workbook(settings, 123).exists())
            self.assertFalse((output_dir / "final_crops").exists())
            self.assertFalse((output_dir / "review_crops").exists())
            self.assertFalse(index_state.exists())
            self.assertIn("SUB-", result.batch_id)
            self.assertIn(result.batch_id, submitted_batches_text(settings, 123))

            archived = load_workbook(result.archived_excel, data_only=True)
            try:
                self.assertIn(INVOICE_EXP_SHEET, archived.sheetnames)
                self.assertEqual(archived[INVOICE_EXP_SHEET].cell(1, 8).value, "Merchant")
                self.assertEqual(archived[INVOICE_EXP_SHEET].cell(2, 8).value, "Cafe")
            finally:
                archived.close()

    def test_sync_creates_status_workbook_without_duplicating_records(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = _settings(Path(temp))
            _write_reimbursement_workbook(settings, 123)

            sync_reimbursement_records(settings, 123)
            sync_reimbursement_records(settings, 123)

            wb = load_workbook(reimbursement_path(settings, 123), data_only=True)
            try:
                self.assertIn("Batches", wb.sheetnames)
                self.assertEqual(wb["Batches"].max_row, 1)
            finally:
                wb.close()


def _write_reimbursement_workbook(settings: Settings, user_id: int) -> None:
    output_dir = telegram_user_output_dir(settings, user_id)
    workbook = telegram_user_workbook(settings, user_id)
    crop = output_dir / "review_crops" / "001_2026-06-12_MXN_100.00_Cafe.jpg"
    crop2 = output_dir / "review_crops" / "002_2026-06-13_MXN_200.00_Pemex.jpg"
    crop.parent.mkdir(parents=True, exist_ok=True)
    crop.write_bytes(b"jpg")
    crop2.write_bytes(b"jpg")
    records = [
        InvoiceRecord(line_no=1, invoice_date="2026-06-12", seller="Cafe", expense_category="Food", total_amount=100, crop_image=str(crop)),
        InvoiceRecord(line_no=2, invoice_date="2026-06-13", seller="Pemex", expense_category="Gas", total_amount=200, crop_image=str(crop2)),
    ]
    ReimbursementWorkbook(workbook).write_records(records)
    (output_dir / "processing_state.json").write_text("{}", encoding="utf-8")


def _settings(root: Path) -> Settings:
    return Settings(
        root=root,
        inbound_dir=root / "data" / "inbound",
        trial_dir=root / "data" / "trial",
        output_dir=root / "data" / "output",
        baseline_dir=root / "data" / "baseline",
    )


if __name__ == "__main__":
    unittest.main()
