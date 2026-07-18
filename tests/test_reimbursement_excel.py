import tempfile
import unittest
import json
from datetime import date, timedelta
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from invoice_system.fx_rates import ExchangeRate
from invoice_system.models import InvoiceRecord
from invoice_system.reimbursement_excel import (
    CHECKED_WORKBOOK_NAME,
    FINAL_CROPS_MANIFEST,
    FOOD_EXP_SHEET,
    INVOICE_EXP_SHEET,
    OTHER_EXP_SHEET,
    REIMBURSEMENT_WORKBOOK_NAME,
    SUMMARY_SHEET,
    TIPS_MXN_HEADER,
    VAT_MXN_HEADER,
    ReimbursementWorkbook,
    assign_available_line_numbers,
    apply_reimbursement_group,
    build_checked_outputs,
    change_reimbursement_record,
    clear_generated_crops,
    confirm_review_repair,
    corrected_crop_names,
    focus_reimbursement_workbook,
    load_reimbursement_records,
    initialize_review_architecture,
    preview_reimbursement_group,
    preview_review_repair,
    reverse_review_sync_for_source,
    rerun_checked_from_finance_edits,
    sync_source_to_review,
)


def _date_text(value):
    if hasattr(value, "date"):
        value = value.date()
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


class ReimbursementExcelTests(unittest.TestCase):
    def test_source_sync_appends_only_new_trace_and_preserves_blank_status_manual_edits(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / REIMBURSEMENT_WORKBOOK_NAME
            old_crop = root / "crops" / "001_2026-10-05_MXN_50.50_Cafe.jpg"
            new_crop = root / "crops" / "002_2026-07-02_MXN_80.00_Pemex.jpg"
            old_source = root / "old.jpg"
            new_source = root / "new.jpg"
            old = InvoiceRecord(
                line_no=1, invoice_date="2026-10-05", seller="Cafe", expense_category="Food",
                currency="MXN", total_amount=50.5, crop_image=str(old_crop), source_image=str(old_source),
            )
            new = InvoiceRecord(
                line_no=2, invoice_date="2026-07-02", seller="Pemex", expense_category="Gas",
                currency="MXN", total_amount=80, crop_image=str(new_crop), source_image=str(new_source),
            )
            ReimbursementWorkbook(workbook).write_records([old])
            wb = load_workbook(workbook)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 3).value = "2026-05-10"
                ws.cell(2, 9).value = "Manual Cafe Name"
                self.assertIsNone(ws.cell(2, 2).value)
                wb.save(workbook)
            finally:
                wb.close()

            sync_source_to_review(root, new_source, [new], [old, new])

            wb = load_workbook(workbook, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.max_row, 3)
                self.assertEqual(_date_text(ws.cell(2, 3).value), "2026-05-10")
                self.assertEqual(ws.cell(2, 9).value, "Manual Cafe Name")
                self.assertEqual(ws.cell(3, 14).value, "002")
                self.assertEqual(ws.cell(3, 9).value, "Pemex")
            finally:
                wb.close()

    def test_source_sync_payment_updates_only_financial_fields_and_is_reversible(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / REIMBURSEMENT_WORKBOOK_NAME
            detail_crop = root / "crops" / "037_2026-05-09_MXN_566.00_Sushi_Roll.jpg"
            payment_crop = root / "crops" / "163_2026-05-09_MXN_622.60_Sushi_Roll_card.jpg"
            detail_source = root / "detail.jpg"
            payment_source = root / "payment.jpg"
            detail = InvoiceRecord(
                line_no=37, invoice_date="2026-05-09", seller="SUSHI ROLL", contents="Restaurant detail receipt",
                expense_category="Food", currency="MXN", total_amount=566, vat_amount=78,
                crop_image=str(detail_crop), source_image=str(detail_source), report_components=True,
            )
            payment = InvoiceRecord(
                line_no=163, invoice_date="2026-05-09", seller="SUSHI ROLL", contents="BBVA tarjeta pago propina",
                expense_category="Food", currency="MXN", total_amount=622.60, tips=56.60,
                crop_image=str(payment_crop), source_image=str(payment_source), report_components=True,
            )
            ReimbursementWorkbook(workbook).write_records([detail])

            result = sync_source_to_review(root, payment_source, [payment], [detail, payment])

            self.assertEqual(result.updated_trace_ids, ("037",))
            self.assertEqual(result.appended_records, ())
            loaded = load_reimbursement_records(workbook)
            self.assertEqual(len(loaded), 1)
            self.assertEqual([Path(value).name for value in loaded[0].supporting_crop_images], [payment_crop.name])
            self.assertIn("supporting crop 163 kept", loaded[0].remarks)
            wb = load_workbook(workbook, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.max_row, 2)
                self.assertEqual(_date_text(ws.cell(2, headers["Date"]).value), "2026-05-09")
                self.assertEqual(ws.cell(2, headers["Merchant"]).value, "SUSHI ROLL")
                self.assertEqual(ws.cell(2, headers["Accounting Category"]).value, "Food")
                self.assertEqual(ws.cell(2, headers[VAT_MXN_HEADER]).value, 78)
                self.assertEqual(ws.cell(2, headers["MXN Amount"]).value, 622.6)
                self.assertEqual(ws.cell(2, headers[TIPS_MXN_HEADER]).value, 56.6)
                self.assertIn("_crop_links", wb.sheetnames)
            finally:
                wb.close()

            reversed_result = reverse_review_sync_for_source(root, payment_source, {"163"})
            self.assertEqual(reversed_result.restored_trace_ids, ("037",))
            wb = load_workbook(workbook, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["MXN Amount"]).value, 566)
                self.assertEqual(ws.cell(2, headers[TIPS_MXN_HEADER]).value, 0)
                self.assertNotIn("_crop_links", wb.sheetnames)
            finally:
                wb.close()

    def test_source_sync_protected_pair_appends_and_warns(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / REIMBURSEMENT_WORKBOOK_NAME
            detail_source = root / "detail.jpg"
            payment_source = root / "payment.jpg"
            detail = InvoiceRecord(
                line_no=1, invoice_date="2026-07-01", seller="Cafe Uno", contents="food detail",
                expense_category="Food", currency="MXN", total_amount=100, vat_amount=16,
                crop_image=str(root / "crops" / "001_detail.jpg"), source_image=str(detail_source),
            )
            payment = InvoiceRecord(
                line_no=2, invoice_date="2026-07-01", seller="Cafe Uno", contents="tarjeta propina",
                expense_category="Food", currency="MXN", total_amount=115,
                crop_image=str(root / "crops" / "002_card.jpg"), source_image=str(payment_source),
            )
            ReimbursementWorkbook(workbook).write_records([detail])
            wb = load_workbook(workbook)
            try:
                wb[INVOICE_EXP_SHEET].cell(2, 2).value = "ok"
                wb.save(workbook)
            finally:
                wb.close()

            result = sync_source_to_review(root, payment_source, [payment], [detail, payment])

            self.assertEqual(len(result.appended_records), 1)
            self.assertTrue(any("protected 001" in warning for warning in result.warnings))
            wb = load_workbook(workbook, data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].max_row, 3)
            finally:
                wb.close()

    def test_source_sync_recovers_excel_saved_before_sync_state_without_duplicate(self):
        import invoice_system.reimbursement_excel as reimbursement_excel

        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "new.jpg"
            record = InvoiceRecord(
                line_no=1, invoice_date="2026-07-01", seller="Cafe", expense_category="Food",
                currency="MXN", total_amount=10, crop_image=str(root / "crops" / "001_cafe.jpg"),
                source_image=str(source),
            )
            initialize_review_architecture(root, [])
            original_save = reimbursement_excel._save_review_sync_state
            calls = 0

            def fail_after_excel(path, data):
                nonlocal calls
                calls += 1
                if calls == 2:
                    raise RuntimeError("simulated crash after Review save")
                original_save(path, data)

            with patch.object(reimbursement_excel, "_save_review_sync_state", side_effect=fail_after_excel):
                with self.assertRaisesRegex(RuntimeError, "simulated crash"):
                    sync_source_to_review(root, source, [record], [record])

            recovered = sync_source_to_review(root, source, [record], [record])
            self.assertTrue(recovered.already_committed)
            wb = load_workbook(root / REIMBURSEMENT_WORKBOOK_NAME, data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].max_row, 2)
            finally:
                wb.close()

    def test_repair_review_requires_unchanged_explicit_preview(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / REIMBURSEMENT_WORKBOOK_NAME
            source = root / "source.jpg"
            crop = root / "crops" / "001_cafe.jpg"
            machine = InvoiceRecord(
                line_no=1, invoice_date="2026-07-01", seller="Machine Cafe", expense_category="Food",
                currency="MXN", total_amount=10, crop_image=str(crop), source_image=str(source),
            )
            ReimbursementWorkbook(workbook).write_records([machine])
            wb = load_workbook(workbook)
            try:
                wb[INVOICE_EXP_SHEET].cell(2, 9).value = "Manual Cafe"
                wb.save(workbook)
            finally:
                wb.close()
            before = workbook.read_bytes()

            preview = preview_review_repair(root, [machine])

            self.assertEqual(workbook.read_bytes(), before)
            self.assertEqual(preview.changed, 1)
            wb = load_workbook(workbook)
            try:
                wb[INVOICE_EXP_SHEET].cell(2, 10).value = "edited after preview"
                wb.save(workbook)
            finally:
                wb.close()
            with self.assertRaisesRegex(RuntimeError, "changed after preview"):
                confirm_review_repair(root, preview.preview_id, [machine])

            fresh = preview_review_repair(root, [machine])
            confirm_review_repair(root, fresh.preview_id, [machine])
            wb = load_workbook(workbook, data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].cell(2, 9).value, "Machine Cafe")
            finally:
                wb.close()

    def test_reimbursement_workbook_name_is_real_chinese(self):
        self.assertEqual(REIMBURSEMENT_WORKBOOK_NAME, "报销明细_2026_xlsx.xlsx")

    def test_usd_receipt_writes_chinese_type_english_accounting_and_original_currency_fields(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(
                path,
                fetch_rates=lambda: [ExchangeRate(date(2026, 6, 12), usd_cny_per_100=700, mxn_per_100_cny=250)],
            )
            record = InvoiceRecord(
                line_no=1,
                invoice_date="2026-06-12",
                expense_category="Food",
                currency="USD",
                total_amount=10,
                vat_amount=1.6,
                tips=2,
                seller="Cafe",
                report_components=True,
            )

            store.write_records([record])

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(1, 2).value, "Manual status")
                self.assertEqual(ws.cell(1, 12).value, "IVA/VAT MXN")
                self.assertEqual(ws.cell(1, 13).value, "Tips MXN")
                self.assertEqual(ws.cell(1, 14).value, "Trace ID")
                self.assertEqual(ws.cell(1, 15).value, "System note")
                self.assertEqual(ws.cell(1, 16).value, "Invoice link")
                self.assertEqual(ws.cell(2, 4).value, 175)
                self.assertEqual(ws.cell(2, 5).value, "餐饮")
                self.assertEqual(ws.cell(2, 6).value, "USD")
                self.assertEqual(ws.cell(2, 7).value, 10)
                self.assertEqual(ws.cell(2, 8).value, 17.5)
                self.assertEqual(ws.cell(2, 11).value, "Food")
                self.assertEqual(ws.cell(2, 12).value, 28)
                self.assertEqual(ws.cell(2, 13).value, 35)
                rates = wb["exchange rate"]
                self.assertEqual(rates.cell(1, 1).value, "日期")
                self.assertEqual(rates.cell(1, 2).value, "USD")
                self.assertEqual(rates.cell(1, 25).value, "MXN")
                self.assertEqual(rates.cell(2, 2).value, 700)
                self.assertEqual(rates.cell(2, 25).value, 250)
            finally:
                wb.close()

    def test_reimbursement_workbook_keeps_manual_rows_in_one_sheet(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(path)

            store.write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe"),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Gas", total_amount=200, seller="Pemex"),
                ]
            )

            wb = load_workbook(path, data_only=True)
            try:
                self.assertEqual(wb[INVOICE_EXP_SHEET].cell(2, 9).value, "Cafe")
                self.assertEqual(wb[INVOICE_EXP_SHEET].cell(3, 9).value, "Pemex")
                self.assertNotIn(FOOD_EXP_SHEET, wb.sheetnames)
                self.assertNotIn(OTHER_EXP_SHEET, wb.sheetnames)
            finally:
                wb.close()

    def test_legacy_record_does_not_backfill_vat_or_tips_columns(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            record = InvoiceRecord(
                line_no=99,
                invoice_date="2026-06-05",
                expense_category="Food",
                total_amount=331.10,
                vat_amount=16,
                tips=30.10,
                seller="Legacy Cafe",
                report_components=False,
            )

            ReimbursementWorkbook(path).write_records([record])

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {cell.value: cell.column for cell in ws[1]}
                self.assertIsNone(ws.cell(2, headers["IVA/VAT MXN"]).value)
                self.assertIsNone(ws.cell(2, headers["Tips MXN"]).value)
            finally:
                wb.close()

    def test_usd_receipt_before_rate_table_uses_nearest_available_rate(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(
                path,
                fetch_rates=lambda: [ExchangeRate(date(2026, 1, 2), usd_cny_per_100=700, mxn_per_100_cny=250)],
            )

            store.write_records(
                [
                    InvoiceRecord(
                        line_no=1,
                        invoice_date="2025-12-11",
                        expense_category="Other",
                        currency="USD",
                        total_amount=40,
                        seller="Target",
                    )
                ]
            )

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(2, 4).value, 700)
                self.assertEqual(ws.cell(2, 6).value, "USD")
                self.assertEqual(ws.cell(2, 7).value, 40)
                self.assertEqual(ws.cell(2, 8).value, 17.5)
            finally:
                wb.close()

    def test_exchange_rate_window_uses_all_invoice_dates_and_stops_at_last_invoice(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            calls = []

            def fetch_rates(start, end):
                calls.append((start, end))
                return [
                    ExchangeRate(day, usd_cny_per_100=700, mxn_per_100_cny=250)
                    for day in {start, end}
                ]

            store = ReimbursementWorkbook(path, fetch_rates=fetch_rates)
            store.write_records(
                [
                    InvoiceRecord(invoice_date="2026-05-01", currency="MXN", total_amount=100, seller="Cafe"),
                    InvoiceRecord(invoice_date="2026-05-05", currency="USD", total_amount=10, seller="Store"),
                ]
            )

            self.assertEqual(calls, [(date(2026, 4, 21), date(2026, 5, 5))])
            wb = load_workbook(path, data_only=True)
            try:
                rate_dates = [
                    _date_text(row[0])
                    for row in wb["exchange rate"].iter_rows(min_row=2, values_only=True)
                    if row[0]
                ]
                self.assertEqual(
                    rate_dates,
                    ["2026-05-01", "2026-05-02", "2026-05-03", "2026-05-04", "2026-05-05"],
                )
            finally:
                wb.close()

    def test_new_foreign_currency_refreshes_full_review_range_and_checked_copies_it(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            calls = []

            def fetch_rates(start, end):
                calls.append((start, end))
                days = (end - start).days + 1
                return [
                    ExchangeRate(start + timedelta(days=offset), usd_cny_per_100=700, mxn_per_100_cny=250)
                    for offset in range(days)
                ]

            old_source = root / "old.jpg"
            new_source = root / "new.jpg"
            old = InvoiceRecord(
                line_no=1,
                invoice_date="2026-05-01",
                currency="MXN",
                total_amount=20,
                seller="Cafe",
                source_image=str(old_source),
                crop_image=str(root / "crops" / "001_old.jpg"),
            )
            new = InvoiceRecord(
                line_no=2,
                invoice_date="2026-07-13",
                currency="USD",
                total_amount=10,
                seller="Store",
                source_image=str(new_source),
                crop_image=str(root / "crops" / "002_new.jpg"),
            )
            ReimbursementWorkbook(path).write_records([old])
            wb = load_workbook(path)
            try:
                old_row = tuple(wb[INVOICE_EXP_SHEET].iter_rows(min_row=2, max_row=2, values_only=True))[0]
            finally:
                wb.close()

            calls.clear()
            sync_source_to_review(root, new_source, [new], [old, new], fetch_rates=fetch_rates)
            self.assertEqual(calls, [(date(2026, 4, 21), date(2026, 7, 13))])
            calls.clear()
            result = build_checked_outputs(root, fetch_rates=fetch_rates)

            self.assertEqual(calls, [(date(2026, 4, 21), date(2026, 7, 13))])
            for workbook_path in (path, result.workbook_path):
                wb = load_workbook(workbook_path, data_only=True)
                try:
                    rate_dates = [
                        row[0].date() if hasattr(row[0], "date") else row[0]
                        for row in wb["exchange rate"].iter_rows(min_row=2, values_only=True)
                        if row[0]
                    ]
                    self.assertEqual(min(rate_dates), date(2026, 5, 1))
                    self.assertEqual(max(rate_dates), date(2026, 7, 13))
                finally:
                    wb.close()
            wb = load_workbook(path, data_only=False)
            try:
                self.assertEqual(
                    tuple(wb[INVOICE_EXP_SHEET].iter_rows(min_row=2, max_row=2, values_only=True))[0],
                    old_row,
                )
            finally:
                wb.close()

    def test_eur_receipt_uses_safe_currency_table_when_available(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(
                path,
                fetch_rates=lambda: [
                    ExchangeRate(
                        date(2026, 6, 12),
                        usd_cny_per_100=700,
                        mxn_per_100_cny=250,
                        rates={"USD": 700, "EUR": 800, "MXN": 250},
                    )
                ],
            )

            store.write_records(
                [
                    InvoiceRecord(
                        line_no=1,
                        invoice_date="2026-06-12",
                        expense_category="Other",
                        currency="EUR",
                        total_amount=10,
                        seller="Hotel",
                    )
                ]
            )

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(2, 4).value, 200)
                self.assertEqual(ws.cell(2, 6).value, "EUR")
                self.assertEqual(ws.cell(2, 8).value, 20)
            finally:
                wb.close()

    def test_corrected_duplicate_record_is_not_written_again(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(path)
            record = InvoiceRecord(
                line_no=1,
                invoice_date="2026-06-12",
                expense_category="Food",
                currency="USD",
                total_amount=10,
                seller="Cafe",
            )
            store.write_records([record])
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "corrected"
                wb.save(path)
            finally:
                wb.close()

            store.write_records([record])

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                rows = [
                    [ws.cell(row, col).value for col in range(1, 12)]
                    for row in range(2, ws.max_row + 1)
                    if any(ws.cell(row, col).value not in (None, "") for col in range(1, 12))
                ]
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0][0], 1)
            finally:
                wb.close()

    def test_protected_crop_id_is_not_written_again_when_values_change(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            crop = root / "review_crops" / "004_2023-10-27_MXN_215.00_Restaurante.jpg"
            crop.parent.mkdir()
            crop.write_bytes(b"jpg")
            store = ReimbursementWorkbook(path)
            store.write_records(
                [
                    InvoiceRecord(
                        line_no=4,
                        invoice_date="2026-04-25",
                        expense_category="Food",
                        total_amount=215,
                        seller="Restaurante",
                        crop_image=str(crop),
                    )
                ]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "ok"
                wb.save(path)
            finally:
                wb.close()

            store.write_records(
                [
                    InvoiceRecord(
                        line_no=5,
                        invoice_date="2023-10-27",
                        expense_category="Food",
                        total_amount=215,
                        seller="Restaurante",
                        crop_image=str(crop),
                    )
                ]
            )

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                rows = [
                    [ws.cell(row, col).value for col in range(1, 13)]
                    for row in range(2, ws.max_row + 1)
                    if any(ws.cell(row, col).value not in (None, "") for col in range(1, 13))
                ]
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0][0], 4)
                self.assertEqual(rows[0][1], "ok")
            finally:
                wb.close()

    def test_marker_outside_manual_status_does_not_lock_row(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(path)
            store.write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe")]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 10).value = "ok"
                wb.save(path)
            finally:
                wb.close()

            replacement = InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Gas", total_amount=50, seller="Pemex")
            store.write_records([replacement])

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(2, 1).value, 2)
                self.assertEqual(ws.cell(2, 9).value, "Pemex")
                self.assertIsNone(ws.cell(2, 2).value)
            finally:
                wb.close()

    def test_deleted_row_is_not_loaded_but_does_not_block_reupload(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(path)
            store.write_records([InvoiceRecord(line_no=2, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe")])
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "deleted"
                wb.save(path)
            finally:
                wb.close()

            self.assertEqual([], load_reimbursement_records(path))

            store.write_records([InvoiceRecord(line_no=3, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe")])

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                rows = [
                    [ws.cell(row, col).value for col in range(1, 13)]
                    for row in range(2, ws.max_row + 1)
                    if any(ws.cell(row, col).value not in (None, "") for col in range(1, 13))
                ]
                self.assertEqual(len(rows), 2)
                self.assertEqual(rows[0][1], "deleted")
                self.assertEqual(rows[1][8], "Cafe")
                self.assertIsNone(rows[1][1])
            finally:
                wb.close()

    def test_ok_and_correct_markers_lock_rows_and_delete_takes_precedence(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(path)
            records = [
                InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe"),
                InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Gas", total_amount=200, seller="Pemex"),
            ]
            store.write_records(records)
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "ok"
                ws.cell(3, 2).value = "ok deleted"
                wb.save(path)
            finally:
                wb.close()

            self.assertEqual(store.locked_numbers(), {1, 2})
            loaded = load_reimbursement_records(path)
            self.assertEqual(len(loaded), 1)
            self.assertEqual(loaded[0].seller, "Cafe")

    def test_build_checked_outputs_filters_deleted_and_reindexes_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            crop1 = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop2 = review / "002_2026-06-13_MXN_200.00_Pemex.jpg"
            crop3 = review / "003_2026-06-14_MXN_50.00_Store.jpg"
            for crop in (crop1, crop2, crop3):
                crop.write_bytes(b"jpg")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=1,
                        invoice_date="2026-06-12",
                        expense_category="Food",
                        total_amount=100,
                        vat_amount=16,
                        tips=10,
                        seller="Cafe",
                        crop_image=str(crop1),
                        report_components=True,
                    ),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Gas", total_amount=200, seller="Pemex", crop_image=str(crop2)),
                    InvoiceRecord(line_no=3, invoice_date="2026-06-14", expense_category="Other", total_amount=50, seller="Store", crop_image=str(crop3)),
                ]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(3, 2).value = "deleted"
                wb.save(path)
            finally:
                wb.close()

            result = build_checked_outputs(root, refresh_exchange_rates=False)

            self.assertEqual(result.records_written, 2)
            self.assertEqual(result.crops_written, 2)
            self.assertEqual(result.workbook_path.name, CHECKED_WORKBOOK_NAME)
            final_crops = sorted((root / "final_crops").rglob("*.jpg"))
            self.assertEqual(
                [path.relative_to(root / "final_crops").as_posix() for path in final_crops],
                ["food/001_trace001_2026-06-12_MXN_100.00_Cafe.jpg", "other/002_trace003_2026-06-14_MXN_50.00_Store.jpg"],
            )
            checked = load_workbook(root / CHECKED_WORKBOOK_NAME, data_only=True)
            try:
                self.assertIn(FOOD_EXP_SHEET, checked.sheetnames)
                self.assertIn(OTHER_EXP_SHEET, checked.sheetnames)
                food_headers = {checked[FOOD_EXP_SHEET].cell(1, col).value: col for col in range(1, checked[FOOD_EXP_SHEET].max_column + 1)}
                other_headers = {checked[OTHER_EXP_SHEET].cell(1, col).value: col for col in range(1, checked[OTHER_EXP_SHEET].max_column + 1)}
                self.assertEqual(checked[FOOD_EXP_SHEET].cell(2, food_headers["Merchant"]).value, "Cafe")
                self.assertEqual(checked[FOOD_EXP_SHEET].cell(2, food_headers["IVA/VAT MXN"]).value, 16)
                self.assertEqual(checked[FOOD_EXP_SHEET].cell(2, food_headers["Tips MXN"]).value, 10)
                self.assertEqual(checked[FOOD_EXP_SHEET].cell(2, food_headers["Trace ID"]).value, "001")
                self.assertEqual(checked[OTHER_EXP_SHEET].cell(2, other_headers["Merchant"]).value, "Store")
                self.assertEqual(checked[OTHER_EXP_SHEET].cell(2, other_headers["Trace ID"]).value, "003")
                self.assertEqual(checked[FOOD_EXP_SHEET].cell(2, food_headers["Invoice link"]).value, "final_crops/food/001_trace001_2026-06-12_MXN_100.00_Cafe.jpg")
                self.assertEqual(checked[OTHER_EXP_SHEET].cell(2, other_headers["Invoice link"]).value, "final_crops/other/002_trace003_2026-06-14_MXN_50.00_Store.jpg")
                ws = checked[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.max_row, 3)
                self.assertEqual(ws.cell(1, 2).value, "Date")
                self.assertEqual(ws.cell(1, headers["Invoice link"]).value, "Invoice link")
                self.assertEqual(ws.cell(2, 1).value, 1)
                self.assertEqual(ws.cell(3, 1).value, 2)
                self.assertEqual(ws.cell(3, headers["Merchant"]).value, "Store")
                self.assertEqual(ws.cell(3, headers["Trace ID"]).value, "003")
                self.assertEqual(ws.cell(3, headers["Invoice link"]).value, "final_crops/other/002_trace003_2026-06-14_MXN_50.00_Store.jpg")
                self.assertEqual(ws.cell(3, headers["Invoice link"]).hyperlink.target, "final_crops/other/002_trace003_2026-06-14_MXN_50.00_Store.jpg")
            finally:
                checked.close()

    def test_checked_summary_lists_each_day_and_only_counts_food(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            crops = root / "crops"
            crops.mkdir()
            crop1 = crops / "001_2026-07-01_MXN_100.00_Cafe.jpg"
            crop2 = crops / "002_2026-07-02_MXN_999.00_Deleted_Food.jpg"
            crop3 = crops / "003_2026-07-03_MXN_200.00_Restaurant.jpg"
            crop4 = crops / "004_2026-07-04_MXN_88.00_Store.jpg"
            for crop in (crop1, crop2, crop3, crop4):
                crop.write_bytes(b"jpg")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-07-01", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(crop1)),
                    InvoiceRecord(line_no=2, invoice_date="2026-07-02", expense_category="Food", total_amount=999, seller="Deleted Food", crop_image=str(crop2)),
                    InvoiceRecord(line_no=3, invoice_date="2026-07-03", expense_category="Food", total_amount=200, seller="Restaurant", crop_image=str(crop3)),
                    InvoiceRecord(line_no=4, invoice_date="2026-07-04", expense_category="Other", total_amount=88, seller="Store", crop_image=str(crop4)),
                ]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(3, 2).value = "deleted"
                wb.save(path)
            finally:
                wb.close()

            build_checked_outputs(root, refresh_exchange_rates=False)

            checked = load_workbook(root / CHECKED_WORKBOOK_NAME, data_only=True)
            try:
                self.assertEqual(checked.sheetnames[0], SUMMARY_SHEET)
                ws = checked[SUMMARY_SHEET]
                self.assertEqual(ws["B3"].value, "2026-07-01 \u81f3 2026-07-04")
                self.assertEqual(ws["B4"].value, "4 \u5929")
                self.assertEqual(ws["B5"].value, "300.00 MXN")
                self.assertEqual(ws["B6"].value, "75.00 MXN")
                self.assertEqual(ws["B7"].value, "2 \u5929")
                self.assertEqual(ws["B8"].value, "2026-07-02, 2026-07-04")
                daily = [
                    (_date_text(ws.cell(row, 1).value), ws.cell(row, 2).value, ws.cell(row, 3).value)
                    for row in range(12, 16)
                ]
                self.assertEqual(
                    daily,
                    [
                        ("2026-07-01", 100, "Yes"),
                        ("2026-07-02", 0, "No"),
                        ("2026-07-03", 200, "Yes"),
                        ("2026-07-04", 0, "No"),
                    ],
                )
            finally:
                checked.close()

    def test_build_checked_outputs_reuses_unchanged_final_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            crop1 = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop2 = review / "002_2026-06-13_MXN_50.00_Store.jpg"
            crop3 = review / "003_2026-06-14_MXN_25.00_Kiosk.jpg"
            crop1.write_bytes(b"one")
            crop2.write_bytes(b"two")
            crop3.write_bytes(b"three")
            records = [
                InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(crop1)),
                InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Other", total_amount=50, seller="Store", crop_image=str(crop2)),
            ]
            store = ReimbursementWorkbook(path)
            store.write_records(records)
            build_checked_outputs(root, refresh_exchange_rates=False)
            first_final = root / "final_crops" / "food" / "001_trace001_2026-06-12_MXN_100.00_Cafe.jpg"
            first_final.write_bytes(b"sentinel")

            store.write_records(
                [
                    *records,
                    InvoiceRecord(line_no=3, invoice_date="2026-06-14", expense_category="Other", total_amount=25, seller="Kiosk", crop_image=str(crop3)),
                ]
            )
            result = build_checked_outputs(root, refresh_exchange_rates=False)

            self.assertEqual(result.crops_written, 3)
            self.assertEqual(first_final.read_bytes(), b"sentinel")
            self.assertTrue((root / "final_crops" / "other" / "003_trace003_2026-06-14_MXN_25.00_Kiosk.jpg").exists())
            self.assertTrue((root / FINAL_CROPS_MANIFEST).exists())

    def test_build_checked_outputs_prunes_deleted_final_crop_from_manifest(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            crop1 = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop2 = review / "002_2026-06-13_MXN_50.00_Store.jpg"
            crop1.write_bytes(b"one")
            crop2.write_bytes(b"two")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(crop1)),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Other", total_amount=50, seller="Store", crop_image=str(crop2)),
                ]
            )
            build_checked_outputs(root, refresh_exchange_rates=False)
            deleted_final = root / "final_crops" / "food" / "001_trace001_2026-06-12_MXN_100.00_Cafe.jpg"
            self.assertTrue(deleted_final.exists())
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "delete"
                wb.save(path)
            finally:
                wb.close()

            build_checked_outputs(root, refresh_exchange_rates=False)

            self.assertFalse(deleted_final.exists())
            manifest = json.loads((root / FINAL_CROPS_MANIFEST).read_text(encoding="utf-8"))
            self.assertNotIn("001", manifest["records"])
            self.assertIn("002", manifest["records"])

    def test_build_checked_outputs_force_rebuild_removes_stale_final_crop(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            crop = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop.write_bytes(b"one")
            stale = root / "final_crops" / "other" / "999_old.jpg"
            stale.parent.mkdir(parents=True)
            stale.write_bytes(b"stale")
            ReimbursementWorkbook(path).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(crop))]
            )

            result = build_checked_outputs(root, force=True, refresh_exchange_rates=False)

            self.assertEqual(result.crops_written, 1)
            self.assertFalse(stale.exists())
            self.assertTrue((root / "final_crops" / "food" / "001_trace001_2026-06-12_MXN_100.00_Cafe.jpg").exists())

    def test_build_checked_outputs_combines_supporting_crops_into_one_final_image(self):
        with tempfile.TemporaryDirectory() as temp:
            from PIL import Image

            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            primary = review / "040_2026-06-12_MXN_126.00_Cafe.jpg"
            supporting = review / "041_2026-06-12_MXN_126.00_Cafe_card.jpg"
            Image.new("RGB", (120, 240), "white").save(primary)
            Image.new("RGB", (100, 200), "white").save(supporting)
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=1,
                        invoice_date="2026-06-12",
                        expense_category="Food",
                        total_amount=126,
                        seller="Cafe",
                        crop_image=str(primary),
                        supporting_crop_images=[str(supporting)],
                    )
                ]
            )

            result = build_checked_outputs(root, refresh_exchange_rates=False)

            self.assertEqual(result.records_written, 1)
            self.assertEqual(result.crops_written, 1)
            final_crops = sorted((root / "final_crops").rglob("*.jpg"))
            self.assertEqual(
                [path.relative_to(root / "final_crops").as_posix() for path in final_crops],
                ["food/001_trace040_2026-06-12_MXN_126.00_Cafe.jpg"],
            )
            with Image.open(final_crops[0]) as combined:
                self.assertGreater(combined.width, 120)
                self.assertGreaterEqual(combined.height, 240)
            checked = load_workbook(root / CHECKED_WORKBOOK_NAME, data_only=True)
            try:
                ws = checked[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.max_row, 2)
                self.assertEqual(ws.cell(2, headers["Trace ID"]).value, "040")
                self.assertEqual(ws.cell(2, headers["Invoice link"]).value, "final_crops/food/001_trace040_2026-06-12_MXN_126.00_Cafe.jpg")
            finally:
                checked.close()

    def test_preview_and_apply_reimbursement_group_merges_rows_and_archives_original_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            from PIL import Image

            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            crops = root / "crops"
            crops.mkdir()
            detail_crop = crops / "044_2026-05-09_MXN_566.00_SUSHI_ROLL.jpg"
            card_crop = crops / "045_2026-05-09_MXN_622.60_SUSHI_ROLL_CARD.jpg"
            Image.new("RGB", (120, 240), "white").save(detail_crop)
            Image.new("RGB", (100, 220), "white").save(card_crop)
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=44,
                        invoice_date="2026-05-09",
                        expense_category="Food",
                        currency="MXN",
                        total_amount=566,
                        seller="SUSHI ROLL",
                        crop_image=str(detail_crop),
                    ),
                    InvoiceRecord(
                        line_no=45,
                        invoice_date="2026-05-09",
                        expense_category="Food",
                        currency="MXN",
                        total_amount=622.60,
                        tips=56.60,
                        seller="SUSHI ROLL MIFEL",
                        contents="venta con propina",
                        crop_image=str(card_crop),
                    ),
                ]
            )

            preview = preview_reimbursement_group(root, ["044", "045"])

            self.assertEqual(preview.primary_id, "044")
            self.assertEqual(preview.total_amount, 622.60)
            self.assertEqual(preview.tips, 56.60)
            self.assertTrue(detail_crop.exists())
            self.assertTrue(card_crop.exists())

            result = apply_reimbursement_group(root, ["044", "045"])

            self.assertEqual(result.primary_id, "044")
            self.assertEqual(result.deleted_ids, ("045",))
            self.assertTrue(result.crop_path.exists())
            self.assertTrue((root / "group_archive" / "044_045" / detail_crop.name).exists())
            self.assertTrue((root / "group_archive" / "044_045" / card_crop.name).exists())
            self.assertFalse(detail_crop.exists())
            self.assertFalse(card_crop.exists())
            records = load_reimbursement_records(path)
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].line_no, 44)
            self.assertEqual(records[0].total_amount, 622.60)
            self.assertEqual(result.tips, 56.60)
            self.assertEqual(Path(records[0].crop_image).name, result.crop_path.name)

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "correct")
                self.assertEqual(ws.cell(3, headers["Manual status"]).value, "delete")
                self.assertEqual(ws.cell(3, headers["System note"]).value, "Grouped into 044")
            finally:
                wb.close()

    def test_change_reimbursement_record_updates_category_and_locks_by_crop_id(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            crop = review / "021_2026-06-12_MXN_100.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(path).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(crop))]
            )

            result = change_reimbursement_record(root, "021", invoice_date="2026-07-01", category="Other")

            self.assertEqual(result.crop_id, "021")
            self.assertEqual(result.status, "ok")
            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, columns["Date"]).value.date(), date(2026, 7, 1))
                self.assertEqual(ws.cell(2, columns["Accounting Category"]).value, "Other")
                self.assertEqual(ws.cell(2, columns["Manual status"]).value, "ok")
            finally:
                wb.close()

    def test_change_reimbursement_record_can_clear_date_and_correct_tax_tip(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            crop = root / "crops" / "021_2026-06-12_MXN_170.00_Nota.jpg"
            crop.parent.mkdir()
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=21,
                        invoice_date="2026-06-12",
                        expense_category="Food",
                        total_amount=170,
                        vat_amount=10,
                        tips=30,
                        seller="Nota De Cuenta",
                        crop_image=str(crop),
                        report_components=True,
                    )
                ]
            )

            change_reimbursement_record(
                root,
                "021",
                invoice_date="",
                amount=140,
                vat_amount=0,
                tips=0,
                status="correct",
            )

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertIsNone(ws.cell(2, columns["Date"]).value)
                self.assertEqual(ws.cell(2, columns["MXN Amount"]).value, 140)
                self.assertEqual(ws.cell(2, columns[VAT_MXN_HEADER]).value, 0)
                self.assertEqual(ws.cell(2, columns[TIPS_MXN_HEADER]).value, 0)
                self.assertEqual(ws.cell(2, columns["Manual status"]).value, "correct")
            finally:
                wb.close()

    def test_manual_no_syncs_to_crop_id_and_change_uses_crop_id(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            old_crop = review / "001_2026-06-10_MXN_10.00_Old.jpg"
            crop_001 = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop_002 = review / "002_2026-06-13_MXN_50.00_Gas.jpg"
            old_crop.write_bytes(b"old")
            crop_001.write_bytes(b"one")
            crop_002.write_bytes(b"two")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-10", expense_category="Food", total_amount=10, seller="Old", crop_image=str(old_crop)),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(crop_001)),
                    InvoiceRecord(line_no=3, invoice_date="2026-06-13", expense_category="Gas", total_amount=50, seller="Gas", crop_image=str(crop_002)),
                ]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                ws.cell(2, columns["Manual status"]).value = "delete"
                ws.cell(3, columns["No."]).value = 2
                ws.cell(4, columns["No."]).value = 3
                wb.save(path)
            finally:
                wb.close()

            change_reimbursement_record(root, "002", category="Other")

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(3, columns["No."]).value, 1)
                self.assertEqual(ws.cell(4, columns["No."]).value, 2)
                self.assertEqual(ws.cell(3, columns["Accounting Category"]).value, "Food")
                self.assertEqual(ws.cell(4, columns["Accounting Category"]).value, "Other")
                self.assertEqual(ws.cell(4, columns["Manual status"]).value, "ok")
            finally:
                wb.close()

    def test_change_reimbursement_record_updates_usd_amount_and_delete_excludes_checked(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            crop = review / "021_2026-06-12_USD_10.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            store = ReimbursementWorkbook(
                path,
                fetch_rates=lambda: [ExchangeRate(date(2026, 6, 12), usd_cny_per_100=700, mxn_per_100_cny=250)],
            )
            store.write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", currency="USD", total_amount=10, seller="Cafe", crop_image=str(crop))]
            )

            change_reimbursement_record(root, "021", amount=20, currency="USD", status="correct")
            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, columns["原金额"]).value, 20)
                self.assertEqual(ws.cell(2, columns["MXN Amount"]).value, 350)
                self.assertEqual(ws.cell(2, columns["Manual status"]).value, "correct")
            finally:
                wb.close()

            change_reimbursement_record(root, "021", status="delete")
            checked = build_checked_outputs(root, refresh_exchange_rates=False)

            self.assertEqual(checked.records_written, 0)
            self.assertEqual(checked.crops_written, 0)

    def test_deleted_row_does_not_block_same_invoice_reupload(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            old_crop = review / "038_2026-06-12_USD_11.00_Store.jpg"
            new_crop = review / "041_2026-06-12_USD_11.00_Store.jpg"
            old_crop.write_bytes(b"old")
            new_crop.write_bytes(b"new")
            store = ReimbursementWorkbook(
                path,
                fetch_rates=lambda: [ExchangeRate(date(2026, 6, 12), usd_cny_per_100=700, mxn_per_100_cny=250)],
            )
            store.write_records(
                [InvoiceRecord(line_no=38, invoice_date="2026-06-12", expense_category="Other", currency="USD", total_amount=11, seller="Store", crop_image=str(old_crop))]
            )
            change_reimbursement_record(root, "038", status="delete")

            store.write_records(
                [
                    InvoiceRecord(line_no=38, invoice_date="2026-06-12", expense_category="Other", currency="USD", total_amount=11, seller="Store", crop_image=str(old_crop)),
                    InvoiceRecord(line_no=41, invoice_date="2026-06-12", expense_category="Other", currency="USD", total_amount=11, seller="Store", crop_image=str(new_crop)),
                ]
            )

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                rows = [
                    (ws.cell(row, columns["Invoice link"]).value, ws.cell(row, columns["Manual status"]).value)
                    for row in range(2, ws.max_row + 1)
                    if ws.cell(row, columns["Invoice link"]).value
                ]
                self.assertEqual(rows, [("review_crops/038_2026-06-12_USD_11.00_Store.jpg", "delete"), ("review_crops/041_2026-06-12_USD_11.00_Store.jpg", None)])
            finally:
                wb.close()

    def test_build_checked_outputs_reads_legacy_final_crops_without_review_copy(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            legacy_final = root / "final_crops" / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            legacy_nested = root / "final_crops" / "other" / "999_2026-06-12_MXN_25.00_Old.jpg"
            legacy_final.parent.mkdir()
            legacy_nested.parent.mkdir()
            legacy_final.write_bytes(b"jpg")
            legacy_nested.write_bytes(b"old")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=1,
                        invoice_date="2026-06-12",
                        expense_category="Food",
                        total_amount=100,
                        seller="Cafe",
                        crop_image=str(legacy_final),
                    )
                ]
            )

            result = build_checked_outputs(root, refresh_exchange_rates=False)

            self.assertEqual(result.records_written, 1)
            self.assertFalse((root / "review_crops" / legacy_final.name).exists())
            self.assertFalse((root / "review_crops" / legacy_nested.name).exists())
            self.assertTrue((root / "final_crops" / "food" / "001_trace001_2026-06-12_MXN_100.00_Cafe.jpg").exists())
            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, columns["Trace ID"]).value, "001")
                self.assertEqual(ws.cell(2, columns["Invoice link"]).value, f"final_crops/{legacy_final.name}")
            finally:
                wb.close()

    def test_build_checked_outputs_reconciles_manual_link_to_original_processing_crop_id(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            old_review = root / "review_crops" / "018_2025-12-03_USD_38.33_ALICIA.jpg"
            raw_crop = root / "crops" / "021_2025-12-03_USD_38.33_ALICIA.jpg"
            old_review.parent.mkdir()
            raw_crop.parent.mkdir()
            old_review.write_bytes(b"old")
            raw_crop.write_bytes(b"raw")
            store = ReimbursementWorkbook(
                path,
                fetch_rates=lambda: [ExchangeRate(date(2025, 12, 3), usd_cny_per_100=700, mxn_per_100_cny=250)],
            )
            store.write_records(
                [
                    InvoiceRecord(
                        line_no=18,
                        invoice_date="2025-12-03",
                        expense_category="Food",
                        currency="USD",
                        total_amount=38.33,
                        seller="ALICIA",
                        crop_image=str(old_review),
                    )
                ]
            )
            (root / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "invoice_date": "2025-12-03",
                                "seller": "ALICIA",
                                "currency": "USD",
                                "total_amount": 38.33,
                                "crop_image": str(raw_crop),
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )

            build_checked_outputs(root, refresh_exchange_rates=False)

            wb = load_workbook(path, data_only=False)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                link = ws.cell(2, columns["Invoice link"]).hyperlink.target
                self.assertEqual(link, f"review_crops/{old_review.name}")
                self.assertEqual(ws.cell(1, columns["Manual status"]).value, "Manual status")
                self.assertFalse((root / "review_crops" / raw_crop.name).exists())
            finally:
                wb.close()

    def test_change_preserves_distinct_trace_ids_for_same_merchant_and_amount(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            crop_111 = root / "crops" / "111_unknown-date_MXN_150.00_Nota_De_Cuenta.jpg"
            crop_207 = root / "crops" / "207_unknown-date_MXN_150.00_Nota_De_Cuenta.jpg"
            crop_111.parent.mkdir()
            crop_111.write_bytes(b"111")
            crop_207.write_bytes(b"207")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=111,
                        expense_category="Food",
                        total_amount=150,
                        seller="Nota De Cuenta",
                        crop_image=str(crop_111),
                    ),
                    InvoiceRecord(
                        line_no=207,
                        expense_category="Food",
                        total_amount=150,
                        seller="Nota De Cuenta",
                        crop_image=str(crop_207),
                    ),
                ]
            )
            (root / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {
                                "invoice_date": "",
                                "seller": "Nota De Cuenta",
                                "currency": "MXN",
                                "total_amount": 150,
                                "crop_image": str(crop_111),
                            },
                            {
                                "invoice_date": "",
                                "seller": "Nota De Cuenta",
                                "currency": "MXN",
                                "total_amount": 150,
                                "crop_image": str(crop_207),
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )

            change_reimbursement_record(root, "207", comment="verified")

            wb = load_workbook(path, data_only=False)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                identities = [
                    (
                        str(ws.cell(row, columns["Trace ID"]).value).zfill(3),
                        Path(str(ws.cell(row, columns["Invoice link"]).value)).name,
                    )
                    for row in range(2, ws.max_row + 1)
                ]
                self.assertEqual(
                    identities,
                    [("111", crop_111.name), ("207", crop_207.name)],
                )
            finally:
                wb.close()

    def test_source_sync_appends_undated_possible_duplicate_and_names_old_trace(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = root / REIMBURSEMENT_WORKBOOK_NAME
            old_source = root / "old.jpg"
            new_source = root / "new.jpg"
            old = InvoiceRecord(
                line_no=111,
                invoice_date="2026-06-06",
                seller="Nota De Cuenta",
                expense_category="Food",
                currency="MXN",
                total_amount=150,
                crop_image=str(root / "crops" / "111_old.jpg"),
                source_image=str(old_source),
            )
            new = InvoiceRecord(
                line_no=207,
                invoice_date="",
                seller="Nota De Cuenta",
                expense_category="Food",
                currency="MXN",
                total_amount=150,
                crop_image=str(root / "crops" / "207_new.jpg"),
                source_image=str(new_source),
            )
            ReimbursementWorkbook(workbook).write_records([old])
            wb = load_workbook(workbook)
            try:
                wb[INVOICE_EXP_SHEET].cell(2, 2).value = "correct"
                wb.save(workbook)
            finally:
                wb.close()

            result = sync_source_to_review(root, new_source, [new], [old, new])

            self.assertEqual(len(result.appended_records), 1)
            self.assertTrue(any("207 may duplicate 111" in warning for warning in result.warnings))
            wb = load_workbook(workbook, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                columns = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                traces = [str(ws.cell(row, columns["Trace ID"]).value).zfill(3) for row in range(2, ws.max_row + 1)]
                self.assertEqual(traces, ["111", "207"])
                self.assertEqual(ws.cell(2, columns["Manual status"]).value, "correct")
                self.assertIn("Possible duplicate with 111", ws.cell(3, columns["System note"]).value)
            finally:
                wb.close()

    def test_rerun_rebuilds_manual_workbook_from_edited_checked_finance_file(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            review = root / "review_crops"
            review.mkdir()
            food_crop = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            other_crop = review / "002_2026-06-13_MXN_200.00_Pemex.jpg"
            food_crop.write_bytes(b"food")
            other_crop.write_bytes(b"other")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe", crop_image=str(food_crop)),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Gas", total_amount=200, seller="Pemex", crop_image=str(other_crop)),
                ]
            )
            build_checked_outputs(root, refresh_exchange_rates=False)
            _move_checked_row(root / CHECKED_WORKBOOK_NAME, FOOD_EXP_SHEET, OTHER_EXP_SHEET, 2)

            result = rerun_checked_from_finance_edits(root)

            self.assertTrue(result.archive_dir.exists())
            self.assertEqual(result.moved, ("001 Food -> Other",))
            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(2, 11).value, "Other")
                self.assertEqual(ws.cell(2, 2).value, "correct")
            finally:
                wb.close()
            checked = load_workbook(root / CHECKED_WORKBOOK_NAME, data_only=True)
            try:
                other_headers = {checked[OTHER_EXP_SHEET].cell(1, col).value: col for col in range(1, checked[OTHER_EXP_SHEET].max_column + 1)}
                self.assertEqual(checked[FOOD_EXP_SHEET].max_row, 1)
                self.assertEqual(checked[OTHER_EXP_SHEET].cell(3, 1).value, 2)
                self.assertEqual(checked[OTHER_EXP_SHEET].cell(2, other_headers["Invoice link"]).value, "final_crops/other/001_trace001_2026-06-12_MXN_100.00_Cafe.jpg")
            finally:
                checked.close()
            self.assertTrue((root / "final_crops" / "other" / "001_trace001_2026-06-12_MXN_100.00_Cafe.jpg").exists())

            _move_checked_row(root / CHECKED_WORKBOOK_NAME, OTHER_EXP_SHEET, FOOD_EXP_SHEET, 2)
            second = rerun_checked_from_finance_edits(root)

            self.assertEqual(second.moved, ("001 Other -> Food",))
            checked = load_workbook(root / CHECKED_WORKBOOK_NAME, data_only=True)
            try:
                self.assertEqual(checked[FOOD_EXP_SHEET].cell(2, 1).value, 1)
                self.assertTrue((root / "final_crops" / "food" / "001_trace001_2026-06-12_MXN_100.00_Cafe.jpg").exists())
            finally:
                checked.close()

    def test_corrected_row_is_preserved_and_new_rows_skip_locked_number(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            store = ReimbursementWorkbook(path)
            store.write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", total_amount=100, seller="Cafe")]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "corrected"
                ws.cell(2, 9).value = "Manual Cafe"
                wb.save(path)
            finally:
                wb.close()

            records = [InvoiceRecord(invoice_date="2026-06-13", expense_category="Gas", total_amount=200, seller="Pemex")]
            assign_available_line_numbers(records, store.locked_numbers())
            store.write_records(records)

            wb = load_workbook(path, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(2, 1).value, 1)
                self.assertEqual(ws.cell(2, 9).value, "Manual Cafe")
                self.assertEqual(ws.cell(3, 1).value, 2)
                self.assertEqual(ws.cell(3, 5).value, "汽油")
                self.assertEqual(ws.cell(3, 9).value, "Pemex")
                self.assertEqual(ws.cell(3, 11).value, "Gas")
            finally:
                wb.close()

    def test_corrected_crop_names_are_preserved_when_clearing_generated_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            path = root / REIMBURSEMENT_WORKBOOK_NAME
            final_dir = root / "final_crops"
            final_dir.mkdir()
            keep = final_dir / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            stale = final_dir / "002_2026-06-12_MXN_50.00_Stale.jpg"
            keep.write_bytes(b"keep")
            stale.write_bytes(b"stale")
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(
                        line_no=1,
                        invoice_date="2026-06-12",
                        expense_category="Food",
                        total_amount=100,
                        seller="Cafe",
                        crop_image=str(keep),
                    )
                ]
            )
            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(2, 2).value = "corrected"
                wb.save(path)
            finally:
                wb.close()

            clear_generated_crops(final_dir, corrected_crop_names(path))

            self.assertTrue(keep.exists())
            self.assertFalse(stale.exists())

    def test_focus_workbook_selects_first_row_matching_target_day(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / REIMBURSEMENT_WORKBOOK_NAME
            ReimbursementWorkbook(path).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", total_amount=100, seller="Cafe"),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-21", expense_category="Gas", total_amount=200, seller="Pemex"),
                ]
            )

            row = focus_reimbursement_workbook(path, date(2026, 6, 21))

            wb = load_workbook(path)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertEqual(row, 3)
                self.assertEqual(wb.active.title, INVOICE_EXP_SHEET)
                self.assertEqual(ws.sheet_view.selection[0].activeCell, "A3")
            finally:
                wb.close()


def _move_checked_row(path: Path, source_sheet: str, target_sheet: str, source_row: int) -> None:
    wb = load_workbook(path)
    try:
        source = wb[source_sheet]
        target = wb[target_sheet]
        values = [source.cell(source_row, col).value for col in range(1, source.max_column + 1)]
        hyperlinks = [source.cell(source_row, col).hyperlink.target if source.cell(source_row, col).hyperlink else None for col in range(1, source.max_column + 1)]
        target.append(values)
        target_row = target.max_row
        for col, hyperlink in enumerate(hyperlinks, start=1):
            if hyperlink:
                target.cell(target_row, col).hyperlink = hyperlink
                target.cell(target_row, col).style = "Hyperlink"
        source.delete_rows(source_row, 1)
        wb.save(path)
    finally:
        wb.close()


if __name__ == "__main__":
    unittest.main()
