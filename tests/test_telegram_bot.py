import tempfile
import unittest
import json
import os
import subprocess
import sys
from unittest.mock import AsyncMock, patch
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from invoice_system.config import Settings
from invoice_system.models import InvoiceRecord, PipelineSummary
from invoice_system.grouping import mark_source_for_group_if_armed
from invoice_system.queue_worker import DONE, PENDING, QueueItem, load_queue_state, save_queue_state, telegram_user_queue_path, telegram_user_workbook, QueueState
from invoice_system.reimbursement_excel import INVOICE_EXP_SHEET, ReimbursementWorkbook
from invoice_system.telegram_bot import (
    append_process_status,
    change_message,
    delete_message,
    download_telegram_media,
    format_telegram_config,
    group_message,
    is_allowed_user,
    processing_failure_message,
    processing_success_message,
    queued_photo_message,
    recent_message,
    report_message,
    resolve_auto_process,
    review_crop_paths,
    rollback_message,
    saved_photo_message,
    scan_completion_message,
    set_user_language,
    submit_confirmation_message,
    submit_message,
    submit_pending_confirmation,
    telegram_batch_source,
    telegram_command_menu,
    telegram_config_ready,
    telegram_image_document_filename,
    telegram_help_message,
    telegram_photo_filename,
    telegram_polling_ready,
    telegram_start_message,
    user_language,
    whoami_message,
    is_supported_telegram_image_document,
    _acquire_telegram_instance_lock,
    _photo_quality_warning,
    _release_telegram_instance_lock,
)


class TelegramBotTests(unittest.TestCase):
    def test_empty_allowed_list_rejects_photos(self):
        self.assertFalse(is_allowed_user(1, set()))

    def test_allowed_user_filter(self):
        self.assertTrue(is_allowed_user(7, {7, 8}))
        self.assertFalse(is_allowed_user(9, {7, 8}))

    def test_resolve_auto_process_uses_env_setting_by_default(self):
        self.assertTrue(resolve_auto_process(Settings(telegram_auto_process=True)))
        self.assertFalse(resolve_auto_process(Settings(telegram_auto_process=False)))

    def test_resolve_auto_process_allows_cli_override(self):
        self.assertFalse(resolve_auto_process(Settings(telegram_auto_process=True), False))
        self.assertTrue(resolve_auto_process(Settings(telegram_auto_process=False), True))

    def test_telegram_config_ready_requires_token_and_allowed_ids(self):
        self.assertFalse(telegram_polling_ready(Settings()))
        self.assertTrue(telegram_polling_ready(Settings(telegram_bot_token="token")))
        self.assertFalse(telegram_config_ready(Settings(telegram_bot_token="token")))
        self.assertFalse(telegram_config_ready(Settings(telegram_allowed_user_ids=frozenset({123}))))
        self.assertTrue(telegram_config_ready(Settings(telegram_bot_token="token", telegram_allowed_user_ids=frozenset({123}))))

    def test_format_telegram_config_reports_missing_setup(self):
        text = format_telegram_config(Settings())

        self.assertIn("Bot token: missing", text)
        self.assertIn("Telegram package:", text)
        self.assertIn("Allowed user IDs: missing", text)
        self.assertIn("Polling startup: NOT READY", text)
        self.assertIn("Photo ingestion: NOT READY", text)
        self.assertIn("Status: NOT READY", text)

    def test_format_telegram_config_reports_setup_mode(self):
        text = format_telegram_config(Settings(telegram_bot_token="token"))

        self.assertIn("Bot token: configured", text)
        self.assertIn("Allowed user IDs: missing", text)
        self.assertIn("Polling startup: READY", text)
        self.assertIn("Photo ingestion: NOT READY", text)
        self.assertIn("Polling can start for /whoami", text)
        self.assertIn("Status: NOT READY", text)

    def test_format_telegram_config_reports_ready_setup(self):
        text = format_telegram_config(
            Settings(
                telegram_bot_token="token",
                telegram_allowed_user_ids=frozenset({123, 456}),
                telegram_auto_process=True,
                openai_api_key="key",
            )
        )

        self.assertIn("Bot token: configured", text)
        self.assertIn("Allowed user IDs: 2", text)
        self.assertIn("Auto process: enabled", text)
        self.assertIn("Company profile: default", text)
        self.assertIn("Qwen OCR: disabled", text)
        self.assertIn("OpenAI fallback: removed", text)
        self.assertIn("Polling startup: READY", text)
        self.assertIn("Photo ingestion: READY", text)
        self.assertIn("Status: READY", text)

    def test_format_telegram_config_reports_missing_opencv_splitter(self):
        settings = Settings(telegram_bot_token="token", telegram_allowed_user_ids=frozenset({123}))
        with patch("invoice_system.telegram_bot.missing_photo_processing_packages", return_value=("cv2",)):
            text = format_telegram_config(settings)

            self.assertFalse(telegram_config_ready(settings))
            self.assertIn("OpenCV splitter: MISSING", text)
            self.assertIn("Photo ingestion: NOT READY", text)
            self.assertIn("opencv-python", text)

    def test_format_telegram_config_warns_when_company_profile_is_missing(self):
        with tempfile.TemporaryDirectory() as temp:
            text = format_telegram_config(Settings(root=Path(temp), company_profile="missing-profile"))

        self.assertIn("Company profile: missing-profile (missing)", text)
        self.assertIn("Company profile warning:", text)

    def test_format_telegram_config_reports_enabled_qwen_scan(self):
        text = format_telegram_config(Settings(qwen_scan_enabled=True, qwen_api_key="key"))

        self.assertIn("Qwen OCR: enabled", text)

    def test_format_telegram_config_never_enables_openai_fallback(self):
        text = format_telegram_config(Settings(codex_scan_enabled=True, openai_api_key="key"))

        self.assertIn("OpenAI fallback: removed", text)
        self.assertNotIn("Codex Scan fallback: enabled", text)

    def test_telegram_start_message_reports_setup_mode_without_allowed_ids(self):
        self.assertIn("setup mode", telegram_start_message(Settings()))
        self.assertIn("photos are rejected", telegram_start_message(Settings()))
        self.assertIn("Invoice bot is ready", telegram_start_message(Settings(telegram_allowed_user_ids=frozenset({123}))))

    def test_whoami_message(self):
        self.assertEqual(whoami_message(123, "marco"), "Your Telegram user ID is 123 (@marco).")
        self.assertEqual(whoami_message(123), "Your Telegram user ID is 123.")

    def test_append_process_status_adds_pid(self):
        self.assertEqual(append_process_status("Status\nPending: 0", pid=1234), "Status\nPending: 0\nTelegram bot PID: 1234")

    def test_saved_photo_message(self):
        path = Path("data/inbound/telegram/2026-06-12/photo.jpg")
        self.assertEqual(saved_photo_message(path), "Saved photo")

    def test_queued_photo_message_reports_worker_status(self):
        path = Path("data/inbound/telegram/123/2026-06-12/photo.jpg")

        self.assertIn("Scanner: started", queued_photo_message(path, True))
        self.assertIn("Scanner: already running", queued_photo_message(path, False))

    def test_help_message_lists_safe_mobile_commands(self):
        text = telegram_help_message()

        self.assertIn("/status", text)
        self.assertIn("tap SD in preview → HD", text)
        self.assertIn("/restart", text)
        self.assertIn("/excel", text)
        self.assertIn("/checked", text)
        self.assertIn("/crops", text)
        self.assertIn("/change", text)
        self.assertIn("/del", text)
        self.assertIn("/group", text)
        self.assertIn("/rollback", text)
        self.assertIn("Available type/category values", text)
        self.assertIn("Food, Gas", text)
        self.assertIn("Office supplies", text)
        self.assertIn("/change 021 other", text)
        self.assertNotIn("/today_excel -", text)
        self.assertIn("/submit", text)

    def test_help_message_explains_hd_camera_in_chinese(self):
        self.assertIn("预览页点 SD → HD", telegram_help_message("zh"))

    def test_sd_four_receipts_warns_but_sd_three_and_hd_four_do_not(self):
        sd_four = QueueItem(path="sd-four.jpg", upload_quality="sd", image_width=960, image_height=1280, detected_receipt_count=4)
        sd_three = QueueItem(path="sd-three.jpg", upload_quality="sd", image_width=960, image_height=1280, detected_receipt_count=3)
        hd_four = QueueItem(path="hd-four.jpg", upload_quality="hd", image_width=1920, image_height=2560, detected_receipt_count=4)

        self.assertIn("SD→HD", _photo_quality_warning(sd_four, "zh"))
        self.assertEqual(_photo_quality_warning(sd_three, "zh"), "")
        self.assertEqual(_photo_quality_warning(hd_four, "zh"), "")

    def test_chinese_help_and_language_preference(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = Settings(root=Path(temp), output_dir=Path(temp) / "out", telegram_language="en")

            self.assertEqual(user_language(settings, 123), "en")
            self.assertEqual(set_user_language(settings, 123, "zh"), "zh")
            self.assertEqual(user_language(settings, 123), "zh")

        text = telegram_help_message("zh")
        self.assertIn("命令", text)
        self.assertIn("/lang", text)
        self.assertIn("切换语言", text)

    def test_telegram_command_menu_lists_popup_commands(self):
        commands = dict(telegram_command_menu())

        self.assertIn("change", commands)
        self.assertIn("del", commands)
        self.assertIn("group", commands)
        self.assertIn("rollback", commands)
        self.assertIn("checked", commands)
        self.assertNotIn("rerun", commands)
        self.assertIn("submit", commands)
        self.assertIn("edit crop", commands["change"])

    def test_telegram_instance_lock_blocks_second_start_and_cleans_stale_pid(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = Settings(root=Path(temp), output_dir=Path(temp) / "out")

            lock = _acquire_telegram_instance_lock(settings)
            try:
                with self.assertRaisesRegex(RuntimeError, "already running"):
                    _acquire_telegram_instance_lock(settings)
            finally:
                _release_telegram_instance_lock(lock)

            lock.write_text("99999999", encoding="utf-8")
            lock = _acquire_telegram_instance_lock(settings)
            try:
                self.assertTrue(lock.exists())
            finally:
                _release_telegram_instance_lock(lock)

    @unittest.skipIf(os.name != "nt", "Windows PID command-line check")
    def test_telegram_instance_lock_ignores_reused_non_telegram_pid(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = Settings(root=Path(temp), output_dir=Path(temp) / "out")
            settings.output_dir.mkdir(parents=True)
            process = subprocess.Popen([sys.executable, "-c", "import time; time.sleep(10)"])
            try:
                lock = settings.output_dir / "telegram_bot.pid"
                lock.write_text(str(process.pid), encoding="utf-8")

                acquired = _acquire_telegram_instance_lock(settings)
                try:
                    self.assertEqual(int(acquired.read_text(encoding="utf-8")), os.getpid())
                finally:
                    _release_telegram_instance_lock(acquired)
            finally:
                process.terminate()
                process.wait(timeout=5)

    def test_empty_change_message_shows_mobile_template(self):
        text = change_message(Settings(), 123, [])

        self.assertIn("/change 021 type Other", text)
        self.assertIn("/change 021 amount 33.35 currency USD", text)
        self.assertIn("/del 021", text)

    def test_empty_del_message_shows_mobile_template(self):
        text = delete_message(Settings(), 123, [])

        self.assertIn("/del 021", text)
        self.assertIn("/del 021 022", text)

    def test_group_message_arms_next_photo(self):
        with tempfile.TemporaryDirectory() as temp:
            settings = Settings(root=Path(temp), output_dir=Path(temp) / "out")
            output_dir = settings.output_dir / "telegram" / "123"
            photo = Path(temp) / "photo.jpg"

            text = group_message(settings, 123, [])

            self.assertIn("Group mode is ready", text)
            self.assertTrue(mark_source_for_group_if_armed(output_dir, photo))
            self.assertFalse(mark_source_for_group_if_armed(output_dir, photo))

    def test_group_message_previews_then_confirms_existing_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            from PIL import Image

            root = Path(temp)
            settings = Settings(root=root, output_dir=root / "out")
            output_dir = settings.output_dir / "telegram" / "123"
            workbook = telegram_user_workbook(settings, 123)
            crops = output_dir / "crops"
            crops.mkdir(parents=True)
            detail_crop = crops / "044_2026-05-09_MXN_566.00_SUSHI_ROLL.jpg"
            card_crop = crops / "045_2026-05-09_MXN_622.60_SUSHI_ROLL_CARD.jpg"
            Image.new("RGB", (120, 240), "white").save(detail_crop)
            Image.new("RGB", (100, 220), "white").save(card_crop)
            ReimbursementWorkbook(workbook).write_records(
                [
                    InvoiceRecord(line_no=44, invoice_date="2026-05-09", expense_category="Food", currency="MXN", total_amount=566, seller="SUSHI ROLL", crop_image=str(detail_crop)),
                    InvoiceRecord(line_no=45, invoice_date="2026-05-09", expense_category="Food", currency="MXN", total_amount=622.60, tips=56.60, seller="SUSHI ROLL MIFEL", contents="venta con propina", crop_image=str(card_crop)),
                ]
            )

            preview = group_message(settings, 123, ["044", "045"])

            self.assertIn("Group preview", preview)
            self.assertIn("Keep Trace ID: 044", preview)
            wb = load_workbook(workbook, data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                self.assertNotEqual(ws.cell(3, 2).value, "delete")
            finally:
                wb.close()

            confirmed = group_message(settings, 123, ["confirm"])

            self.assertIn("Group saved", confirmed)
            self.assertIn("Mark delete: 045", confirmed)
            records = load_workbook(workbook, data_only=True)
            try:
                ws = records[INVOICE_EXP_SHEET]
                self.assertEqual(ws.cell(2, 2).value, "correct")
                self.assertEqual(ws.cell(3, 2).value, "delete")
            finally:
                records.close()

    def test_rollback_message_removes_last_pending_photo(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            photo = root / "data" / "inbound" / "telegram" / "123" / "2026-07-01" / "bad.jpg"
            photo.parent.mkdir(parents=True)
            photo.write_bytes(b"jpg")
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([QueueItem(path=str(photo), status=PENDING)]))

            text = rollback_message(settings, 123, "zh")

            self.assertIn("已撤销最近一张照片", text)
            self.assertNotIn("bad.jpg", text)
            self.assertFalse(photo.exists())
            self.assertEqual(load_queue_state(telegram_user_queue_path(settings, 123)).items, [])

    def test_processing_success_message(self):
        summary = PipelineSummary(
            source_images=1,
            crops=2,
            records_written=2,
            workbook_path=Path("data/output/production/报销明细_2026_xlsx.xlsx"),
        )
        self.assertEqual(
            processing_success_message(summary),
            "Saved photo and processed current Telegram batch. Sources: 1. Rows: 2. Excel: data\\output\\production\\报销明细_2026_xlsx.xlsx",
        )

    def test_telegram_batch_source_uses_day_folder(self):
        day = Path("data/inbound/telegram/2026-06-12")

        self.assertEqual(telegram_batch_source(day), day)

    def test_processing_failure_message_includes_saved_path_and_error(self):
        path = Path("data/inbound/telegram/2026-06-12/photo.jpg")
        self.assertEqual(
            processing_failure_message(path, RuntimeError("OCR unavailable")),
            "Saved photo\nProcessing failed: OCR unavailable",
        )

    def test_processing_failure_message_handles_empty_error_text(self):
        path = Path("photo.jpg")
        self.assertEqual(processing_failure_message(path, RuntimeError()), "Saved photo\nProcessing failed: RuntimeError")

    def test_scan_completion_message_reports_crop_lines_and_today_totals(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            records = [
                InvoiceRecord(
                    line_no=1,
                    invoice_date="2026-06-20",
                    expense_category="Food",
                    seller="Cafe",
                    total_amount=150,
                    crop_image=str(root / "crops" / "021_2026-06-20_MXN_150.00_Cafe.jpg"),
                )
            ]
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(records)
            photo = root / "photo.jpg"
            photo.write_bytes(b"jpg")
            item = QueueItem(
                path=str(photo),
                status=DONE,
                row_count=1,
                total_amount=150,
                category_totals={"Food": 150},
                updated_at=datetime.now().isoformat(timespec="seconds"),
            )
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([item]))
            output = settings.output_dir / "telegram" / "123"
            (output / "processing_state.json").write_text(
                json.dumps({"records": [{"source_image": str(photo), "crop_image": records[0].crop_image}]}),
                encoding="utf-8",
            )

            text = scan_completion_message(settings, 123, item, records)

            self.assertIn("Scan complete", text)
            self.assertNotIn("Photo:", text)
            self.assertNotIn("photo.jpg", text)
            self.assertIn("This scan", text)
            self.assertIn("Crops:\n- 021 2026-06-20 Food: MXN 150.00 | Cafe", text)
            self.assertIn("Original totals:\n- MXN: 150.00", text)
            self.assertIn("Today\nRows: 1\nMXN total: 150.00", text)
            self.assertNotIn("Since last submit", text)
            self.assertIn("- Food: 150.00", text)
            self.assertIn("Queue", text)
            self.assertIn("Review: /excel", text)

    def test_scan_completion_message_explains_automatic_payment_pairing(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            record = InvoiceRecord(
                line_no=37,
                invoice_date="2026-05-16",
                expense_category="Food",
                seller="Las Palmas",
                total_amount=616.40,
                tips=80.40,
                crop_image=str(root / "crops" / "037_2026-05-16_MXN_536.00_Las_Palmas.jpg"),
                supporting_crop_images=[str(root / "crops" / "163_2026-05-16_MXN_616.40_Las_Palmas.jpg")],
                remarks="Combined payment slip; tips calculated as 80.40; supporting crop 163 kept",
            )
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records([record])
            photo = root / "photo.jpg"
            photo.write_bytes(b"jpg")
            item = QueueItem(path=str(photo), status=DONE, row_count=1, total_amount=616.40)
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState([item]))

            text = scan_completion_message(settings, 123, item, [record])

            self.assertIn("Automatic pairing:", text)
            self.assertIn("163 merged into 037", text)
            self.assertIn("payment includes tips 80.40", text)
            self.assertIn("will not be reused", text)

    def test_change_message_edits_manual_workbook_by_crop_id(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop = review / "021_2026-06-20_MXN_150.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", seller="Cafe", total_amount=150, crop_image=str(crop))]
            )

            text = change_message(settings, 123, ["021", "correct", "2026-7-1", "type", "Other", "+", "wrong", "OCR"])

            self.assertIn("Change saved", text)
            self.assertIn("Crop: 021", text)
            self.assertIn("Status: correct", text)
            self.assertFalse((output / "报销_checked_2026.xlsx").exists())
            self.assertFalse((output / "final_crops").exists())
            wb = load_workbook(telegram_user_workbook(settings, 123), data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Date"]).value.date(), date(2026, 7, 1))
                self.assertEqual(ws.cell(2, headers["Accounting Category"]).value, "Other")
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "correct")
                self.assertIn("wrong OCR", ws.cell(2, headers["Detail"]).value)
            finally:
                wb.close()

    def test_report_message_uses_current_checked_crops_not_historical_queue_done(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop1 = review / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop2 = review / "002_2026-06-13_MXN_200.00_Pemex.jpg"
            crop1.write_bytes(b"jpg")
            crop2.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", seller="Cafe", total_amount=100, crop_image=str(crop1)),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-13", expense_category="Gas", seller="Pemex", total_amount=200, crop_image=str(crop2)),
                ]
            )
            items = [
                QueueItem(path=str(root / f"{idx}.jpg"), status=DONE, received_at="2026-06-21T11:00:00", row_count=1)
                for idx in range(48)
            ]
            save_queue_state(telegram_user_queue_path(settings, 123), QueueState(items))

            text = report_message(settings, 123)

            self.assertIn("Photos: 2", text)
            self.assertNotIn("Photos: 48", text)
            self.assertIn("Days: 2; Food average/day: 50.00 MXN", text)
            self.assertFalse((output / "报销_checked_2026.xlsx").exists())
            self.assertFalse((output / "final_crops").exists())

    def test_change_message_accepts_wide_date_formats(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop = review / "022_2026-06-20_MXN_150.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", seller="Cafe", total_amount=150, crop_image=str(crop))]
            )

            text = change_message(settings, 123, ["022", "date", "July", "1", "2026"])

            self.assertIn("Change saved", text)
            wb = load_workbook(telegram_user_workbook(settings, 123), data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Date"]).value.date(), date(2026, 7, 1))
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "ok")
            finally:
                wb.close()

    def test_change_message_accepts_category_shorthand(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop = review / "023_2026-06-20_MXN_150.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", seller="Cafe", total_amount=150, crop_image=str(crop))]
            )

            text = change_message(settings, 123, ["023", "other"])

            self.assertIn("Change saved", text)
            self.assertTrue(load_queue_state(telegram_user_queue_path(settings, 123)).rollback_blocked)
            wb = load_workbook(telegram_user_workbook(settings, 123), data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Accounting Category"]).value, "Other")
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "ok")
            finally:
                wb.close()

    def test_change_message_accepts_chinese_category_shorthand(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop = review / "024_2026-06-20_MXN_150.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", seller="Cafe", total_amount=150, crop_image=str(crop))]
            )

            text = change_message(settings, 123, ["024", "\u6c7d\u6cb9"])

            self.assertIn("Change saved", text)
            wb = load_workbook(telegram_user_workbook(settings, 123), data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Accounting Category"]).value, "Gas")
                self.assertEqual(ws.cell(2, headers["Type"]).value, "\u6c7d\u6cb9")
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "ok")
            finally:
                wb.close()

    def test_change_message_rejects_delete_and_del_deletes_crop(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop = review / "021_2026-06-20_MXN_150.00_Cafe.jpg"
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", seller="Cafe", total_amount=150, crop_image=str(crop))]
            )

            rejected = change_message(settings, 123, ["021", "delete"])
            deleted = delete_message(settings, 123, ["021"])

            self.assertIn("Use /del 021", rejected)
            self.assertIn("Deleted", deleted)
            self.assertFalse((output / "报销_checked_2026.xlsx").exists())
            self.assertFalse((output / "final_crops").exists())
            wb = load_workbook(telegram_user_workbook(settings, 123), data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "delete")
            finally:
                wb.close()

    def test_del_message_deletes_multiple_crops(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
                telegram_allowed_user_ids=frozenset({123}),
            )
            output = settings.output_dir / "telegram" / "123"
            review = output / "review_crops"
            review.mkdir(parents=True)
            crop_38 = review / "038_2026-06-20_MXN_150.00_Cafe.jpg"
            crop_39 = review / "039_2026-06-20_MXN_50.00_Gas.jpg"
            crop_38.write_bytes(b"jpg")
            crop_39.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [
                    InvoiceRecord(line_no=1, invoice_date="2026-06-20", expense_category="Food", seller="Cafe", total_amount=150, crop_image=str(crop_38)),
                    InvoiceRecord(line_no=2, invoice_date="2026-06-20", expense_category="Gas", seller="Gas", total_amount=50, crop_image=str(crop_39)),
                ]
            )

            text = delete_message(settings, 123, ["038", "039"])

            self.assertIn("Deleted", text)
            self.assertIn("Crops: 038, 039", text)
            self.assertIn("Count: 2", text)
            self.assertFalse((output / "报销_checked_2026.xlsx").exists())
            self.assertFalse((output / "final_crops").exists())
            wb = load_workbook(telegram_user_workbook(settings, 123), data_only=True)
            try:
                ws = wb[INVOICE_EXP_SHEET]
                headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
                self.assertEqual(ws.cell(2, headers["Manual status"]).value, "delete")
                self.assertEqual(ws.cell(3, headers["Manual status"]).value, "delete")
            finally:
                wb.close()

    def test_telegram_photo_filename_uses_timestamp_and_file_id(self):
        stamp = datetime(2026, 6, 12, 8, 9, 10)

        self.assertEqual(telegram_photo_filename(stamp, "AgACAgQAAxkBAAIB"), "080910_AgACAgQAAxkBAAIB.jpg")

    def test_telegram_photo_filename_is_windows_safe(self):
        stamp = datetime(2026, 6, 12, 8, 9, 10)

        self.assertEqual(telegram_photo_filename(stamp, 'file/id:with*bad?chars'), "080910_file_id_with_bad_chars.jpg")

    def test_telegram_photo_filename_falls_back_to_unique_id(self):
        stamp = datetime(2026, 6, 12, 8, 9, 10)

        self.assertEqual(telegram_photo_filename(stamp, "???", "unique-id"), "080910_unique-id.jpg")

    def test_telegram_image_document_filename_preserves_supported_extension(self):
        stamp = datetime(2026, 6, 12, 8, 9, 10)

        self.assertEqual(telegram_image_document_filename(stamp, "file/id", "receipt.PNG"), "080910_file_id.png")

    def test_telegram_image_document_filter_accepts_images_only(self):
        image_document = type("Document", (), {"mime_type": "image/jpeg", "file_name": "receipt.jpg"})()
        pdf_document = type("Document", (), {"mime_type": "application/pdf", "file_name": "receipt.pdf"})()

        self.assertTrue(is_supported_telegram_image_document(image_document))
        self.assertFalse(is_supported_telegram_image_document(pdf_document))

    def test_review_crop_paths_returns_review_images(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            user_dir = settings.output_dir / "telegram" / "123"
            crop = user_dir / "review_crops" / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop.parent.mkdir(parents=True)
            crop.write_bytes(b"jpg")

            paths = review_crop_paths(settings, 123)

            self.assertEqual(paths, [crop])

    def test_review_crop_paths_returns_latest_completed_photo_only(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            user_dir = settings.output_dir / "telegram" / "123"
            review = user_dir / "review_crops"
            review.mkdir(parents=True)
            old_crop = review / "001_2026-06-12_MXN_100.00_Old.jpg"
            new_crop = review / "002_2026-06-13_MXN_200.00_New.jpg"
            old_crop.write_bytes(b"old")
            new_crop.write_bytes(b"new")
            old_photo = root / "old.jpg"
            new_photo = root / "171139_AgACAgEAAxkBAAIBLWpAWKnEY0vG5Y5o1q32EVFoD6YmAAJTDGsbNc0AAUZLghjrS_Ao9AEAAwIAA3kAAzwE.jpg"
            old_photo.write_bytes(b"old")
            new_photo.write_bytes(b"new")
            save_queue_state(
                telegram_user_queue_path(settings, 123),
                QueueState(
                    [
                        QueueItem(path=str(old_photo), status=DONE, received_at="2026-06-21T10:00:00", row_count=1),
                        QueueItem(path=str(new_photo), status=DONE, received_at="2026-06-21T11:00:00", row_count=1),
                    ]
                ),
            )
            (user_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {"source_image": str(old_photo), "crop_image": str(old_crop)},
                            {"source_image": str(new_photo), "crop_image": str(new_crop)},
                        ]
                    }
                ),
                encoding="utf-8",
            )

            paths = review_crop_paths(settings, 123)

            self.assertEqual(paths, [new_crop])

    def test_review_crop_paths_falls_back_to_recent_rows_when_state_names_are_stale(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            user_dir = settings.output_dir / "telegram" / "123"
            review = user_dir / "review_crops"
            review.mkdir(parents=True)
            old_crop = review / "001_2026-06-12_MXN_100.00_Old.jpg"
            new_crop_1 = review / "002_2026-06-13_MXN_200.00_New.jpg"
            new_crop_2 = review / "003_2026-06-13_MXN_300.00_New.jpg"
            old_crop.write_bytes(b"old")
            new_crop_1.write_bytes(b"new1")
            new_crop_2.write_bytes(b"new2")
            photo = root / "new.jpg"
            photo.write_bytes(b"photo")
            save_queue_state(
                telegram_user_queue_path(settings, 123),
                QueueState([QueueItem(path=str(photo), status=DONE, received_at="2026-06-21T11:00:00", row_count=2)]),
            )
            (user_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {"source_image": str(photo), "crop_image": "999_missing.jpg"},
                            {"source_image": str(photo), "crop_image": "998_missing.jpg"},
                        ]
                    }
                ),
                encoding="utf-8",
            )

            paths = review_crop_paths(settings, 123)

            self.assertEqual(paths, [new_crop_1, new_crop_2])

    def test_review_crop_paths_uses_raw_crop_paths_when_review_names_are_missing(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            user_dir = settings.output_dir / "telegram" / "123"
            review = user_dir / "review_crops"
            raw = user_dir / "crops"
            review.mkdir(parents=True)
            raw.mkdir(parents=True)
            old_review = review / "001_2026-06-12_MXN_100.00_Old.jpg"
            raw_crop_1 = raw / "019_2026-06-13_MXN_200.00_New.jpg"
            raw_crop_2 = raw / "020_2026-06-13_MXN_300.00_New.jpg"
            old_review.write_bytes(b"old")
            raw_crop_1.write_bytes(b"raw1")
            raw_crop_2.write_bytes(b"raw2")
            photo = root / "new.jpg"
            photo.write_bytes(b"photo")
            save_queue_state(
                telegram_user_queue_path(settings, 123),
                QueueState([QueueItem(path=str(photo), status=DONE, received_at="2026-06-21T11:00:00", row_count=2)]),
            )
            (user_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {"source_image": str(photo), "crop_image": str(raw_crop_1)},
                            {"source_image": str(photo), "crop_image": str(raw_crop_2)},
                        ]
                    }
                ),
                encoding="utf-8",
            )

            paths = review_crop_paths(settings, 123)

            self.assertEqual(paths, [raw_crop_1, raw_crop_2])

    def test_recent_message_reports_latest_two_uploads_and_excel_rows(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            user_dir = settings.output_dir / "telegram" / "123"
            review = user_dir / "review_crops"
            review.mkdir(parents=True)
            crop_40 = review / "040_2026-06-12_MXN_126.00_Cafe.jpg"
            crop_41 = review / "041_2026-06-12_MXN_11.00_Store.jpg"
            crop_39 = review / "039_2026-06-11_MXN_24.00_Old.jpg"
            for crop in (crop_39, crop_40, crop_41):
                crop.write_bytes(b"jpg")
            old_photo = root / "old.jpg"
            new_photo = root / "new.jpg"
            old_photo.write_bytes(b"old")
            new_photo.write_bytes(b"new")
            save_queue_state(
                telegram_user_queue_path(settings, 123),
                QueueState(
                    [
                        QueueItem(path=str(old_photo), status=DONE, received_at="2026-06-21T10:00:00", row_count=1),
                        QueueItem(path=str(new_photo), status=DONE, received_at="2026-06-21T11:00:00", row_count=2),
                    ]
                ),
            )
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [
                    InvoiceRecord(line_no=39, invoice_date="2026-06-11", expense_category="Food", seller="Old", total_amount=24, crop_image=str(crop_39)),
                    InvoiceRecord(line_no=40, invoice_date="2026-06-12", expense_category="Food", seller="Cafe", total_amount=126, crop_image=str(crop_40)),
                    InvoiceRecord(line_no=41, invoice_date="2026-06-12", expense_category="Other", seller="Store", total_amount=11, crop_image=str(crop_41)),
                ]
            )
            wb = load_workbook(telegram_user_workbook(settings, 123))
            try:
                ws = wb[INVOICE_EXP_SHEET]
                ws.cell(4, 2).value = "delete"
                wb.save(telegram_user_workbook(settings, 123))
            finally:
                wb.close()
            (user_dir / "processing_state.json").write_text(
                json.dumps(
                    {
                        "records": [
                            {"source_image": str(old_photo), "crop_image": str(crop_39)},
                            {"source_image": str(new_photo), "crop_image": str(crop_40)},
                            {"source_image": str(new_photo), "crop_image": str(crop_41)},
                        ]
                    }
                ),
                encoding="utf-8",
            )

            text = recent_message(settings, 123)

            self.assertIn("Recent uploads", text)
            self.assertIn("Input 1: done | Excel rows 2 | crops 2", text)
            self.assertNotIn("Photo:", text)
            self.assertNotIn("171139_AgACAgEAAxkBAAIB", text)
            self.assertIn("040 -> Excel row 3, No. 040", text)
            self.assertIn("041 -> Excel row 4, No. 041", text)
            self.assertIn("| delete", text)
            self.assertIn("Input 2: done | Excel rows 1 | crops 1", text)
            self.assertIn("039 -> Excel row 2, No. 039", text)

    def test_submit_message_requires_confirmation(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            settings = Settings(
                root=root,
                inbound_dir=root / "data" / "inbound",
                trial_dir=root / "data" / "trial",
                output_dir=root / "data" / "output",
                baseline_dir=root / "data" / "baseline",
            )
            crop = settings.output_dir / "telegram" / "123" / "review_crops" / "001_2026-06-12_MXN_100.00_Cafe.jpg"
            crop.parent.mkdir(parents=True)
            crop.write_bytes(b"jpg")
            ReimbursementWorkbook(telegram_user_workbook(settings, 123)).write_records(
                [InvoiceRecord(line_no=1, invoice_date="2026-06-12", expense_category="Food", seller="Cafe", total_amount=100, crop_image=str(crop))]
            )

            preview = submit_message(settings, 123, [])

            self.assertIn("Submit preview", preview)
            self.assertIn("Reply confirm", preview)
            self.assertTrue(telegram_user_workbook(settings, 123).exists())
            self.assertTrue(submit_pending_confirmation(settings, 123))

            slash_confirm = submit_message(settings, 123, ["confirm"])

            self.assertIn("Use /submit first", slash_confirm)
            self.assertTrue(telegram_user_workbook(settings, 123).exists())

            handled, pending = submit_confirmation_message(settings, 123, "maybe")

            self.assertFalse(handled)
            self.assertIn("Submit pending", pending)

            handled, submitted = submit_confirmation_message(settings, 123, "confirm")

            self.assertTrue(handled)
            self.assertIn("Submit /", submitted)
            self.assertIn("Batch:", submitted)
            self.assertIn("Records: 1", submitted)

    def test_submit_confirmation_can_cancel(self):
        handled, text = submit_confirmation_message(Settings(), 123, "cancel")

        self.assertTrue(handled)
        self.assertIn("cancelled", text)


class TelegramDownloadRetryTests(unittest.IsolatedAsyncioTestCase):
    async def test_download_retries_transient_timeout_then_succeeds(self):
        timed_out = type("TimedOut", (Exception,), {})
        with tempfile.TemporaryDirectory() as temp:
            target = Path(temp) / "photo.jpg"
            telegram_file = type("TelegramFile", (), {})()

            async def save_file(custom_path):
                Path(custom_path).write_bytes(b"photo")

            telegram_file.download_to_drive = AsyncMock(side_effect=save_file)
            bot = type("Bot", (), {})()
            bot.get_file = AsyncMock(side_effect=[timed_out("network timeout"), telegram_file])

            with patch("invoice_system.telegram_bot.asyncio.sleep", new_callable=AsyncMock) as sleep:
                await download_telegram_media(bot, "file-id", target)

            self.assertEqual(bot.get_file.await_count, 2)
            sleep.assert_awaited_once()
            self.assertEqual(target.read_bytes(), b"photo")


if __name__ == "__main__":
    unittest.main()
